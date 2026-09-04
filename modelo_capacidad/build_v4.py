"""Genera Modelo_Capacidad_CUADRO_v4.xlsx a partir de la v3.

Cambio de v3 a v4: en los escenarios B, C y D los dos overlock ya no se
modelan como reemplazo de las maquinas de Union/Montaje de las lineas 1 y 2,
sino como maquinas DE APOYO que agregan una posicion nueva a cada una de esas
lineas. Esto suma capacidad de overlock en lugar de solo mejorarla un 8%.
"""

from copy import copy

import openpyxl

SRC = "Modelo_Capacidad_CUADRO_v3.xlsx"
DST = "Modelo_Capacidad_CUADRO_v4.xlsx"

# Fila del bloque "no existe / nueva" que se usa como plantilla de estilos.
TEMPLATE_ROW = 23
APOYO_ROWS = (25, 26)
GAIN = "(1+Supuestos!$B$13)"


def copy_row_style(ws, src_row, dst_row, max_col=15):
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        dst._style = copy(src._style)


def build_lineas(ws):
    ws.insert_rows(APOYO_ROWS[0], 2)

    for row, linea in zip(APOYO_ROWS, ("Línea 1", "Línea 2")):
        copy_row_style(ws, TEMPLATE_ROW, row)
        ws.cell(row, 1).value = linea
        ws.cell(row, 2).value = 6
        ws.cell(row, 3).value = "Overlock"
        ws.cell(row, 4).value = "Apoyo Unión / Montaje"
        ws.cell(row, 5).value = 190
        ws.cell(row, 6).value = "No existe"
        ws.cell(row, 7).value = 0
        ws.cell(row, 8).value = "No existe"
        ws.cell(row, 9).value = 0
        for estado_col, ops_col in ((10, 11), (12, 13), (14, 15)):
            ws.cell(row, estado_col).value = "NUEVA (apoyo)"
            ws.cell(row, ops_col).value = f"=$E{row}*{GAIN}"

    # El overlock de Union/Montaje de L1 se queda como esta: ya no se compra
    # una maquina para reemplazarlo.
    for estado_col, ops_col in ((10, 11), (12, 13), (14, 15)):
        ws.cell(5, estado_col).value = "Disponible"
        ws.cell(5, ops_col).value = 190
        ws.cell(5, ops_col)._style = copy(ws.cell(5, 7)._style)
        ws.cell(5, estado_col)._style = copy(ws.cell(5, 6)._style)

    # Linea 6 bajo dos filas: sus formulas relativas hay que reescribirlas.
    for row in range(27, 32):
        ws.cell(row, 15).value = f"=$E{row}*{GAIN}"

    # Bloque de totales (desplazado dos filas por la insercion).
    cedida = "(1-Supuestos!$B$14)"
    line_rows = {
        "Línea 1": [5, 6, 7, 8, 23, 25],
        "Línea 2": [9, 10, 11, 12, 24, 26],
        "Línea 3": [13, 14, 15, 16, 17],
        "Línea 4": [18, 19, 20, 21, 22],
        "Línea 6": [27, 28, 29, 30, 31],
    }
    # col letra por escenario -> (columna de la formula, columna de ops, cede L3)
    escenarios = [
        ("F", "G", True),
        ("H", "I", True),
        ("J", "K", True),
        ("L", "M", False),
        ("N", "O", False),
    ]
    for i, (linea, rows) in enumerate(line_rows.items()):
        out_row = 35 + i
        ws.cell(out_row, 1).value = linea
        for dest_col, ops_col, cede in escenarios:
            suma = "+".join(f"{ops_col}{r}" for r in rows)
            factor = cedida if (linea == "Línea 3" and cede) else ""
            ws[f"{dest_col}{out_row}"] = f"=({suma})" + (f"*{factor}" if factor else "")
    for dest_col, _, _ in escenarios:
        ws[f"{dest_col}40"] = f"=SUM({dest_col}35:{dest_col}39)"

    ws["A44"] = (
        "Los 2 overlock del escenario B son DE APOYO: no reemplazan al overlock de "
        "Unión / Montaje de L1 y L2, agregan una sexta posición a cada una de esas "
        "líneas (posición 6). Por eso la capacidad de la línea sube en una máquina "
        "completa y no solo en el 8% de ganancia electrónica."
    )
    ws["A44"]._style = copy(ws["A43"]._style)


def build_capacidad(ws):
    overlock = [5, 7, 9, 11, 13, 15, 18, 20, 25, 26, 27, 29]
    cuello = [6, 10, 14, 19, 28]
    ruedo = [8, 12, 16, 17, 21, 22, 23, 24, 30, 31]
    l3_rows = set(range(13, 18))
    cedida = "*(1-Supuestos!$B$14)"

    # (columna destino en Capacidad, columna de ops en Lineas, aplica cesion L3)
    escenarios = [("B", "G", True), ("C", "I", True), ("D", "K", True),
                  ("E", "M", False), ("F", "O", False)]
    for dest_col, ops_col, cede in escenarios:
        for out_row, rows in ((5, overlock), (6, cuello), (7, ruedo)):
            partes = []
            for r in rows:
                ref = f"'Lineas por Escenario'!{ops_col}{r}"
                if cede and r in l3_rows:
                    ref += cedida
                partes.append(ref)
            ws[f"{dest_col}{out_row}"] = "=" + "+".join(partes)


def build_supuestos(ws):
    ws["A1"] = "Modelo de Capacidad — Taller CÚADRO (v4)"
    ws["A2"] = (
        "Base de medición: Operaciones_Por_Linea (2). Celdas AMARILLAS editables. "
        "v4: los 2 overlock de los escenarios B, C y D son DE APOYO (posición nueva "
        "en L1 y en L2), no reemplazo del overlock de Unión / Montaje."
    )
    ws["D10"] = (
        "SUPUESTO. Se renuevan 7 posiciones averiadas o inactivas (6 collareteras de "
        "ruedo + el overlock de L2). Los 2 overlock de apoyo son posiciones nuevas, "
        "no renovación, por eso la tasa no baja más."
    )
    ws["D11"] = "SUPUESTO. Igual que B (la collaretera de L5 no toca las líneas 1-4)"
    ws["D12"] = (
        "SUPUESTO. 7 posiciones renovadas, 2 overlock de apoyo y una línea 100% nueva"
    )

    ws["A29"] = "Cambio de la versión 3 a la versión 4"
    ws["A29"]._style = copy(ws["A25"]._style)
    notas = [
        "v3: el escenario B compraba 2 overlock para reemplazar el de Unión / Montaje "
        "de L1 y el de L2. Reemplazar una máquina sana solo aporta el 8% de ganancia "
        "electrónica, así que el techo del taller casi no se movía.",
        "v4: esos 2 overlock son DE APOYO. Se instalan como posición 6 de L1 y de L2 y "
        "suman una máquina completa de capacidad a cada línea. El overlock averiado de "
        "L2 se sigue reponiendo, porque eso ya venía en el escenario A.",
        "Overlock por lo tanto pasa de 2 a 3 máquinas en B (1 reposición + 2 de apoyo) "
        "y la inversión de B sube de US$ 20.800 a US$ 22.200.",
        "Como el overlock es la restricción activa del taller en todos los escenarios, "
        "este cambio es el que más mueve la capacidad de todo el modelo.",
    ]
    for i, texto in enumerate(notas):
        cell = ws.cell(30 + i, 1)
        cell.value = texto
        cell._style = copy(ws["A2"]._style)


def build_escenarios(ws):
    ws["B6"] = (
        "6 collareteras de ruedo + 1 overlock de reposición (L2) + 2 overlock DE APOYO "
        "(1 en L1 y 1 en L2)"
    )
    ws["C6"] = (
        "Además de reponer lo averiado, iguala las cuatro líneas y refuerza el cuello "
        "de botella. Hoy L3 y L4 tienen dos posiciones de ruedo y L1 y L2 solo una: se "
        "reemplazan las cuatro collareteras de L3 y L4 y se agrega una quinta posición "
        "de ruedo a L1 y otra a L2. Los 2 overlock NO son reemplazo: entran como "
        "máquinas de apoyo en una sexta posición de L1 y de L2, de modo que cada una de "
        "esas líneas gana una máquina completa de Unión / Montaje. El overlock averiado "
        "de L2 se repone igual que en A. Overlock es la restricción activa del taller, "
        "así que estas dos máquinas de apoyo son las que realmente levantan el techo."
    )
    ws["C8"] = (
        "Suma una quinta línea de confección con la estructura de 5 máquinas de las "
        "líneas 3 y 4: 2 overlock + 1 collaret de cuello + 2 collaret de ruedo. Es el "
        "único escenario que amplía la planta en lugar de repararla o reequilibrarla, y "
        "el único que alcanza a cubrir el promedio mensual de la zafra. Requiere "
        "dotación de personal para la línea nueva, costo no incluido en la cifra de "
        "inversión."
    )
    ws["D6"] = 9
    ws["D7"] = 10
    ws["D8"] = 15
    ws["A10"] = (
        "Los escenarios son acumulativos: A ⊂ B ⊂ C ⊂ D. Comprar A no invalida pasar "
        "después a B, C o D. B incluye la reposición del overlock de L2 que ya venía en A."
    )


def build_inversion(ws):
    # ruedo, cuello, overlock
    compras = {5: (3, 0, 1), 6: (6, 0, 3), 7: (7, 0, 3), 8: (9, 1, 5)}
    for row, (ruedo, cuello, over) in compras.items():
        ws.cell(row, 2).value = ruedo
        ws.cell(row, 3).value = cuello
        ws.cell(row, 4).value = over
    ws["A16"] = (
        "El escenario D no incluye el costo de la dotación de personal de la línea "
        "nueva. B, C y D incluyen los 2 overlock de apoyo además del overlock de "
        "reposición de L2."
    )


def build_cambio(wb):
    ws = wb.create_sheet("Cambio v3 a v4")
    src = wb["Capacidad"]
    title_style = copy(src["A1"]._style)
    sub_style = copy(src["A2"]._style)
    head_style = copy(src["A4"]._style)
    label_style = copy(src["A5"]._style)
    num_style = copy(src["B12"]._style)
    tot_style = copy(src["B15"]._style)

    for col, width in (("A", 46), ("B", 16), ("C", 16), ("D", 16), ("E", 60)):
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 16.5
    ws.row_dimensions[4].height = 25.5

    ws["A1"] = "Qué cambió de la v3 a la v4"
    ws["A1"]._style = title_style
    ws["A2"] = (
        "Los 2 overlock de los escenarios B, C y D pasan de reemplazo en L1 y L2 a "
        "máquinas de apoyo (posición nueva en cada línea). Capacidad comparada = "
        "Modelo 2, balance de estaciones."
    )
    ws["A2"]._style = sub_style

    headers = ["Concepto", "v3 — reemplazo", "v4 — de apoyo", "Diferencia", "Por qué"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(4, i)
        c.value = h
        c._style = head_style

    # Valores de la v3 congelados como referencia historica.
    v3 = {
        "B": dict(over=1543.6, dia=220.514285714286, mes=4233.87428571429,
                  deficit=-325.625714285714, inv=20800, maq=8),
        "C": dict(over=1582.4, dia=226.057142857143, mes=4340.29714285714,
                  deficit=-219.202857142856, inv=23800, maq=9),
        "D": dict(over=2001.44, dia=285.92, mes=5518.256,
                  deficit=958.756000000000, inv=34500, maq=14),
    }
    # concepto, clave v3, celda viva v4, formato
    filas = [
        ("Overlock — capacidad neta (ops/día)", "over", "=Capacidad!{cap}5", "#,##0.0",
         "Las 2 máquinas de apoyo suman posición nueva en vez de mejorar una existente"),
        ("Piezas/día (restricción activa)", "dia", "=Capacidad!{cap}23", "#,##0.0",
         "Overlock es la restricción activa: al ampliarla sube la producción"),
        ("Piezas/mes netas de reproceso", "mes", "=Capacidad!{cap}26", "#,##0",
         "20 días laborables, netas de reproceso"),
        ("Déficit / superávit — mes regular", "deficit",
         "='Demanda y Deficit'!{dem}5", '#,##0;[Red]\\(#,##0\\)',
         "Contra 4.559 pzas/mes con stock de seguridad"),
        ("Máquinas a comprar", "maq", "=Escenarios!D{esc}", "#,##0",
         "Un overlock más: 3 en total (1 de reposición en L2 + 2 de apoyo) en vez de 2"),
        ("Inversión US$", "inv", "=Inversion!H{inv}", "$#,##0",
         "El overlock adicional cuesta US$ 1.400"),
    ]
    cap_col = {"B": "D", "C": "E", "D": "F"}
    dem_col = {"B": "E", "C": "F", "D": "G"}
    esc_row = {"B": 6, "C": 7, "D": 8}

    row = 5
    for esc in ("B", "C", "D"):
        ws.cell(row, 1).value = f"ESCENARIO {esc}"
        ws.cell(row, 1)._style = label_style
        for col in range(2, 6):
            ws.cell(row, col)._style = label_style
        row += 1
        for concepto, key, plantilla, fmt, porque in filas:
            ws.cell(row, 1).value = concepto
            ws.cell(row, 1)._style = label_style
            viva = plantilla.format(cap=cap_col[esc], dem=dem_col[esc],
                                    esc=esc_row[esc], inv=esc_row[esc])
            b = ws.cell(row, 2)
            b.value = v3[esc][key]
            b._style = num_style
            b.number_format = fmt
            c = ws.cell(row, 3)
            c.value = viva
            c._style = tot_style
            c.number_format = fmt
            d = ws.cell(row, 4)
            d.value = f"=C{row}-B{row}"
            d._style = num_style
            d.number_format = fmt
            e = ws.cell(row, 5)
            e.value = porque
            e._style = sub_style
            row += 1
        row += 1

    ws.cell(row, 1).value = (
        "El escenario A no cambia: sigue siendo reposición estricta de las 4 máquinas "
        "averiadas, US$ 10.400."
    )
    ws.cell(row, 1)._style = sub_style
    ws.cell(row + 1, 1).value = (
        "La estructura de L1 y L2 en B, C y D queda con 6 posiciones: 2 overlock de "
        "línea + 1 overlock de apoyo + 1 collaret de cuello + 2 collaret de ruedo."
    )
    ws.cell(row + 1, 1)._style = sub_style


def main():
    wb = openpyxl.load_workbook(SRC)
    build_lineas(wb["Lineas por Escenario"])
    build_capacidad(wb["Capacidad"])
    build_supuestos(wb["Supuestos"])
    build_escenarios(wb["Escenarios"])
    build_inversion(wb["Inversion"])
    build_cambio(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.save(DST)
    print(f"escrito {DST}")


if __name__ == "__main__":
    main()
