#!/usr/bin/env python3
"""Genera SOLICITUD DE CONSUMIBLES Agosto 2026 en formato estándar."""
import re
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CUADRO = '/home/ubuntu/.cursor/projects/workspace/uploads/Consumibles_2Meses_Somos_Cuadro_d069.xlsx'
TEMPLATE = '/home/ubuntu/.cursor/projects/workspace/uploads/SOLICITUD_DE_CONSUMIBLES___SUM_AGOSTO_1_2561.xlsx'
OUTPUT = '/workspace/SOLICITUD_DE_CONSUMIBLES_AGOSTO_2026.xlsx'

ALIASES = {
    'BOLSAS BLANCAS PAPELERA': 'BOLSA DE PAPELERA',
    'CAFE': 'CAFÉ',
    'LAPICES': 'LAPIZ',
    'PILAS AA': 'BATERIAS AA',
    'PILAS AAA': 'BATERIAS AAA',
    'CELOVEN TRANSPARENTE CINTA ADHESIVA': 'CINTA ADHESIVA TRANSPARENTE',
    'CARTUCHO DE TONER 05A/80A': 'TONER 05A/80A',
    'POST IT': 'POST ITS',
    'CUTTERS EXACTOS': 'CUTTERS (EXACTO)',
    'MARCADOR ROJO PUNTA GRUESA': 'MARCADOR PERMANENTE PUNTA GRUESA',
}


def parse_num(val):
    if pd.isna(val) or val in ['-', '—', '']:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip().replace(',', ''))
    except ValueError:
        return 0


def clean_art(x):
    if pd.isna(x):
        return ''
    return re.sub(r'\s+', ' ', str(x).strip().upper())


def get_col(df, *patterns):
    for p in patterns:
        for c in df.columns:
            if p.lower() in str(c).lower():
                return c
    return None


def load_template_ref():
    template_items = pd.read_excel(TEMPLATE, sheet_name='25052026', header=5)
    ref = {}
    for _, row in template_items.iterrows():
        if pd.notna(row.get('ARTICULO')) and pd.notna(row.get('CANTIDAD')):
            art = clean_art(row['ARTICULO'])
            cant = str(row['CANTIDAD']).strip()
            obs = str(row.get('OBSERVACION', '')) if pd.notna(row.get('OBSERVACION')) else ''
            m_u = re.match(r'^([\d.,]+)\s*(.+)$', cant, re.I)
            if m_u:
                ref[art] = {
                    'qty': float(m_u.group(1).replace(',', '.')),
                    'unit': m_u.group(2).strip().upper(),
                    'obs': obs,
                }
    return ref


def format_qty(qty, articulo, template_ref):
    qty = float(qty)
    if qty <= 0:
        return None
    art = clean_art(articulo)

    for t_art, t_data in template_ref.items():
        mapped = ALIASES.get(t_art, t_art)
        if mapped == art or art in mapped or mapped in art:
            unit = t_data['unit']
            if unit in ('BULTOS', 'BULTO'):
                n = max(1, int(np.ceil(qty / 7)))
                return f'{n} BULTOS'
            if unit == 'GAL':
                return f'{max(1, int(np.ceil(qty)))} GAL'
            if unit in ('CAJAS', 'CAJA'):
                n = max(1, int(np.ceil(qty / 6)))
                return f'{n} {"CAJAS" if n > 1 else "CAJA"}'
            if unit == 'PAQ':
                n = max(1, int(np.ceil(qty / 10)))
                return f'{n} PAQ'
            return f'{max(1, int(np.ceil(qty)))} UND'

    rules = [
        (lambda a: 'CAFÉ' in a or a == 'CAFE', lambda q: (max(1, int(np.ceil(q / 7))), 'BULTOS')),
        (lambda a: a in {'DESINFECTANTE', 'CLORO', 'ALCOHOL', 'GEL DE BAÑO', 'LAVATODO', 'VENSOL', 'PRIDE'},
         lambda q: (max(1, int(np.ceil(q))), 'GAL')),
        (lambda a: 'SERVILLETAS' in a, lambda q: (max(1, int(np.ceil(q / 10))), 'PAQ')),
        (lambda a: 'RESMA' in a, lambda q: (max(1, int(np.ceil(q / 5))), 'PAQ')),
        (lambda a: 'PAPEL SANITARIO' in a, lambda q: (max(1, int(np.ceil(q / 5))), 'PAQ')),
        (lambda a: 'ROLLO' in a and 'FISCAL' in a, lambda q: (max(1, int(np.ceil(q / 5))), 'PAQ')),
        (lambda a: a in {'BOLIGRAFOS', 'LAPIZ', 'GRAPAS', 'CINTA DE EMBALAJE', 'CINTA TERMICA'},
         lambda q: (max(1, int(np.ceil(q / 6))), 'CAJAS')),
        (lambda a: 'BATERIAS' in a, lambda q: (max(1, int(np.ceil(q / 4))), 'CAJA')),
        (lambda a: 'TONER' in a, lambda q: (max(1, int(np.ceil(q))), 'UND')),
    ]
    for pred, conv in rules:
        if pred(art):
            n, unit = conv(qty)
            if unit == 'CAJAS' and n == 1:
                return '1 CAJA'
            return f'{n} {unit}'
    return f'{max(1, int(np.ceil(qty)))} UND'


def build_obs(row):
    parts = []
    abc = row.get('ABC', '')
    pct_nd = row.get('PCT_ND', 0)
    pct_aten = row.get('PCT_ATEN', 100)
    prom_mes = row.get('PROM_MES', 0)
    qty_nd = row.get('QTY_ND', 0)
    pend = row.get('QTY_PEND', 0)
    ped = row.get('PEDIR', 0)
    taller = row.get('PED_TALLER', 0)
    tiendas = row.get('PED_TIENDAS', 0)

    if abc == 'A':
        parts.append('PRIORIDAD A — ALTA ROTACIÓN')
    elif abc == 'B':
        parts.append('PRIORIDAD B — ROTACIÓN MEDIA')

    if qty_nd > 0:
        parts.append(f'INCLUYE DEMANDA NO ATENDIDA ({pct_nd:.0f}% ND · {int(qty_nd)} UND SIN STOCK)')
    if pend > 0:
        parts.append(f'CUBRE {int(pend)} UND PENDIENTES')
    if taller > 0 and tiendas > 0:
        parts.append(f'DEMANDA TALLER {int(taller)} + TIENDAS {int(tiendas)} UND/MES')
    elif tiendas > 0:
        parts.append(f'DEMANDA TIENDAS {int(tiendas)} UND/MES')
    elif taller > 0:
        parts.append(f'DEMANDA TALLER {int(taller)} UND/MES')

    parts.append(f'JUSTIFICADO: PROM {prom_mes:.0f} UND/MES (JUN–JUL 2026) +10% = {int(ped)} UND')

    if pct_aten < 70:
        parts.append('🔴 URGENTE — BAJO NIVEL DE SERVICIO')
    elif pct_aten < 85:
        parts.append('🟠 CRÍTICO — PEDIDO PRIORITARIO')
    else:
        parts.append('REABASTECER SEGÚN CONSUMO HISTÓRICO')

    return '. '.join(parts)


def style_header_row(ws, row, cols, header_font, header_fill, border, wrap):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = wrap


def main():
    template_ref = load_template_ref()
    mensual = pd.read_excel(CUADRO, sheet_name='📋 MENSUAL', header=8)
    unif = pd.read_excel(CUADRO, sheet_name='🔗 UNIFICADO', header=8)
    semanal = pd.read_excel(CUADRO, sheet_name='📋 SEMANAL', header=8)
    quincenal = pd.read_excel(CUADRO, sheet_name='📋 QUINCENAL', header=8)

    u = unif.copy()
    u.columns = [str(c).strip() for c in u.columns]
    art_col = get_col(u, 'ARTÍCULO', 'ARTICULO')
    u = u[u[art_col].notna() & ~u[art_col].astype(str).str.contains('TOTALES', na=False)].copy()
    u['ARTICULO'] = u[art_col].apply(clean_art)

    m = mensual.copy()
    m.columns = [str(c).strip() for c in m.columns]
    m_art = get_col(m, 'ARTÍCULO', 'ARTICULO')
    m = m[m[m_art].notna() & ~m[m_art].astype(str).str.contains('TOTALES', na=False)].copy()
    m['ARTICULO'] = m[m_art].apply(clean_art)

    ped_total_col = get_col(m, 'PEDIR\nTOTAL', 'PEDIR TOTAL')
    ped_taller_col = get_col(m, 'PEDIR\nPED_MENSUAL')
    ped_tiendas_col = get_col(m, 'PEDIR\nPED_MENSUAL.1')
    abc_col_m = get_col(m, 'ABC')

    records = []
    for _, mr in m.iterrows():
        art = mr['ARTICULO']
        ped_total = parse_num(mr[ped_total_col])
        if ped_total <= 0:
            continue

        ur = u[u['ARTICULO'] == art]
        if len(ur):
            ur = ur.iloc[0]
            cat_col = get_col(u, 'CATEGORÍA', 'CATEGORIA')
            abc_col = get_col(u, 'ABC')
            pct_aten_col = get_col(u, '% LÍNEAS\nATENDIDAS')
            pct_nd_col = get_col(u, '% LÍNEAS\nNO DISP')
            prom_col = get_col(u, 'PROM MES\nDEMANDA')
            qty_nd_col = get_col(u, 'QTY\nNO DISP')
            qty_pend_col = get_col(u, 'QTY\nPENDIENTE')
            row_data = {
                'ABC': ur[abc_col] if abc_col else mr.get(abc_col_m, ''),
                'CATEGORIA': ur[cat_col] if cat_col else '',
                'PCT_ATEN': parse_num(ur[pct_aten_col]) if pct_aten_col else 100,
                'PCT_ND': parse_num(ur[pct_nd_col]) if pct_nd_col else 0,
                'PROM_MES': parse_num(ur[prom_col]) if prom_col else ped_total / 1.1,
                'QTY_ND': parse_num(ur[qty_nd_col]) if qty_nd_col else 0,
                'QTY_PEND': parse_num(ur[qty_pend_col]) if qty_pend_col else 0,
            }
        else:
            row_data = {
                'ABC': mr.get(abc_col_m, ''),
                'CATEGORIA': '',
                'PCT_ATEN': 100,
                'PCT_ND': 0,
                'PROM_MES': ped_total / 1.1,
                'QTY_ND': 0,
                'QTY_PEND': 0,
            }

        row_data.update({
            'PEDIR': ped_total,
            'PED_TALLER': parse_num(mr[ped_taller_col]),
            'PED_TIENDAS': parse_num(mr[ped_tiendas_col]),
        })

        cant = format_qty(ped_total, art, template_ref)
        records.append({
            'CANTIDAD': cant,
            'ARTICULO': art,
            'OBSERVACION': build_obs(row_data),
            '_PED': ped_total,
            '_ABC': row_data['ABC'],
            '_CAT': row_data['CATEGORIA'],
        })

    abc_order = {'A': 0, 'B': 1, 'C': 2}
    records.sort(key=lambda r: (abc_order.get(r['_ABC'], 3), -r['_PED']))

    header_font = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='D9E1F2')
    title_fill = PatternFill('solid', fgColor='1F4E79')
    title_font = Font(bold=True, size=14, color='FFFFFF')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical='top')

    wb = Workbook()

    # Hoja principal — formato solicitud
    ws = wb.active
    ws.title = '05082026'
    for r, val in [
        (2, 'DPTO: CADENA DE SUMINISTROS-LOGISTICA'),
        (3, 'SOLICITANTE:  SAMUEL GRISANTI / SUPERVISOR LOGÍSTICA & INVENTARIOS'),
        (4, 'FECHA PEDIDO: 05/08/2026'),
        (5, ' FECHA ENTREGA:  (según lead time proveedor)'),
    ]:
        ws.merge_cells(f'A{r}:F{r}')
        ws.cell(row=r, column=1, value=val).font = Font(bold=True)

    headers = ['CANTIDAD', 'ARTICULO', 'CODIGO (OPCIONAL)', 'IMAGEN MUESTRA/LINK', 'OBSERVACION', 'ENTREGADO']
    for i, h in enumerate(headers, 1):
        ws.cell(row=6, column=i, value=h)
    style_header_row(ws, 6, 6, header_font, header_fill, border, wrap)

    row = 7
    for rec in records:
        ws.cell(row=row, column=1, value=rec['CANTIDAD'])
        ws.cell(row=row, column=2, value=rec['ARTICULO'])
        ws.cell(row=row, column=5, value=rec['OBSERVACION'])
        for c in range(1, 7):
            ws.cell(row=row, column=c).alignment = wrap
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['E'].width = 72
    ws.freeze_panes = 'A7'

    # VS pedido anterior
    ws5 = wb.create_sheet('VS PEDIDO JUN 2026')
    ws5['A1'] = 'COMPARATIVO: PEDIDO JUNIO 2026 vs AGOSTO 2026 (JUSTIFICADO 2 MESES)'
    ws5['A1'].font = title_font
    ws5['A1'].fill = title_fill
    ws5.merge_cells('A1:G1')

    new_map = {r['ARTICULO']: r for r in records}
    comp_vs = []
    matched_new = set()

    for t_art, t_data in template_ref.items():
        mapped = ALIASES.get(t_art, t_art)
        match = new_map.get(mapped)
        if not match:
            for k, v in new_map.items():
                if mapped in k or k in mapped:
                    match = v
                    break
        if match:
            matched_new.add(match['ARTICULO'])
        comp_vs.append({
            'Artículo Jun 2026': t_art,
            'Cant. Jun': f"{int(t_data['qty'])} {t_data['unit']}",
            'Artículo Ago 2026': match['ARTICULO'] if match else '— SIN DEMANDA 2M —',
            'Cant. Ago': match['CANTIDAD'] if match else '—',
            'Und/mes Cuadro': int(match['_PED']) if match else 0,
        })

    for rec in records:
        if rec['ARTICULO'] not in matched_new:
            comp_vs.append({
                'Artículo Jun 2026': '— NUEVO EN DEMANDA —',
                'Cant. Jun': '—',
                'Artículo Ago 2026': rec['ARTICULO'],
                'Cant. Ago': rec['CANTIDAD'],
                'Und/mes Cuadro': int(rec['_PED']),
            })

    headers5 = list(comp_vs[0].keys())
    for c, h in enumerate(headers5, 1):
        ws5.cell(row=3, column=c, value=h)
    style_header_row(ws5, 3, len(headers5), header_font, header_fill, border, wrap)
    for ri, rd in enumerate(comp_vs, 4):
        for ci, v in enumerate(rd.values(), 1):
            ws5.cell(row=ri, column=ci, value=v)
    ws5.freeze_panes = 'A4'

    # 3 frecuencias
    ws3 = wb.create_sheet('3 FRECUENCIAS')
    ws3['A1'] = 'SEMANAL · QUINCENAL · MENSUAL (SOMOS CUADRO — UNIFICADO)'
    ws3['A1'].font = title_font
    ws3['A1'].fill = title_fill
    ws3.merge_cells('A1:E1')

    freq_data = {}
    for df_raw, label in [(semanal, 'SEMANAL'), (quincenal, 'QUINCENAL'), (mensual, 'MENSUAL')]:
        df = df_raw.copy()
        df.columns = [str(c).strip() for c in df.columns]
        ac = get_col(df, 'ARTÍCULO', 'ARTICULO')
        pc = get_col(df, 'PEDIR\nTOTAL', 'PEDIR TOTAL')
        sub = df[df[ac].notna() & ~df[ac].astype(str).str.contains('TOTALES', na=False)]
        for _, r in sub.iterrows():
            art = clean_art(r[ac])
            p = parse_num(r[pc])
            if p > 0:
                freq_data.setdefault(art, {'Artículo': r[ac]})[label] = int(p)

    pivot = pd.DataFrame(freq_data.values()).fillna(0)
    for col in ['SEMANAL', 'QUINCENAL', 'MENSUAL']:
        if col not in pivot.columns:
            pivot[col] = 0
    pivot = pivot.sort_values('MENSUAL', ascending=False)

    for c, h in enumerate(pivot.columns, 1):
        ws3.cell(row=3, column=c, value=h)
    style_header_row(ws3, 3, len(pivot.columns), header_font, header_fill, border, wrap)
    for ri, rd in enumerate(pivot.itertuples(index=False), 4):
        for ci, v in enumerate(rd, 1):
            ws3.cell(row=ri, column=ci, value=v)

    tr = len(pivot) + 5
    ws3.cell(row=tr, column=1, value='TOTALES').font = Font(bold=True)
    for i, col in enumerate(['SEMANAL', 'QUINCENAL', 'MENSUAL'], 2):
        ws3.cell(row=tr, column=i, value=int(pivot[col].sum())).font = Font(bold=True)

    # Comparativo cuadro
    ws2 = wb.create_sheet('COMPARATIVO CUADRO')
    ws2['A1'] = 'TRAZABILIDAD: SOMOS CUADRO → SOLICITUD AGOSTO'
    ws2['A1'].font = title_font
    ws2['A1'].fill = title_fill
    ws2.merge_cells('A1:H1')

    comp_rows = []
    for rec in records:
        ur = u[u['ARTICULO'] == rec['ARTICULO']]
        comp_rows.append({
            'Artículo': rec['ARTICULO'],
            'ABC': rec['_ABC'],
            'Categoría': rec['_CAT'],
            'Pedido Mensual Cuadro (und)': int(rec['_PED']),
            'Cantidad Solicitud': rec['CANTIDAD'],
            'Demanda 2m (und)': int(parse_num(ur.iloc[0][get_col(u, 'QTY\nDEMANDA')])) if len(ur) else '',
            '% Atendidas': ur.iloc[0][get_col(u, '% LÍNEAS\nATENDIDAS')] if len(ur) else '',
            '% NO DISP': ur.iloc[0][get_col(u, '% LÍNEAS\nNO DISP')] if len(ur) else '',
        })
    comp_df = pd.DataFrame(comp_rows)
    for c, h in enumerate(comp_df.columns, 1):
        ws2.cell(row=3, column=c, value=h)
    style_header_row(ws2, 3, len(comp_df.columns), header_font, header_fill, border, wrap)
    for ri, rd in enumerate(comp_df.itertuples(index=False), 4):
        for ci, v in enumerate(rd, 1):
            ws2.cell(row=ri, column=ci, value=v)

    # Metodología
    ws4 = wb.create_sheet('METODOLOGIA')
    ws4['A1'] = 'METODOLOGÍA'
    ws4['A1'].font = title_font
    ws4['A1'].fill = title_fill
    notes = [
        ('Fuente', 'Consumibles_2Meses_Somos_Cuadro.xlsx + formato SOLICITUD SUM AGOSTO 1'),
        ('Período', 'Jun–Jul 2026 (2 meses historial desde cero)'),
        ('Cantidades', 'Hoja PEDIDO MENSUAL → PEDIR TOTAL (promedio + 10% buffer)'),
        ('Formato', 'Réplica estructura: CANTIDAD | ARTICULO | CODIGO | IMAGEN | OBSERVACION | ENTREGADO'),
        ('Unidades', 'UND / CAJAS / PAQ / GAL / BULTOS según convención pedido Jun 2026 y categoría'),
        ('Items', f'{len(records)} artículos · {sum(r["_PED"] for r in records):,.0f} und totales'),
        ('Servicio global', '76.1% líneas atendidas · 19.0% NO DISP · 91.0% qty cumplida'),
    ]
    for i, (k, v) in enumerate(notes, 3):
        ws4.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws4.cell(row=i, column=2, value=v)

    wb.save(OUTPUT)
    print(f'Generado: {OUTPUT}')
    print(f'Items: {len(records)} | Total und: {sum(r["_PED"] for r in records):,.0f}')


if __name__ == '__main__':
    main()
