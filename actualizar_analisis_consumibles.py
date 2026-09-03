#!/usr/bin/env python3
"""Análisis y solicitud de consumibles basados en consumo real (sin buffer)."""
import re
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Rutas ---
TIENDAS_FILE = '/home/ubuntu/.cursor/projects/workspace/uploads/_Solicitudes__TIENDAS__consumibles_2meses_99a4.xlsx'
TALLER_MOV_FILE = '/home/ubuntu/.cursor/projects/workspace/uploads/movimientos_taller_actualizado_0a2c.xlsx'
TEMPLATE = '/home/ubuntu/.cursor/projects/workspace/uploads/SOLICITUD_DE_CONSUMIBLES___SUM_AGOSTO_1_2561.xlsx'
ANALISIS_OUT = '/workspace/analisis_pedidos_consumibles_2meses.xlsx'
SOLICITUD_OUT = '/workspace/SOLICITUD_DE_CONSUMIBLES_AGOSTO_2026.xlsx'

MESES = 2.0
SEMANAS_MES = 4.33

ALIASES = {
    'BOLSAS BLANCAS PAPELERA': 'BOLSA DE PAPELERA',
    'CAFE': 'CAFÉ',
    'LAPICES': 'LAPIZ',
    'PILAS AA': 'BATERIAS AA',
    'PILAS AAA': 'BATERIAS AAA',
    'PINZA DEVASTADO': 'PINZA DE DEVASTADO',
    'PAÑOS AMARILLOS': 'PAÑO AMARILLO',
    'CELOVEN TRANSPARENTE CINTA ADHESIVA': 'CINTA ADHESIVA TRANSPARENTE',
    'CARTUCHO DE TONER 05A/80A': 'TONER 05A/80A',
    'POST IT': 'POST ITS',
    'CUTTERS EXACTOS': 'CUTTERS (EXACTO)',
    'MARCADOR ROJO PUNTA GRUESA': 'MARCADOR PERMANENTE PUNTA GRUESA',
}


def clean_art(x):
    if pd.isna(x):
        return ''
    s = re.sub(r'\s+', ' ', str(x).strip().upper())
    return ALIASES.get(s, s)


def round_qty(q):
    """Redondeo al entero más cercano; mínimo 1 si hubo consumo."""
    if q <= 0:
        return 0
    return max(1, int(round(q)))


def load_tiendas(path):
    xl = pd.ExcelFile(path)
    frames = []
    for sheet in xl.sheet_names:
        if sheet == 'INVENTARIO DE CONSUMIBLES':
            continue
        df = pd.read_excel(path, sheet_name=sheet, header=1)
        df.columns = ['FECHA', 'NOMBRE', 'ARTICULO', 'CARACTER_ADICIONAL', 'CANTIDAD', 'ESTADO', 'SUCURSAL', 'NOTAS']
        df = df[df['FECHA'].notna() & (df['FECHA'].astype(str) != 'FECHA')]
        df['ORIGEN'] = 'TIENDAS'
        frames.append(df)
    return pd.concat(frames, ignore_index=True)


def load_taller_movimientos(path):
    df = pd.read_excel(path, sheet_name='Hoja 1')
    df.columns = ['FECHA', 'CODIGO', 'MOVIMIENTO', 'ARTICULO', 'CATEGORIA', 'UND_MED', 'SUCURSAL']
    df['FECHA'] = pd.to_datetime(df['FECHA'], errors='coerce')
    df['CANTIDAD'] = df['MOVIMIENTO'].abs()
    df['ORIGEN'] = 'TALLER'
    df['ESTADO'] = 'CONSUMIDO'
    df['SUCURSAL'] = 'TALLER'
    df['NOMBRE'] = 'MOVIMIENTO INVENTARIO'
    df['NOTAS'] = df['CODIGO']
    df['CARACTER_ADICIONAL'] = df['UND_MED']
    return df


def prepare_data(df):
    out = df.copy()
    out['FECHA'] = pd.to_datetime(out['FECHA'], errors='coerce')
    out['CANTIDAD'] = pd.to_numeric(out['CANTIDAD'], errors='coerce').fillna(0)
    out['ESTADO_NORM'] = out['ESTADO'].astype(str).str.strip().str.upper()
    out['ARTICULO_NORM'] = out['ARTICULO'].apply(clean_art)
    out['MES'] = out['FECHA'].dt.to_period('M').astype(str)

    def clasificar(estado):
        if estado == 'NO DISPONIBLE':
            return 'NO ATENDIDA'
        if estado in {'RECIBIDO', 'ENTRGADO', 'ENTREGADO', 'ENVIADO', 'CONSUMIDO'}:
            return 'ATENDIDA'
        if estado == 'SOLICITADO':
            return 'PENDIENTE'
        return 'OTRO'

    out['CLASIFICACION'] = out['ESTADO_NORM'].apply(clasificar)
    out['ES_ATENDIDA'] = out['CLASIFICACION'].isin({'ATENDIDA', 'PENDIENTE'})
    out['ES_NO_DISPONIBLE'] = out['ESTADO_NORM'] == 'NO DISPONIBLE'
    return out


def calc_pedidos(und_total):
    """Promedios reales sin buffer."""
    if und_total <= 0:
        return 0, 0, 0, 0.0
    prom_mes = und_total / MESES
    prom_sem = prom_mes / SEMANAS_MES
    prom_quin = prom_mes / 2
    return (
        round_qty(prom_sem),
        round_qty(prom_quin),
        round_qty(prom_mes),
        round(prom_mes, 2),
    )


def agg_producto(df, origen_label=None):
    rows = []
    for key, sub in df.groupby('ARTICULO_NORM'):
        und_total = sub['CANTIDAD'].sum()
        movs = len(sub)
        und_no = sub.loc[sub['ES_NO_DISPONIBLE'], 'CANTIDAD'].sum()
        sol_no = sub['ES_NO_DISPONIBLE'].sum()
        und_med = sub['CARACTER_ADICIONAL'].mode().iloc[0] if len(sub['CARACTER_ADICIONAL'].mode()) else 'UND'
        if pd.isna(und_med):
            und_med = 'UND'

        ped_sem, ped_quin, ped_mes, prom_mes = calc_pedidos(und_total)

        rows.append({
            'Producto': key,
            'Und medida': und_med,
            'Registros': movs,
            'Consumo real (2 meses)': int(und_total),
            'Prom mensual real': prom_mes,
            'Pedido SEMANAL': ped_sem,
            'Pedido QUINCENAL': ped_quin,
            'Pedido MENSUAL': ped_mes,
            'Und NO DISPONIBLE': int(und_no),
            '% NO DISP (líneas)': round(100 * sol_no / movs, 1) if movs else 0,
        })
    res = pd.DataFrame(rows)
    if origen_label:
        res.insert(0, 'Origen', origen_label)
    return res.sort_values('Consumo real (2 meses)', ascending=False)


def build_unified(prod_tiendas, prod_taller):
    t = prod_tiendas.set_index('Producto')
    m = prod_taller.set_index('Producto')
    rows = []
    for p in sorted(set(t.index) | set(m.index)):
        pt = t.loc[p] if p in t.index else None
        pm = m.loc[p] if p in m.index else None

        und_t = int(pt['Consumo real (2 meses)']) if pt is not None else 0
        und_m = int(pm['Consumo real (2 meses)']) if pm is not None else 0
        und_tot = und_t + und_m

        ped_sem, ped_quin, ped_mes, prom_mes = calc_pedidos(und_tot)
        _, _, ped_mes_t, _ = calc_pedidos(und_t)
        _, _, ped_mes_m, _ = calc_pedidos(und_m)

        rows.append({
            'Producto': p,
            'Consumo Tiendas (2m)': und_t,
            'Consumo Taller (2m)': und_m,
            'Consumo Total (2m)': und_tot,
            'Prom mensual real': prom_mes,
            'Pedido SEMANAL': ped_sem,
            'Pedido QUINCENAL': ped_quin,
            'Pedido MENSUAL': ped_mes,
            'Pedido Mensual Tiendas': ped_mes_t,
            'Pedido Mensual Taller': ped_mes_m,
        })
    return pd.DataFrame(rows).sort_values('Consumo Total (2m)', ascending=False)


def resumen(df, nombre, fuente):
    n = len(df)
    und = df['CANTIDAD'].sum()
    no_disp = df['ES_NO_DISPONIBLE'].sum()
    _, _, ped_mes, _ = calc_pedidos(und)
    return pd.DataFrame([
        {'Indicador': 'Segmento', 'Valor': nombre},
        {'Indicador': 'Fuente', 'Valor': fuente},
        {'Indicador': 'Desde', 'Valor': df['FECHA'].min().strftime('%Y-%m-%d')},
        {'Indicador': 'Hasta', 'Valor': df['FECHA'].max().strftime('%Y-%m-%d')},
        {'Indicador': 'Registros', 'Valor': n},
        {'Indicador': 'Consumo real (und)', 'Valor': int(und)},
        {'Indicador': 'Prom mensual real (und)', 'Valor': round(und / MESES, 1)},
        {'Indicador': 'Líneas NO DISPONIBLE', 'Valor': int(no_disp)},
        {'Indicador': '% NO DISPONIBLE', 'Valor': f"{round(100 * no_disp / n, 1)}%" if n else '0%'},
        {'Indicador': 'Pedido mensual sugerido (sum)', 'Valor': int(agg_producto(df)['Pedido MENSUAL'].sum())},
    ])


# --- Excel helpers ---
def write_analysis_sheet(writer, name, df, intro=None):
    df.to_excel(writer, sheet_name=name, index=False, startrow=3 if intro else 0)
    ws = writer.sheets[name]
    fmt_title = writer.book.add_format({'bold': True, 'font_size': 14, 'bg_color': '#1F4E79', 'font_color': 'white'})
    fmt_header = writer.book.add_format({'bold': True, 'bg_color': '#D9E1F2', 'border': 1, 'text_wrap': True})
    fmt_note = writer.book.add_format({'italic': True, 'text_wrap': True, 'font_color': '#555555'})
    ws.write(0, 0, name, fmt_title)
    row = 1
    if intro:
        for line in intro:
            ws.write(row, 0, line, fmt_note)
            row += 1
    sr = 3 if intro else 0
    for col_num, col in enumerate(df.columns):
        ws.write(sr, col_num, col, fmt_header)
        ws.set_column(col_num, col_num, 40 if col == 'Producto' else 16)
    ws.freeze_panes(sr + 1, 0)


def load_template_ref():
    template_items = pd.read_excel(TEMPLATE, sheet_name='25052026', header=5)
    ref = {}
    for _, row in template_items.iterrows():
        if pd.notna(row.get('ARTICULO')) and pd.notna(row.get('CANTIDAD')):
            art = clean_art(row['ARTICULO'])
            cant = str(row['CANTIDAD']).strip()
            m_u = re.match(r'^([\d.,]+)\s*(.+)$', cant, re.I)
            if m_u:
                ref[art] = {'unit': m_u.group(2).strip().upper()}
    return ref


def format_qty(qty, articulo, template_ref):
    art = clean_art(articulo)
    qty = float(qty)
    if qty <= 0:
        return None

    for t_art, t_data in template_ref.items():
        mapped = ALIASES.get(t_art, t_art)
        if mapped == art or art in mapped or mapped in art:
            unit = t_data['unit']
            if unit in ('BULTOS', 'BULTO'):
                return f'{max(1, round(qty / 7))} BULTOS'
            if unit == 'GAL':
                return f'{max(1, round(qty))} GAL'
            if unit in ('CAJAS', 'CAJA'):
                n = max(1, round(qty / 6))
                return f'{n} {"CAJAS" if n > 1 else "CAJA"}'
            if unit == 'PAQ':
                return f'{max(1, round(qty / 10))} PAQ'
            return f'{max(1, round(qty))} UND'

    rules = [
        (lambda a: 'CAFÉ' in a, lambda q: f'{max(1, round(q / 7))} BULTOS'),
        (lambda a: a in {'DESINFECTANTE', 'CLORO', 'ALCOHOL', 'GEL DE BAÑO', 'LAVATODO', 'VENSOL'},
         lambda q: f'{max(1, round(q))} GAL'),
        (lambda a: 'SERVILLETAS' in a, lambda q: f'{max(1, round(q / 10))} PAQ'),
        (lambda a: 'RESMA' in a, lambda q: f'{max(1, round(q / 5))} PAQ'),
        (lambda a: a in {'BOLIGRAFOS', 'LAPIZ', 'GRAPAS', 'CINTA DE EMBALAJE', 'CINTA TERMICA'},
         lambda q: f'{max(1, round(q / 6))} {"CAJAS" if round(q/6) > 1 else "CAJA"}'),
        (lambda a: 'BATERIAS' in a, lambda q: f'{max(1, round(q / 4))} CAJA'),
    ]
    for pred, fmt in rules:
        if pred(art):
            return fmt(qty)
    return f'{max(1, round(qty))} UND'


def build_obs(row, freq_label, qty_col):
    """Observación factual, sin buffer ni alertas de ajuste."""
    und_t = int(row['Consumo Tiendas (2m)'])
    und_m = int(row['Consumo Taller (2m)'])
    und_tot = int(row['Consumo Total (2m)'])
    prom = row['Prom mensual real']
    qty = int(row[qty_col])

    partes = [f'CONSUMO REAL 2 MESES: {und_tot} UND']
    if und_t > 0 and und_m > 0:
        partes.append(f'TIENDAS {und_t} + TALLER {und_m}')
    elif und_t > 0:
        partes.append(f'TIENDAS {und_t}')
    elif und_m > 0:
        partes.append(f'TALLER {und_m}')

    partes.append(f'PROM MENSUAL REAL: {prom} UND')
    partes.append(f'PEDIDO {freq_label}: {qty} UND')
    return '. '.join(partes)


def write_solicitud_sheet(wb, sheet_name, unified, qty_col, freq_label, template_ref, header_font, border, wrap):
    ws = wb.create_sheet(sheet_name)

    for rn, val in [
        (2, 'DPTO: CADENA DE SUMINISTROS-LOGISTICA'),
        (3, 'SOLICITANTE:  SAMUEL GRISANTI / SUPERVISOR LOGÍSTICA & INVENTARIOS'),
        (4, 'FECHA PEDIDO: 06/08/2026'),
        (5, f' FRECUENCIA: {freq_label}'),
        (6, ' FECHA ENTREGA:  (según lead time proveedor)'),
    ]:
        ws.merge_cells(f'A{rn}:F{rn}')
        ws.cell(row=rn, column=1, value=val).font = Font(bold=True)

    headers = ['CANTIDAD', 'ARTICULO', 'CODIGO (OPCIONAL)', 'IMAGEN MUESTRA/LINK', 'OBSERVACION', 'ENTREGADO']
    hr = 8
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h).font = header_font

    records = []
    for _, r in unified.iterrows():
        qty = r[qty_col]
        if qty <= 0:
            continue
        cant = format_qty(qty, r['Producto'], template_ref)
        records.append({
            'CANTIDAD': cant,
            'ARTICULO': r['Producto'],
            'OBSERVACION': build_obs(r, freq_label, qty_col),
            '_qty': qty,
        })
    records.sort(key=lambda x: -x['_qty'])

    row = hr + 1
    for rec in records:
        ws.cell(row=row, column=1, value=rec['CANTIDAD'])
        ws.cell(row=row, column=2, value=rec['ARTICULO'])
        ws.cell(row=row, column=5, value=rec['OBSERVACION'])
        for c in range(1, 7):
            ws.cell(row=row, column=c).alignment = wrap
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['E'].width = 65
    ws.freeze_panes = f'A{hr + 1}'
    return records


def generate_solicitud(unified, template_ref):
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='D9E1F2')
    title_font = Font(bold=True, size=14, color='FFFFFF')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical='top')

    freqs = [
        ('PEDIDO SEMANAL', 'Pedido SEMANAL', 'SEMANAL'),
        ('PEDIDO QUINCENAL', 'Pedido QUINCENAL', 'QUINCENAL'),
        ('PEDIDO MENSUAL', 'Pedido MENSUAL', 'MENSUAL'),
    ]
    totals = {}
    for sheet_name, col, label in freqs:
        recs = write_solicitud_sheet(wb, sheet_name, unified, col, label, template_ref, header_font, border, wrap)
        totals[label] = {'items': len(recs), 'und': sum(r['_qty'] for r in recs)}

    # Referencia 3 frecuencias
    ws_ref = wb.create_sheet('3 FRECUENCIAS', 0)
    ws_ref['A1'] = 'REFERENCIA — CONSUMO REAL Y 3 OPCIONES DE PEDIDO (SIN BUFFER)'
    ws_ref['A1'].font = title_font
    ws_ref['A1'].fill = PatternFill('solid', fgColor='1F4E79')
    ws_ref.merge_cells('A1:J1')

    ref_cols = ['Producto', 'Consumo Tiendas (2m)', 'Consumo Taller (2m)', 'Consumo Total (2m)',
                'Prom mensual real', 'Pedido SEMANAL', 'Pedido QUINCENAL', 'Pedido MENSUAL']
    ref_df = unified[ref_cols]
    for c, h in enumerate(ref_cols, 1):
        ws_ref.cell(row=3, column=c, value=h).font = header_font
    for ri, rd in enumerate(ref_df.itertuples(index=False), 4):
        for ci, v in enumerate(rd, 1):
            ws_ref.cell(row=ri, column=ci, value=v)

    tr = len(ref_df) + 5
    ws_ref.cell(row=tr, column=1, value='TOTALES').font = Font(bold=True)
    for i, col in enumerate(['Consumo Total (2m)', 'Pedido SEMANAL', 'Pedido QUINCENAL', 'Pedido MENSUAL'], 4):
        ws_ref.cell(row=tr, column=i, value=int(ref_df[col].sum())).font = Font(bold=True)

    # Metodología
    ws_m = wb.create_sheet('METODOLOGIA')
    ws_m['A1'] = 'METODOLOGÍA'
    ws_m['A1'].font = title_font
    notes = [
        ('Principio', 'Cantidades basadas en consumo/movimientos reales. Sin buffer ni ajustes artificiales.'),
        ('Taller', 'movimientos_taller_actualizado.xlsx — salidas de inventario (consumo efectivo).'),
        ('Tiendas', 'Solicitudes 2 meses — cantidades registradas por sucursal.'),
        ('Período', 'Jun–Ago 2026 (~2 meses).'),
        ('Prom mensual', 'Consumo total 2 meses ÷ 2'),
        ('Pedido semanal', 'Prom mensual ÷ 4.33, redondeado'),
        ('Pedido quincenal', 'Prom mensual ÷ 2, redondeado'),
        ('Pedido mensual', 'Prom mensual, redondeado'),
        ('Totales pedido mensual', f"{totals['MENSUAL']['items']} items · {totals['MENSUAL']['und']:,} und"),
        ('Totales pedido quincenal', f"{totals['QUINCENAL']['items']} items · {totals['QUINCENAL']['und']:,} und"),
        ('Totales pedido semanal', f"{totals['SEMANAL']['items']} items · {totals['SEMANAL']['und']:,} und"),
    ]
    for i, (k, v) in enumerate(notes, 3):
        ws_m.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws_m.cell(row=i, column=2, value=v)

    wb.save(SOLICITUD_OUT)
    return totals


def main():
    tiendas = prepare_data(load_tiendas(TIENDAS_FILE))
    taller = prepare_data(load_taller_movimientos(TALLER_MOV_FILE))

    prod_tiendas = agg_producto(tiendas, 'TIENDAS')
    prod_taller = agg_producto(taller, 'TALLER')
    unified = build_unified(prod_tiendas, prod_taller)

    # --- Análisis Excel ---
    writer = pd.ExcelWriter(ANALISIS_OUT, engine='xlsxwriter')

    metodologia = pd.DataFrame({
        'Sección': ['Enfoque', 'Taller', 'Tiendas', 'Período', 'Fórmulas', 'Buffer'],
        'Descripción': [
            'Consumo real registrado en 2 meses. Sin proyecciones ni colchones.',
            'Movimientos de inventario (salidas = consumo).',
            'Solicitudes por sucursal (cantidades pedidas/recibidas).',
            f"{tiendas['FECHA'].min().date()} a {max(tiendas['FECHA'].max(), taller['FECHA'].max()).date()}",
            'Prom mensual = und/2 · Semanal = prom/4.33 · Quincenal = prom/2',
            'NO se aplica buffer (+10%). Solo realidad de los datos.',
        ],
    })
    write_analysis_sheet(writer, '0-METODOLOGIA', metodologia)

    exec_df = pd.concat([
        resumen(tiendas, 'TIENDAS', 'Solicitudes sucursales').assign(Segmento='TIENDAS'),
        resumen(taller, 'TALLER', 'Movimientos inventario').assign(Segmento='TALLER'),
        resumen(pd.concat([tiendas, taller]), 'UNIFICADO', 'Tiendas + Taller').assign(Segmento='UNIFICADO'),
    ], ignore_index=True)
    write_analysis_sheet(writer, '1-RESUMEN', exec_df)

    write_analysis_sheet(writer, '2-TIENDAS', prod_tiendas.drop(columns=['Origen'], errors='ignore'))
    write_analysis_sheet(writer, '3-TALLER', prod_taller.drop(columns=['Origen'], errors='ignore'))
    write_analysis_sheet(writer, '4-UNIFICADO', unified,
                         intro=['Consumo real 2 meses. Tres opciones de pedido sin buffer.'])

    # Detalle movimientos taller
    det = taller[['FECHA', 'ARTICULO_NORM', 'CANTIDAD', 'CATEGORIA', 'CARACTER_ADICIONAL', 'NOTAS']].copy()
    det.columns = ['FECHA', 'PRODUCTO', 'CANTIDAD', 'CATEGORIA', 'UND_MED', 'CODIGO']
    write_analysis_sheet(writer, '5-MOVIMIENTOS TALLER', det.sort_values('FECHA', ascending=False))

    det_t = tiendas[['FECHA', 'SUCURSAL', 'ARTICULO_NORM', 'CANTIDAD', 'ESTADO_NORM']].copy()
    det_t.columns = ['FECHA', 'SUCURSAL', 'PRODUCTO', 'CANTIDAD', 'ESTADO']
    write_analysis_sheet(writer, '6-SOLICITUDES TIENDAS', det_t.sort_values('FECHA', ascending=False))

    writer.close()

    # --- Solicitud ---
    template_ref = load_template_ref()
    totals = generate_solicitud(unified, template_ref)

    print('=' * 60)
    print('SOLICITUD ACTUALIZADA — CONSUMO REAL, SIN BUFFER')
    print('=' * 60)
    print(f"Consumo taller (movimientos): {int(taller['CANTIDAD'].sum())} und")
    print(f"Consumo tiendas (solicitudes): {int(tiendas['CANTIDAD'].sum())} und")
    print(f"Consumo total 2 meses: {int(unified['Consumo Total (2m)'].sum())} und")
    print()
    for label in ['SEMANAL', 'QUINCENAL', 'MENSUAL']:
        t = totals[label]
        print(f"  {label:12} → {t['items']} items · {t['und']:,} und")
    print(f"\nArchivos:\n  {ANALISIS_OUT}\n  {SOLICITUD_OUT}")


if __name__ == '__main__':
    main()
