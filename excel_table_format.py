"""Export DataFrames with the standard Cuadro table format."""
from __future__ import annotations

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MESES_ES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}

STANDARD_VENTAS_COLUMNS = [
    "Año",
    "Mes",
    "Producto",
    "Variante del producto",
    "SKU",
    "GENERO",
    "COLOR",
    "TALLA",
    "tienda / ubicación",
    "Cant. ordenada",
]

HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(size=11)
ALT_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def prepare_standard_ventas_df(df: pd.DataFrame) -> pd.DataFrame:
    """Map processed sales data to the standard 10-column layout."""
    out = df.copy()
    fechas = pd.to_datetime(out["Fecha de la orden"], errors="coerce")

    producto = out["Producto"] if "Producto" in out.columns else out.get("modelo")

    cant = out["Cant. ordenada"]
    if pd.api.types.is_float_dtype(cant):
        cant = cant.round().astype("Int64")

    prepared = pd.DataFrame(
        {
            "Año": fechas.dt.year,
            "Mes": fechas.dt.month.map(MESES_ES),
            "Producto": producto,
            "Variante del producto": out["Variante del producto"],
            "SKU": out["SKU"],
            "GENERO": out["GENERO"],
            "COLOR": out["COLOR"],
            "TALLA": out["TALLA"],
            "tienda / ubicación": out["tienda / ubicación"],
            "Cant. ordenada": cant,
        }
    )
    return prepared[STANDARD_VENTAS_COLUMNS]


def _apply_table_style(ws, n_rows: int, n_cols: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows}"

    widths = {
        "Año": 8,
        "Mes": 14,
        "Producto": 24,
        "Variante del producto": 46,
        "SKU": 16,
        "GENERO": 10,
        "COLOR": 28,
        "TALLA": 8,
        "tienda / ubicación": 18,
        "Cant. ordenada": 14,
    }

    for col_idx in range(1, n_cols + 1):
        header = ws.cell(row=1, column=col_idx).value
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(str(header), 14)

        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.fill = HEADER_FILL
        header_cell.font = HEADER_FONT
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.border = THIN_BORDER

    for row_idx in range(2, n_rows + 1):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            header = ws.cell(row=1, column=col_idx).value
            if header == "Cant. ordenada":
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 22


def export_formatted_ventas(
    df: pd.DataFrame,
    output_path: str,
    sheet_name: str = "Ventas",
) -> pd.DataFrame:
    """Write sales data using the standard formatted table layout."""
    from process_ventas_report import sanitize_dataframe_for_excel

    prepared = prepare_standard_ventas_df(df)
    prepared = sanitize_dataframe_for_excel(prepared)

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
        datetime_format="yyyy-mm-dd hh:mm:ss",
        date_format="yyyy-mm-dd",
    ) as writer:
        prepared.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        _apply_table_style(ws, len(prepared) + 1, len(prepared.columns))

    return prepared
