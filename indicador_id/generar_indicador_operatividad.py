#!/usr/bin/env python3
"""
Genera el libro Excel de Indicador de Operatividad I+D (quincenal).

Diseño:
- Hoja 1_Datos: pegar/actualizar operaciones (fuente viva).
- Hoja 2_Parametros: quincena activa, pesos de prioridad y metas.
- Hojas de KPI con fórmulas Excel (SUMIFS/COUNTIFS/AVERAGEIFS) para que
  al renovar datos se recalculen automáticamente al abrir en Excel/Sheets.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
OUT_REPO = ROOT / "Indicador_Operatividad_ID_Quincenal.xlsx"
OUT_ARTIFACT = Path("/opt/cursor/artifacts/Indicador_Operatividad_ID_Quincenal.xlsx")

# Columnas canónicas esperadas al pegar desde KPIS I+D.xlsx
HEADERS = [
    "ID",
    "Operacion",
    "Responsable",
    "Fecha_Inicio",
    "Fecha_Finalizacion",
    "Prioridad",
    "Estado",
    "Area",
    "Observaciones",
]

THIN = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_TITLE = PatternFill("solid", fgColor="0B3A5C")
FILL_SOFT = PatternFill("solid", fgColor="E8F1F8")
FILL_WARN = PatternFill("solid", fgColor="FFF4CE")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_BAD = PatternFill("solid", fgColor="FFC7CE")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")
FONT_WHITE = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_TITLE = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
FONT_H2 = Font(name="Calibri", bold=True, color="1F4E79", size=13)
FONT_BODY = Font(name="Calibri", size=11)
FONT_BOLD = Font(name="Calibri", bold=True, size=11)


def style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = FILL_HEADER
        cell.font = FONT_WHITE
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def autosize(ws, min_width: int = 12, max_width: int = 42) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, length + 2))


def demo_operations() -> list[list]:
    """Datos demostrativos con estructura típica I+D (reemplazar por datos reales)."""
    base = date(2026, 7, 1)
    responsables = [
        "Ana Pérez",
        "Luis Gómez",
        "María Ruiz",
        "Carlos Díaz",
        "Sofía Vargas",
    ]
    prioridades = ["Alta", "Media", "Baja"]
    areas = ["Diseño", "Innovación", "Prototipo", "Validación"]
    estados = ["Completada", "Completada", "Completada", "En proceso", "Pendiente"]
    rows = []
    for i in range(1, 61):
        resp = responsables[(i - 1) % len(responsables)]
        prio = prioridades[(i * 2) % len(prioridades)]
        area = areas[(i * 3) % len(areas)]
        # Distribuye inicios a lo largo de julio-agosto 2026
        start = base + timedelta(days=((i - 1) * 2) % 50)
        # Ciclo según prioridad: Alta más corta esperada
        planned = {"Alta": 5, "Media": 9, "Baja": 14}[prio]
        # Variación de cumplimiento: algunos se atrasan
        slip = 0 if i % 4 else (2 if i % 3 else 5)
        end = start + timedelta(days=planned + slip - 1)
        estado = estados[i % len(estados)]
        if estado != "Completada":
            end = ""  # sin cierre aún
        rows.append(
            [
                f"OP-{i:03d}",
                f"Operación I+D #{i}",
                resp,
                start,
                end if end else None,
                prio,
                estado,
                area,
                "Dato demo — reemplazar con KPIS I+D.xlsx",
            ]
        )
    return rows


def build_workbook() -> Workbook:
    wb = Workbook()

    # ------------------------------------------------------------------
    # 0_Instrucciones
    # ------------------------------------------------------------------
    ws0 = wb.active
    ws0.title = "0_Instrucciones"
    ws0["A1"] = "Indicador de Operatividad I+D — Medición Quincenal"
    ws0["A1"].font = FONT_TITLE
    ws0["A1"].fill = FILL_TITLE
    ws0.merge_cells("A1:F1")
    ws0.row_dimensions[1].height = 28

    instrucciones = [
        "",
        "OBJETIVO",
        "Medir la operatividad de cada responsable del departamento I+D (Innovación y Diseño) cada quincena,",
        "usando fechas de inicio/finalización y el nivel de prioridad de cada operación.",
        "",
        "CÓMO ACTUALIZAR (enfoque auto-renovable)",
        "1) Abra la hoja 1_Datos.",
        "2) Borre solo las filas de datos (no borre la fila 1 de encabezados).",
        "3) Pegue las operaciones actuales desde KPIS I+D.xlsx (mismas columnas o mapee a estas).",
        "4) En 2_Parametros elija Año, Mes y Quincena a evaluar (1 = días 1-15; 2 = días 16-fin).",
        "5) Las hojas 3_Calc, 4_KPI_Responsables y 5_Ranking se actualizan solas con fórmulas Excel.",
        "",
        "COLUMNAS REQUERIDAS EN 1_Datos",
        "ID | Operacion | Responsable | Fecha_Inicio | Fecha_Finalizacion | Prioridad | Estado | Area | Observaciones",
        "Prioridad aceptada: Alta / Media / Baja (también P1/P2/P3 o 1/2/3 se normalizan en 3_Calc).",
        "Estado: Completada / En proceso / Pendiente / Cancelada (ajustable).",
        "",
        "INDICADOR PRINCIPAL: IOI — Índice de Operatividad I+D",
        "IOI = (0.70 × Cumplimiento ponderado a tiempo) + (0.30 × Eficiencia de ciclo ponderada)",
        "  • Cumplimiento ponderado: operaciones cerradas a tiempo / operaciones a evaluar, ponderadas por prioridad.",
        "  • Eficiencia de ciclo: meta de días por prioridad / días reales, tope 100%, ponderada por prioridad.",
        "  • Pesos por defecto: Alta=3, Media=2, Baja=1 (editables en 2_Parametros).",
        "",
        "QUINCENA DE EVALUACIÓN",
        "Se incluyen operaciones cuya Fecha_Finalizacion cae en la quincena, o (si siguen abiertas)",
        "cuya Fecha_Inicio está en la quincena / siguen activas al cierre de la quincena.",
        "Regla operativa recomendada (editable): evaluar principalmente cierres de la quincena + atrasos abiertos.",
        "",
        "SEMAFORO",
        "Cumple ≥ Meta IOI (default 85%) | Riesgo entre Umbral y Meta | Critico < Umbral (default 70%).",
    ]
    for i, line in enumerate(instrucciones, start=2):
        ws0.cell(i, 1, line).font = FONT_H2 if line.isupper() and line else FONT_BODY
    ws0.column_dimensions["A"].width = 120

    # ------------------------------------------------------------------
    # 1_Datos
    # ------------------------------------------------------------------
    ws1 = wb.create_sheet("1_Datos")
    ws1["A1"] = "PEGUE AQUÍ LOS DATOS RENOVADOS DE KPIS I+D (no modifique los encabezados)"
    ws1["A1"].fill = FILL_WARN
    ws1["A1"].font = FONT_BOLD
    ws1.merge_cells("A1:I1")

    for col, h in enumerate(HEADERS, start=1):
        ws1.cell(2, col, h)
    style_header_row(ws1, 2, len(HEADERS))

    demo = demo_operations()
    for r_idx, row in enumerate(demo, start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws1.cell(r_idx, c_idx, val)
            cell.border = THIN
            cell.font = FONT_BODY
            if c_idx in (4, 5) and isinstance(val, date):
                cell.number_format = "DD/MM/YYYY"

    last_data_row = 2 + len(demo)
    table = Table(displayName="TablaDatos", ref=f"A2:I{last_data_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws1.add_table(table)

    # Validación prioridad / estado
    dv_prio = DataValidation(
        type="list", formula1='"Alta,Media,Baja,P1,P2,P3,1,2,3"', allow_blank=True
    )
    dv_estado = DataValidation(
        type="list",
        formula1='"Completada,En proceso,Pendiente,Cancelada"',
        allow_blank=True,
    )
    ws1.add_data_validation(dv_prio)
    ws1.add_data_validation(dv_estado)
    dv_prio.add(f"F3:F{max(last_data_row, 500)}")
    dv_estado.add(f"G3:G{max(last_data_row, 500)}")

    # Filas vacías preformateadas para pegar más datos sin romper formato
    for r in range(last_data_row + 1, last_data_row + 40):
        for c in range(1, 10):
            cell = ws1.cell(r, c, None)
            cell.border = THIN
            if c in (4, 5):
                cell.number_format = "DD/MM/YYYY"

    autosize(ws1)
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["I"].width = 40
    ws1.freeze_panes = "A3"

    # ------------------------------------------------------------------
    # 2_Parametros
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("2_Parametros")
    ws2["A1"] = "Parámetros de medición (editables)"
    ws2["A1"].font = FONT_TITLE
    ws2["A1"].fill = FILL_TITLE
    ws2.merge_cells("A1:D1")

    params = [
        ("A3", "Año evaluación", 2026),
        ("A4", "Mes evaluación (1-12)", 8),
        ("A5", "Quincena (1 o 2)", 1),
        ("A7", "Peso prioridad Alta / P1 / 1", 3),
        ("A8", "Peso prioridad Media / P2 / 2", 2),
        ("A9", "Peso prioridad Baja / P3 / 3", 1),
        ("A11", "Meta días ciclo Alta", 5),
        ("A12", "Meta días ciclo Media", 9),
        ("A13", "Meta días ciclo Baja", 14),
        ("A15", "Peso IOI — Cumplimiento a tiempo", 0.70),
        ("A16", "Peso IOI — Eficiencia de ciclo", 0.30),
        ("A18", "Meta IOI (%)", 0.85),
        ("A19", "Umbral alerta IOI (%)", 0.70),
        ("A21", "Incluir abiertas atrasadas (1=Sí, 0=No)", 1),
    ]
    for cell_ref, label, value in params:
        row = int(cell_ref[1:])
        ws2.cell(row, 1, label).font = FONT_BOLD
        ws2.cell(row, 1).fill = FILL_SOFT
        val_cell = ws2.cell(row, 2, value)
        val_cell.fill = FILL_INPUT
        val_cell.font = FONT_BOLD
        val_cell.border = THIN
        if "IOI" in label and "%" in label or "Peso IOI" in label:
            val_cell.number_format = "0%"
        if label.startswith("Meta IOI") or label.startswith("Umbral"):
            val_cell.number_format = "0%"
        if "Peso IOI" in label:
            val_cell.number_format = "0%"

    # Nombres amigables / fechas de quincena calculadas
    ws2["A23"] = "Inicio quincena"
    ws2["B23"] = '=DATE(B3,B4,IF(B5=1,1,16))'
    ws2["B23"].number_format = "DD/MM/YYYY"
    ws2["A24"] = "Fin quincena"
    ws2["B24"] = '=IF(B5=1,DATE(B3,B4,15),EOMONTH(DATE(B3,B4,1),0))'
    ws2["B24"].number_format = "DD/MM/YYYY"
    ws2["A25"] = "Etiqueta quincena"
    ws2["B25"] = '="Q"&B5&" "&TEXT(DATE(B3,B4,1),"MMM-YYYY")'
    ws2["A23"].font = ws2["A24"].font = ws2["A25"].font = FONT_BOLD

    ws2["D3"] = "Leyenda de actualización"
    ws2["D3"].font = FONT_H2
    ws2["D4"] = "Celdas amarillas = parámetros editables."
    ws2["D5"] = "Al cambiar Año/Mes/Quincena, todo el KPI se recalcula."
    ws2["D6"] = "Al pegar nuevos datos en 1_Datos, no hace falta tocar fórmulas."
    ws2["D7"] = "Si su archivo usa otros nombres de columna, mapee a los encabezados de 1_Datos."

    for col, w in {"A": 46, "B": 16, "C": 3, "D": 70}.items():
        ws2.column_dimensions[col].width = w

    # ------------------------------------------------------------------
    # 3_Calc — fila a fila con fórmulas referenciando TablaDatos / rangos
    # Usamos rangos A3:Ix para compatibilidad amplia (tablas + filas extra)
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("3_Calc")
    calc_headers = [
        "ID",
        "Operacion",
        "Responsable",
        "Fecha_Inicio",
        "Fecha_Finalizacion",
        "Prioridad_Raw",
        "Prioridad_Norm",
        "Estado",
        "Peso",
        "Dias_Ciclo",
        "Meta_Dias",
        "A_Tiempo",
        "Eficiencia_Ciclo",
        "En_Quincena",
        "Es_Cerrada",
        "Es_Atraso_Abierto",
        "Incluir_KPI",
        "Peso_x_A_Tiempo",
        "Peso_x_Eficiencia",
    ]
    for c, h in enumerate(calc_headers, start=1):
        ws3.cell(1, c, h)
    style_header_row(ws3, 1, len(calc_headers))

    # Pre-cargar fórmulas para filas 2..N (demo + buffer). Usamos 200 filas útiles.
    max_calc = max(last_data_row - 1, 120)  # filas de datos posibles
    for i in range(2, max_calc + 2):
        src = i + 1  # 1_Datos empieza en fila 3
        # ID .. Estado mapeo directo
        ws3.cell(i, 1, f'=IF(\'1_Datos\'!A{src}="","",\'1_Datos\'!A{src})')
        ws3.cell(i, 2, f'=IF(A{i}="","",\'1_Datos\'!B{src})')
        ws3.cell(i, 3, f'=IF(A{i}="","",\'1_Datos\'!C{src})')
        ws3.cell(i, 4, f'=IF(A{i}="","",\'1_Datos\'!D{src})')
        ws3.cell(i, 4).number_format = "DD/MM/YYYY"
        ws3.cell(i, 5, f'=IF(A{i}="","",IF(\'1_Datos\'!E{src}="","",\'1_Datos\'!E{src}))')
        ws3.cell(i, 5).number_format = "DD/MM/YYYY"
        ws3.cell(i, 6, f'=IF(A{i}="","",\'1_Datos\'!F{src})')
        # Prioridad normalizada
        ws3.cell(
            i,
            7,
            f'=IF(A{i}="","",IF(OR(UPPER(F{i})="ALTA",UPPER(F{i})="P1",F{i}=1),"Alta",'
            f'IF(OR(UPPER(F{i})="MEDIA",UPPER(F{i})="P2",F{i}=2),"Media",'
            f'IF(OR(UPPER(F{i})="BAJA",UPPER(F{i})="P3",F{i}=3),"Baja","Media"))))',
        )
        ws3.cell(i, 8, f'=IF(A{i}="","",\'1_Datos\'!G{src})')
        # Peso
        ws3.cell(
            i,
            9,
            f'=IF(A{i}="","",IF(G{i}="Alta",\'2_Parametros\'!$B$7,'
            f'IF(G{i}="Media",\'2_Parametros\'!$B$8,\'2_Parametros\'!$B$9)))',
        )
        # Días ciclo (si hay fin; si no, días hasta fin quincena para abiertas)
        ws3.cell(
            i,
            10,
            f'=IF(A{i}="","",IF(E{i}<>"",E{i}-D{i}+1,'
            f'IF(D{i}="","",\'2_Parametros\'!$B$24-D{i}+1)))',
        )
        # Meta días
        ws3.cell(
            i,
            11,
            f'=IF(A{i}="","",IF(G{i}="Alta",\'2_Parametros\'!$B$11,'
            f'IF(G{i}="Media",\'2_Parametros\'!$B$12,\'2_Parametros\'!$B$13)))',
        )
        # A tiempo: cerrada y dias_ciclo <= meta
        ws3.cell(
            i,
            12,
            f'=IF(A{i}="","",IF(AND(UPPER(H{i})="COMPLETADA",E{i}<>"",J{i}<=K{i}),1,0))',
        )
        # Eficiencia ciclo (tope 1)
        ws3.cell(
            i,
            13,
            f'=IF(A{i}="","",IF(OR(J{i}="",J{i}=0),0,MIN(1,K{i}/J{i})))',
        )
        ws3.cell(i, 13).number_format = "0.0%"
        # En quincena por fecha fin o inicio
        ws3.cell(
            i,
            14,
            f'=IF(A{i}="","",IF(OR(AND(E{i}<>"",E{i}>=\'2_Parametros\'!$B$23,E{i}<=\'2_Parametros\'!$B$24),'
            f'AND(D{i}<>"",D{i}>=\'2_Parametros\'!$B$23,D{i}<=\'2_Parametros\'!$B$24)),1,0))',
        )
        # Es cerrada
        ws3.cell(
            i,
            15,
            f'=IF(A{i}="","",IF(AND(UPPER(H{i})="COMPLETADA",E{i}<>""),1,0))',
        )
        # Atraso abierto: no completada, inicio+meta < fin quincena, aún abierta
        ws3.cell(
            i,
            16,
            f'=IF(A{i}="","",IF(AND(UPPER(H{i})<>"COMPLETADA",UPPER(H{i})<>"CANCELADA",'
            f'D{i}<>"",D{i}+K{i}-1<\'2_Parametros\'!$B$24),1,0))',
        )
        # Incluir en KPI de la quincena:
        # (cerrada con fin en quincena) OR (atraso abierto si param=1)
        ws3.cell(
            i,
            17,
            f'=IF(A{i}="","",IF(OR(AND(O{i}=1,E{i}>=\'2_Parametros\'!$B$23,E{i}<=\'2_Parametros\'!$B$24),'
            f'AND(\'2_Parametros\'!$B$21=1,P{i}=1,N{i}=1)),1,0))',
        )
        # Peso x a tiempo / eficiencia (solo si Incluir_KPI)
        ws3.cell(i, 18, f'=IF(A{i}="","",IF(Q{i}=1,I{i}*L{i},0))')
        ws3.cell(i, 19, f'=IF(A{i}="","",IF(Q{i}=1,I{i}*M{i},0))')

        for c in range(1, 20):
            ws3.cell(i, c).border = THIN
            ws3.cell(i, c).font = FONT_BODY

    autosize(ws3, min_width=11, max_width=18)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:S{max_calc + 1}"

    # ------------------------------------------------------------------
    # 4_KPI_Responsables
    # ------------------------------------------------------------------
    ws4 = wb.create_sheet("4_KPI_Responsables")
    ws4["A1"] = "KPI de Operatividad por Responsable"
    ws4["A1"].font = FONT_TITLE
    ws4["A1"].fill = FILL_TITLE
    ws4.merge_cells("A1:L1")
    ws4["A2"] = '="Quincena activa: "&\'2_Parametros\'!B25&"  |  Periodo: "&TEXT(\'2_Parametros\'!B23,"DD/MM/YYYY")&" - "&TEXT(\'2_Parametros\'!B24,"DD/MM/YYYY")'
    ws4["A2"].font = FONT_BOLD
    ws4.merge_cells("A2:L2")

    kpi_headers = [
        "Responsable",
        "Ops_Evaluadas",
        "Ops_Completadas",
        "Ops_A_Tiempo",
        "Peso_Total",
        "Cumplimiento_Ponderado",
        "Eficiencia_Ciclo_Ponderada",
        "IOI",
        "Semaforo",
        "Dias_Ciclo_Promedio",
        "Ops_Alta_Prioridad",
        "Brecha_vs_Meta",
    ]
    for c, h in enumerate(kpi_headers, start=1):
        ws4.cell(4, c, h)
    style_header_row(ws4, 4, len(kpi_headers))

    # Lista única de responsables (hasta 25) con fórmula UNIQUE si hay Excel 365;
    # además dejamos nombres demo fijos + filas dinámicas con INDEX/MATCH fallback.
    # Usamos lista derivada de demo + celdas para pegar extras.
    demo_resps = sorted({r[2] for r in demo})
    for idx, resp in enumerate(demo_resps, start=5):
        r = idx
        ws4.cell(r, 1, resp).font = FONT_BOLD
        # Ops evaluadas
        ws4.cell(r, 2, f'=COUNTIFS(\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1)')
        # Completadas en evaluación
        ws4.cell(
            r,
            3,
            f'=COUNTIFS(\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1,\'3_Calc\'!$O:$O,1)',
        )
        # A tiempo
        ws4.cell(
            r,
            4,
            f'=SUMIFS(\'3_Calc\'!$L:$L,\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1)',
        )
        # Peso total
        ws4.cell(
            r,
            5,
            f'=SUMIFS(\'3_Calc\'!$I:$I,\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1)',
        )
        # Cumplimiento ponderado
        ws4.cell(
            r,
            6,
            f'=IF(E{r}=0,"",SUMIFS(\'3_Calc\'!$R:$R,\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1)/E{r})',
        )
        ws4.cell(r, 6).number_format = "0.0%"
        # Eficiencia ponderada
        ws4.cell(
            r,
            7,
            f'=IF(E{r}=0,"",SUMIFS(\'3_Calc\'!$S:$S,\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1)/E{r})',
        )
        ws4.cell(r, 7).number_format = "0.0%"
        # IOI
        ws4.cell(
            r,
            8,
            f'=IF(E{r}=0,"",F{r}*\'2_Parametros\'!$B$15+G{r}*\'2_Parametros\'!$B$16)',
        )
        ws4.cell(r, 8).number_format = "0.0%"
        # Semáforo
        ws4.cell(
            r,
            9,
            f'=IF(H{r}="","",IF(H{r}>=\'2_Parametros\'!$B$18,"Cumple",'
            f'IF(H{r}>=\'2_Parametros\'!$B$19,"Riesgo","Critico")))',
        )
        # Días ciclo promedio (solo cerradas evaluadas)
        ws4.cell(
            r,
            10,
            f'=IFERROR(AVERAGEIFS(\'3_Calc\'!$J:$J,\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1,\'3_Calc\'!$O:$O,1),"")',
        )
        ws4.cell(r, 10).number_format = "0.0"
        # Ops alta prioridad evaluadas
        ws4.cell(
            r,
            11,
            f'=COUNTIFS(\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1,\'3_Calc\'!$G:$G,"Alta")',
        )
        # Brecha vs meta
        ws4.cell(r, 12, f'=IF(H{r}="","",H{r}-\'2_Parametros\'!$B$18)')
        ws4.cell(r, 12).number_format = "0.0%"

        for c in range(1, 13):
            ws4.cell(r, c).border = THIN

    # Filas extra vacías para nuevos responsables: usuario escribe el nombre en col A
    for r in range(5 + len(demo_resps), 5 + len(demo_resps) + 10):
        ws4.cell(r, 1).fill = FILL_INPUT
        ws4.cell(r, 2, f'=IF(A{r}="","",COUNTIFS(\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1))')
        ws4.cell(
            r,
            3,
            f'=IF(A{r}="","",COUNTIFS(\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1,\'3_Calc\'!$O:$O,1))',
        )
        ws4.cell(
            r,
            4,
            f'=IF(A{r}="","",SUMIFS(\'3_Calc\'!$L:$L,\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1))',
        )
        ws4.cell(
            r,
            5,
            f'=IF(A{r}="","",SUMIFS(\'3_Calc\'!$I:$I,\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1))',
        )
        ws4.cell(
            r,
            6,
            f'=IF(OR(A{r}="",E{r}=0),"",SUMIFS(\'3_Calc\'!$R:$R,\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1)/E{r})',
        )
        ws4.cell(r, 6).number_format = "0.0%"
        ws4.cell(
            r,
            7,
            f'=IF(OR(A{r}="",E{r}=0),"",SUMIFS(\'3_Calc\'!$S:$S,\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1)/E{r})',
        )
        ws4.cell(r, 7).number_format = "0.0%"
        ws4.cell(
            r,
            8,
            f'=IF(OR(A{r}="",E{r}=0),"",F{r}*\'2_Parametros\'!$B$15+G{r}*\'2_Parametros\'!$B$16)',
        )
        ws4.cell(r, 8).number_format = "0.0%"
        ws4.cell(
            r,
            9,
            f'=IF(H{r}="","",IF(H{r}>=\'2_Parametros\'!$B$18,"Cumple",'
            f'IF(H{r}>=\'2_Parametros\'!$B$19,"Riesgo","Critico")))',
        )
        ws4.cell(
            r,
            10,
            f'=IF(A{r}="","",IFERROR(AVERAGEIFS(\'3_Calc\'!$J:$J,\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1,\'3_Calc\'!$O:$O,1),""))',
        )
        ws4.cell(r, 10).number_format = "0.0"
        ws4.cell(
            r,
            11,
            f'=IF(A{r}="","",COUNTIFS(\'3_Calc\'!$C:$C,$A{r},\'3_Calc\'!$Q:$Q,1,\'3_Calc\'!$G:$G,"Alta"))',
        )
        ws4.cell(r, 12, f'=IF(H{r}="","",H{r}-\'2_Parametros\'!$B$18)')
        ws4.cell(r, 12).number_format = "0.0%"
        for c in range(1, 13):
            ws4.cell(r, c).border = THIN

    # Totales departamento
    total_row = 5 + len(demo_resps) + 11
    ws4.cell(total_row, 1, "TOTAL DEPARTAMENTO I+D").font = FONT_WHITE
    ws4.cell(total_row, 1).fill = FILL_HEADER
    end_resp = 5 + len(demo_resps) + 9
    ws4.cell(total_row, 2, f"=SUM(B5:B{end_resp})")
    ws4.cell(total_row, 3, f"=SUM(C5:C{end_resp})")
    ws4.cell(total_row, 4, f"=SUM(D5:D{end_resp})")
    ws4.cell(total_row, 5, f"=SUM(E5:E{end_resp})")
    ws4.cell(total_row, 6, f'=IF(E{total_row}=0,"",SUMPRODUCT((E5:E{end_resp})*(F5:F{end_resp}))/E{total_row})')
    ws4.cell(total_row, 6).number_format = "0.0%"
    ws4.cell(total_row, 7, f'=IF(E{total_row}=0,"",SUMPRODUCT((E5:E{end_resp})*(G5:G{end_resp}))/E{total_row})')
    ws4.cell(total_row, 7).number_format = "0.0%"
    ws4.cell(
        total_row,
        8,
        f'=IF(E{total_row}=0,"",F{total_row}*\'2_Parametros\'!$B$15+G{total_row}*\'2_Parametros\'!$B$16)',
    )
    ws4.cell(total_row, 8).number_format = "0.0%"
    ws4.cell(
        total_row,
        9,
        f'=IF(H{total_row}="","",IF(H{total_row}>=\'2_Parametros\'!$B$18,"Cumple",'
        f'IF(H{total_row}>=\'2_Parametros\'!$B$19,"Riesgo","Critico")))',
    )
    ws4.cell(total_row, 11, f"=SUM(K5:K{end_resp})")
    ws4.cell(total_row, 12, f'=IF(H{total_row}="","",H{total_row}-\'2_Parametros\'!$B$18)')
    ws4.cell(total_row, 12).number_format = "0.0%"
    for c in range(1, 13):
        ws4.cell(total_row, c).fill = FILL_HEADER
        ws4.cell(total_row, c).font = FONT_WHITE
        ws4.cell(total_row, c).border = THIN

    # Formato condicional IOI
    ws4.conditional_formatting.add(
        f"H5:H{end_resp}",
        ColorScaleRule(
            start_type="num",
            start_value=0.5,
            start_color="FFC7CE",
            mid_type="num",
            mid_value=0.7,
            mid_color="FFEB9C",
            end_type="num",
            end_value=0.9,
            end_color="C6EFCE",
        ),
    )

    # Nota metodológica corta
    note_row = total_row + 2
    ws4.cell(note_row, 1, "Fórmulas vivas: al renovar 1_Datos o cambiar 2_Parametros, esta hoja se actualiza sola.").font = FONT_BODY
    ws4.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)

    autosize(ws4, min_width=12, max_width=28)
    ws4.freeze_panes = "B5"

    # ------------------------------------------------------------------
    # 5_Ranking
    # ------------------------------------------------------------------
    ws5 = wb.create_sheet("5_Ranking")
    ws5["A1"] = "Ranking de Operatividad I+D (quincena activa)"
    ws5["A1"].font = FONT_TITLE
    ws5["A1"].fill = FILL_TITLE
    ws5.merge_cells("A1:D1")
    ws5["A3"] = "Posición"
    ws5["B3"] = "Responsable"
    ws5["C3"] = "IOI"
    ws5["D3"] = "Interpretación"
    style_header_row(ws5, 3, 4)

    # Ranking manual ordenado por IOI usando LARGE + INDEX/MATCH
    n_rank = len(demo_resps)
    for i in range(1, n_rank + 1):
        r = 3 + i
        ws5.cell(r, 1, i)
        # Responsable del i-ésimo mayor IOI
        ws5.cell(
            r,
            2,
            f'=IFERROR(INDEX(\'4_KPI_Responsables\'!$A$5:$A${4+n_rank},'
            f'MATCH(LARGE(\'4_KPI_Responsables\'!$H$5:$H${4+n_rank},{i}),'
            f'\'4_KPI_Responsables\'!$H$5:$H${4+n_rank},0)),"")',
        )
        ws5.cell(
            r,
            3,
            f'=IFERROR(LARGE(\'4_KPI_Responsables\'!$H$5:$H${4+n_rank},{i}),"")',
        )
        ws5.cell(r, 3).number_format = "0.0%"
        ws5.cell(
            r,
            4,
            f'=IF(C{r}="","",IF(C{r}>=\'2_Parametros\'!$B$18,"Operatividad alta",'
            f'IF(C{r}>=\'2_Parametros\'!$B$19,"Operatividad media — revisar atrasos",'
            f'"Operatividad baja — priorizar cierres Alta")))',
        )
        for c in range(1, 5):
            ws5.cell(r, c).border = THIN

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "IOI por responsable (quincena)"
    chart.y_axis.title = None
    chart.x_axis.title = "IOI"
    data = Reference(ws5, min_col=3, min_row=3, max_row=3 + n_rank)
    cats = Reference(ws5, min_col=2, min_row=4, max_row=3 + n_rank)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 18
    chart.height = 10
    ws5.add_chart(chart, "F3")

    autosize(ws5)
    ws5.column_dimensions["D"].width = 48

    # ------------------------------------------------------------------
    # 6_Metodologia
    # ------------------------------------------------------------------
    ws6 = wb.create_sheet("6_Metodologia")
    ws6["A1"] = "Propuesta metodológica — Indicador de Operatividad de Responsables I+D"
    ws6["A1"].font = FONT_TITLE
    ws6["A1"].fill = FILL_TITLE
    ws6.merge_cells("A1:B1")

    metodologia = [
        ("Nombre del indicador", "IOI — Índice de Operatividad I+D"),
        ("Unidad", "Porcentaje (0%–100%)"),
        ("Frecuencia", "Quincenal (Q1: días 1–15; Q2: días 16–fin de mes)"),
        ("Alcance", "Todos los responsables con operaciones en el departamento I+D"),
        (
            "Insumos",
            "Fecha_Inicio, Fecha_Finalizacion, Prioridad, Responsable, Estado (desde KPIS I+D.xlsx)",
        ),
        (
            "Fórmula IOI",
            "IOI = 70% × Cumplimiento_Ponderado + 30% × Eficiencia_Ciclo_Ponderada",
        ),
        (
            "Cumplimiento_Ponderado",
            "Σ(Peso_prioridad × A_Tiempo) / Σ(Peso_prioridad) sobre ops evaluadas en la quincena",
        ),
        (
            "A_Tiempo",
            "1 si Estado=Completada y (Fecha_Finalizacion − Fecha_Inicio + 1) ≤ Meta_días(prioridad); si no, 0",
        ),
        (
            "Eficiencia_Ciclo_Ponderada",
            "Σ(Peso × min(1, Meta_días / Días_ciclo)) / Σ(Peso)",
        ),
        (
            "Pesos de prioridad (default)",
            "Alta=3, Media=2, Baja=1 — refleja que un atraso en Alta impacta más la operatividad",
        ),
        (
            "Metas de ciclo (default)",
            "Alta=5 días, Media=9 días, Baja=14 días — ajustables en 2_Parametros",
        ),
        (
            "Ops evaluadas en la quincena",
            "Cierres con Fecha_Finalizacion en la quincena + (opcional) abiertas atrasadas activas en la quincena",
        ),
        (
            "Por qué este diseño",
            "Combina puntualidad y velocidad real, ponderadas por criticidad; evita premiar solo volumen sin calidad de cierre",
        ),
        (
            "Actualización de datos",
            "Renovar filas en 1_Datos; los KPI se recalculan por fórmulas. No hardcodear resultados.",
        ),
        (
            "Uso gerencial",
            "Comparar IOI entre responsables cada quincena; intervenir primero en Critico y en Alta prioridad atrasada",
        ),
        (
            "Limitación / nota",
            "Si KPIS I+D.xlsx trae columnas con otros nombres, mapear al encabezado de 1_Datos. Los datos demo son ilustrativos.",
        ),
    ]
    ws6["A3"] = "Concepto"
    ws6["B3"] = "Definición"
    style_header_row(ws6, 3, 2)
    for i, (k, v) in enumerate(metodologia, start=4):
        ws6.cell(i, 1, k).font = FONT_BOLD
        ws6.cell(i, 1).fill = FILL_SOFT
        ws6.cell(i, 1).border = THIN
        ws6.cell(i, 2, v).font = FONT_BODY
        ws6.cell(i, 2).alignment = Alignment(wrap_text=True, vertical="center")
        ws6.cell(i, 2).border = THIN
        ws6.row_dimensions[i].height = 36

    ws6.column_dimensions["A"].width = 34
    ws6.column_dimensions["B"].width = 100

    # ------------------------------------------------------------------
    # 7_Snapshot — resultados precalculados (útiles al abrir sin recalcular)
    # ------------------------------------------------------------------
    ws7 = wb.create_sheet("7_Snapshot_Resultados")
    ws7["A1"] = "Snapshot calculado (datos demo / última generación del script)"
    ws7["A1"].font = FONT_TITLE
    ws7["A1"].fill = FILL_TITLE
    ws7.merge_cells("A1:H1")
    ws7["A2"] = (
        "Esta hoja muestra valores ya calculados en Python. "
        "Para operación diaria use 4_KPI_Responsables (fórmulas Excel). "
        "Al renovar 1_Datos, vuelva a ejecutar el script o confíe en las fórmulas."
    )
    ws7["A2"].alignment = Alignment(wrap_text=True)
    ws7.merge_cells("A2:H2")
    ws7.row_dimensions[2].height = 36

    snap_headers = [
        "Responsable",
        "Ops_Evaluadas",
        "Ops_Completadas",
        "Ops_A_Tiempo",
        "Cumplimiento_Ponderado",
        "Eficiencia_Ciclo_Ponderada",
        "IOI",
        "Semaforo",
    ]
    for c, h in enumerate(snap_headers, start=1):
        ws7.cell(4, c, h)
    style_header_row(ws7, 4, len(snap_headers))

    # Parámetros del snapshot (= defaults del libro)
    year, month, q = 2026, 8, 1
    start_q = date(year, month, 1 if q == 1 else 16)
    if q == 1:
        end_q = date(year, month, 15)
    else:
        end_q = (date(year, month + 1, 1) - timedelta(days=1)) if month < 12 else date(year, 12, 31)
    weights = {"Alta": 3, "Media": 2, "Baja": 1}
    metas = {"Alta": 5, "Media": 9, "Baja": 14}
    w_c, w_e, meta_ioi, umbral = 0.70, 0.30, 0.85, 0.70

    def norm_p(p) -> str:
        s = str(p).strip().upper()
        if s in ("ALTA", "P1", "1"):
            return "Alta"
        if s in ("MEDIA", "P2", "2"):
            return "Media"
        if s in ("BAJA", "P3", "3"):
            return "Baja"
        return "Media"

    stats: dict = defaultdict(
        lambda: {
            "peso": 0.0,
            "px_at": 0.0,
            "px_ef": 0.0,
            "ops": 0,
            "comp": 0,
            "at": 0,
        }
    )
    for id_, op, resp, fi, ff, prio, estado, area, obs in demo:
        if not resp or not fi:
            continue
        g = norm_p(prio)
        peso = weights[g]
        meta_d = metas[g]
        estado_u = str(estado or "").upper()
        closed = estado_u == "COMPLETADA" and ff not in (None, "")
        if closed:
            dias = (ff - fi).days + 1
        else:
            dias = (end_q - fi).days + 1
        a_tiempo = 1 if closed and dias <= meta_d else 0
        ef = min(1.0, meta_d / dias) if dias else 0.0
        atraso_abierto = estado_u not in ("COMPLETADA", "CANCELADA") and (
            fi + timedelta(days=meta_d - 1) < end_q
        )
        en_quincena = (closed and start_q <= ff <= end_q) or (start_q <= fi <= end_q)
        incluir = (closed and start_q <= ff <= end_q) or (atraso_abierto and en_quincena)
        if not incluir:
            continue
        s = stats[resp]
        s["ops"] += 1
        s["peso"] += peso
        s["px_at"] += peso * a_tiempo
        s["px_ef"] += peso * ef
        s["at"] += a_tiempo
        if closed:
            s["comp"] += 1

    ws7["A3"] = f"Periodo snapshot: Q{q} {month:02d}/{year} ({start_q.isoformat()} a {end_q.isoformat()})"
    ws7["A3"].font = FONT_BOLD

    r = 5
    for resp in sorted(stats.keys()):
        s = stats[resp]
        cumpl = s["px_at"] / s["peso"] if s["peso"] else 0
        efic = s["px_ef"] / s["peso"] if s["peso"] else 0
        ioi = cumpl * w_c + efic * w_e
        if ioi >= meta_ioi:
            sem = "Cumple"
            fill = FILL_OK
        elif ioi >= umbral:
            sem = "Riesgo"
            fill = FILL_WARN
        else:
            sem = "Critico"
            fill = FILL_BAD
        values = [resp, s["ops"], s["comp"], s["at"], cumpl, efic, ioi, sem]
        for c, val in enumerate(values, start=1):
            cell = ws7.cell(r, c, val)
            cell.border = THIN
            cell.font = FONT_BODY
            if c in (5, 6, 7):
                cell.number_format = "0.0%"
            if c == 8:
                cell.fill = fill
                cell.font = FONT_BOLD
        r += 1

    # Totales
    if stats:
        tot_peso = sum(s["peso"] for s in stats.values())
        tot_cumpl = sum(s["px_at"] for s in stats.values()) / tot_peso if tot_peso else 0
        tot_efic = sum(s["px_ef"] for s in stats.values()) / tot_peso if tot_peso else 0
        tot_ioi = tot_cumpl * w_c + tot_efic * w_e
        ws7.cell(r, 1, "TOTAL I+D").font = FONT_WHITE
        ws7.cell(r, 2, sum(s["ops"] for s in stats.values()))
        ws7.cell(r, 3, sum(s["comp"] for s in stats.values()))
        ws7.cell(r, 4, sum(s["at"] for s in stats.values()))
        ws7.cell(r, 5, tot_cumpl).number_format = "0.0%"
        ws7.cell(r, 6, tot_efic).number_format = "0.0%"
        ws7.cell(r, 7, tot_ioi).number_format = "0.0%"
        ws7.cell(r, 8, "Cumple" if tot_ioi >= meta_ioi else ("Riesgo" if tot_ioi >= umbral else "Critico"))
        for c in range(1, 9):
            ws7.cell(r, c).fill = FILL_HEADER
            ws7.cell(r, c).font = FONT_WHITE
            ws7.cell(r, c).border = THIN

    autosize(ws7)

    # Orden de hojas
    order = [
        "0_Instrucciones",
        "1_Datos",
        "2_Parametros",
        "3_Calc",
        "4_KPI_Responsables",
        "5_Ranking",
        "6_Metodologia",
        "7_Snapshot_Resultados",
    ]
    for i, name in enumerate(order):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    return wb


def main() -> None:
    wb = build_workbook()
    OUT_REPO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_REPO)
    OUT_ARTIFACT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_ARTIFACT)
    print(f"OK -> {OUT_REPO}")
    print(f"OK -> {OUT_ARTIFACT}")


if __name__ == "__main__":
    main()
