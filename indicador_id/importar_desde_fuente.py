#!/usr/bin/env python3
"""
Importa operaciones desde un Excel fuente (p.ej. KPIS I+D.xlsx) hacia
Indicador_Operatividad_ID_Quincenal.xlsx (hoja 1_Datos), preservando fórmulas.

Uso:
  python3 indicador_id/importar_desde_fuente.py "/ruta/KPIS I+D.xlsx"

Si los nombres de columna difieren, edite COLUMN_ALIASES abajo.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

TARGET = Path(__file__).resolve().parent / "Indicador_Operatividad_ID_Quincenal.xlsx"
ARTIFACT = Path("/opt/cursor/artifacts/Indicador_Operatividad_ID_Quincenal.xlsx")

CANONICAL = [
    "ID",
    "Operacion",
    "Responsable",
    "Fecha_Inicio",
    "Fecha_Finalizacion",
    "Prioridad",
    "Estado",
    "Area",
    "Observaciones",
]

COLUMN_ALIASES = {
    "ID": ["id", "codigo", "código", "nro", "numero", "número", "op"],
    "Operacion": [
        "operacion",
        "operación",
        "actividad",
        "tarea",
        "descripcion",
        "descripción",
        "nombre",
    ],
    "Responsable": ["responsable", "owner", "asignado", "encargado", "analista"],
    "Fecha_Inicio": [
        "fecha_inicio",
        "fecha inicio",
        "inicio",
        "f_inicio",
        "start",
        "fecha de inicio",
    ],
    "Fecha_Finalizacion": [
        "fecha_finalizacion",
        "fecha finalizacion",
        "fecha_fin",
        "fecha fin",
        "finalizacion",
        "finalización",
        "cierre",
        "fecha de finalizacion",
        "fecha de finalización",
        "end",
    ],
    "Prioridad": ["prioridad", "priority", "nivel", "nivel_prioridad", "p"],
    "Estado": ["estado", "status", "situacion", "situación"],
    "Area": ["area", "área", "proceso", "categoria", "categoría"],
    "Observaciones": ["observaciones", "obs", "notas", "comentario", "comentarios"],
}


def _norm(s: str) -> str:
    return (
        str(s)
        .strip()
        .lower()
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )


def map_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapping = {}
    cols_norm = {_norm(c): c for c in df.columns}
    for canon, aliases in COLUMN_ALIASES.items():
        if canon in df.columns:
            mapping[canon] = canon
            continue
        for alias in aliases + [_norm(canon)]:
            if alias in cols_norm:
                mapping[canon] = cols_norm[alias]
                break
    missing = [c for c in ("Responsable", "Fecha_Inicio", "Prioridad") if c not in mapping]
    if missing:
        raise SystemExit(
            f"No se pudieron mapear columnas obligatorias: {missing}. "
            f"Columnas encontradas: {list(df.columns)}"
        )
    out = pd.DataFrame()
    for canon in CANONICAL:
        out[canon] = df[mapping[canon]] if canon in mapping else None
    return out


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Uso: python3 importar_desde_fuente.py \"/ruta/KPIS I+D.xlsx\" [hoja]")
    src = Path(sys.argv[1])
    sheet = sys.argv[2] if len(sys.argv) > 2 else 0
    if not src.exists():
        raise SystemExit(f"No existe: {src}")
    if not TARGET.exists():
        raise SystemExit(f"Primero genere el indicador: falta {TARGET}")

    df = pd.read_excel(src, sheet_name=sheet)
    mapped = map_columns(df)

    wb = load_workbook(TARGET)
    ws = wb["1_Datos"]
    # Limpia datos previos (conserva encabezado fila 2)
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)
    # Quita tabla previa si existe para recrearla simple
    if "TablaDatos" in ws.tables:
        del ws.tables["TablaDatos"]

    for r_idx, row in enumerate(mapped.itertuples(index=False), start=3):
        for c_idx, val in enumerate(row, start=1):
            if pd.isna(val):
                val = None
            cell = ws.cell(r_idx, c_idx, val)
            if c_idx in (4, 5) and val is not None:
                cell.number_format = "DD/MM/YYYY"

    from openpyxl.worksheet.table import Table, TableStyleInfo

    last = 2 + len(mapped)
    if last >= 3:
        table = Table(displayName="TablaDatos", ref=f"A2:I{last}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)

    wb.save(TARGET)
    wb.save(ARTIFACT)
    print(f"Importadas {len(mapped)} filas desde {src.name} -> {TARGET}")
    print("Abra Excel y revise 4_KPI_Responsables / o regenere snapshot con generar_indicador_operatividad.py")


if __name__ == "__main__":
    main()
