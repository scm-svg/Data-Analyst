#!/usr/bin/env python3
"""
Análisis de solicitudes de consumibles - TALLER y TIENDAS.
Genera Excel con demanda, min/max, pedidos recomendados (semanal, quincenal, mensual).
"""

import re
import unicodedata
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TALLER_FILE = "/home/ubuntu/.cursor/projects/workspace/uploads/Solicitudes__TALLER__consumibles_eab4.xlsx"
TIENDAS_FILE = "/home/ubuntu/.cursor/projects/workspace/uploads/Solicitudes__TIENDAS__consumibles_ca36.xlsx"
OUTPUT_FILE = "/workspace/Analisis_Consumibles_Pedidos_Recomendados.xlsx"

FECHA_MIN = pd.Timestamp("2024-01-01")
FECHA_MAX = pd.Timestamp("2027-12-31")
SAFETY_FACTOR = 1.15  # 15% colchón sobre demanda promedio


def normalize_text(s):
    if pd.isna(s):
        return ""
    s = str(s).strip().upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s


def parse_cantidad(x):
    if pd.isna(x):
        return np.nan
    s = str(x).strip().replace(",", ".")
    m = re.search(r"(\d+(?:\.\d+)?)", s)
    return float(m.group(1)) if m else np.nan


def parse_fecha(x):
    if pd.isna(x):
        return pd.NaT
    dt = pd.to_datetime(x, errors="coerce")
    return dt


def classify_estado(estado_norm):
    e = str(estado_norm).strip().upper()
    if not e:
        return "SIN ESTADO"
    if any(k in e for k in ("ENTREG", "RECIB", "DISPON", "ENVIAD", "PROCES")):
        return "ATENDIDO"
    if any(k in e for k in ("NO DISP", "SOLICIT", "PEND", "CANCEL", "RECHAZ")):
        return "NO ATENDIDO"
    return "OTRO"


def load_taller():
    df = pd.read_excel(TALLER_FILE, sheet_name="SOLICITUDES TALLER")
    df = df[
        ["FECHA", "Nombre", "Articulo", "Cantidad", "Estado", "SUCURSAL", "Notas y Comentarios"]
    ].copy()
    df = df[df["Articulo"].notna() & (df["Articulo"].astype(str).str.strip() != "")]
    df["FECHA"] = df["FECHA"].apply(parse_fecha)
    df = df[(df["FECHA"] >= FECHA_MIN) & (df["FECHA"] <= FECHA_MAX)]
    df["CANTIDAD"] = df["Cantidad"].apply(parse_cantidad)
    df["ARTICULO"] = df["Articulo"].apply(lambda x: str(x).strip())
    df["ARTICULO_NORM"] = df["ARTICULO"].apply(normalize_text)
    df["ESTADO_RAW"] = df["Estado"].fillna("").astype(str)
    df["ESTADO_NORM"] = df["ESTADO_RAW"].apply(normalize_text)
    df["CLASIF_ESTADO"] = df["ESTADO_NORM"].apply(classify_estado)
    df["SOLICITANTE"] = df["Nombre"].fillna("").astype(str)
    df["SUCURSAL"] = df["SUCURSAL"].fillna("TALLER").astype(str)
    df["NOTAS"] = df["Notas y Comentarios"].fillna("").astype(str)
    df["ORIGEN"] = "TALLER"
    df["TIENDA_HOJA"] = ""
    return df[
        [
            "FECHA",
            "ORIGEN",
            "SUCURSAL",
            "TIENDA_HOJA",
            "SOLICITANTE",
            "ARTICULO",
            "ARTICULO_NORM",
            "CANTIDAD",
            "ESTADO_RAW",
            "CLASIF_ESTADO",
            "NOTAS",
        ]
    ]


def load_tiendas():
    xl = pd.ExcelFile(TIENDAS_FILE)
    store_sheets = [s for s in xl.sheet_names if s != "INVENTARIO DE CONSUMIBLES"]
    frames = []
    for sheet in store_sheets:
        raw = pd.read_excel(TIENDAS_FILE, sheet_name=sheet, header=None)
        header_row = None
        for i, row in raw.iterrows():
            if str(row.iloc[0]).strip().upper() == "FECHA":
                header_row = i
                break
        if header_row is None:
            continue
        df = pd.read_excel(TIENDAS_FILE, sheet_name=sheet, header=header_row)
        df.columns = [
            "FECHA",
            "NOMBRE",
            "ARTICULO",
            "CARACTER_ADICIONAL",
            "CANTIDAD",
            "ESTADO",
            "SUCURSAL",
            "NOTAS",
        ]
        df = df[df["ARTICULO"].notna() & (df["ARTICULO"].astype(str).str.strip() != "")]
        df["FECHA"] = df["FECHA"].apply(parse_fecha)
        df = df[df["FECHA"].notna()]
        df = df[(df["FECHA"] >= FECHA_MIN) & (df["FECHA"] <= FECHA_MAX)]
        df["CANTIDAD"] = df["CANTIDAD"].apply(parse_cantidad)
        df["ARTICULO"] = df["ARTICULO"].apply(lambda x: str(x).strip())
        if df["CARACTER_ADICIONAL"].notna().any():
            mask = df["CARACTER_ADICIONAL"].notna() & (
                df["CARACTER_ADICIONAL"].astype(str).str.strip() != ""
            )
            df.loc[mask, "ARTICULO"] = (
                df.loc[mask, "ARTICULO"] + " (" + df.loc[mask, "CARACTER_ADICIONAL"].astype(str) + ")"
            )
        df["ARTICULO_NORM"] = df["ARTICULO"].apply(normalize_text)
        df["ESTADO_RAW"] = df["ESTADO"].fillna("").astype(str)
        df["ESTADO_NORM"] = df["ESTADO_RAW"].apply(normalize_text)
        df["CLASIF_ESTADO"] = df["ESTADO_NORM"].apply(classify_estado)
        df["SOLICITANTE"] = df["NOMBRE"].fillna("").astype(str)
        df["SUCURSAL"] = df["SUCURSAL"].fillna(sheet).astype(str)
        df["NOTAS"] = df["NOTAS"].fillna("").astype(str)
        df["ORIGEN"] = "TIENDAS"
        df["TIENDA_HOJA"] = sheet
        frames.append(
            df[
                [
                    "FECHA",
                    "ORIGEN",
                    "SUCURSAL",
                    "TIENDA_HOJA",
                    "SOLICITANTE",
                    "ARTICULO",
                    "ARTICULO_NORM",
                    "CANTIDAD",
                    "ESTADO_RAW",
                    "CLASIF_ESTADO",
                    "NOTAS",
                ]
            ]
        )
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def load_catalogo():
    inv_t = pd.read_excel(TALLER_FILE, sheet_name="INVENTARIO DE CONSUMIBLES")
    inv_t = inv_t.rename(columns={"PRODUCTO": "PRODUCTO_CAT", "CATEGORÍA / FAMILIA": "CATEGORIA"})
    inv_t["ARTICULO_NORM"] = inv_t["PRODUCTO_CAT"].apply(normalize_text)
    inv_t = inv_t[["ARTICULO_NORM", "CÓDIGO", "PRODUCTO_CAT", "CATEGORIA"]].drop_duplicates("ARTICULO_NORM")

    inv_s = pd.read_excel(TIENDAS_FILE, sheet_name="INVENTARIO DE CONSUMIBLES")
    inv_s = inv_s.rename(columns={"PRODUNTOS": "PRODUCTO_CAT", "ID": "CÓDIGO"})
    inv_s["ARTICULO_NORM"] = inv_s["PRODUCTO_CAT"].apply(normalize_text)
    inv_s = inv_s[["ARTICULO_NORM", "CÓDIGO", "PRODUCTO_CAT", "CATEGORIA"]].drop_duplicates("ARTICULO_NORM")

    cat = pd.concat([inv_t, inv_s]).drop_duplicates("ARTICULO_NORM", keep="first")
    return cat


def period_days(df):
    if df.empty:
        return 1
    return max(1, (df["FECHA"].max() - df["FECHA"].min()).days + 1)


def build_demand_analysis(df, origen_label):
    if df.empty:
        return pd.DataFrame()

    days = period_days(df)
    months = max(days / 30.4375, 1)

    # Demanda atendida (consumo real entregado)
    atendido = df[df["CLASIF_ESTADO"] == "ATENDIDO"].copy()
    # Demanda total (todas las solicitudes como señal de necesidad)
    all_req = df.copy()

    def agg_by_product(sub, prefix):
        g = sub.groupby("ARTICULO_NORM", as_index=False).agg(
            CANTIDAD_TOTAL=( "CANTIDAD", "sum"),
            N_SOLICITUDES=("CANTIDAD", "count"),
            PROM_SOLICITUD=("CANTIDAD", "mean"),
            MAX_SOLICITUD=("CANTIDAD", "max"),
            MIN_SOLICITUD=("CANTIDAD", "min"),
        )
        g.columns = ["ARTICULO_NORM"] + [f"{prefix}_{c}" for c in g.columns[1:]]
        return g

    agg_all = agg_by_product(all_req, "TOTAL")
    agg_at = agg_by_product(atendido, "ATENDIDO")

    # Consumo mensual por producto
    tmp = atendido.copy()
    tmp["MES"] = tmp["FECHA"].dt.to_period("M")
    monthly = (
        tmp.groupby(["ARTICULO_NORM", "MES"])["CANTIDAD"]
        .sum()
        .reset_index()
        .groupby("ARTICULO_NORM")
        .agg(
            PROM_MENSUAL=("CANTIDAD", "mean"),
            MAX_MENSUAL=("CANTIDAD", "max"),
            MIN_MENSUAL=("CANTIDAD", "min"),
            STD_MENSUAL=("CANTIDAD", "std"),
            MESES_CON_DEMANDA=("CANTIDAD", "count"),
        )
        .reset_index()
    )

    result = agg_all.merge(agg_at, on="ARTICULO_NORM", how="left")
    result = result.merge(monthly, on="ARTICULO_NORM", how="left")

    # Nombre display (más frecuente)
    names = (
        df.groupby("ARTICULO_NORM")["ARTICULO"]
        .agg(lambda s: s.mode().iloc[0] if len(s.mode()) else s.iloc[0])
        .reset_index()
        .rename(columns={"ARTICULO": "ARTICULO_DISPLAY"})
    )
    result = names.merge(result, on="ARTICULO_NORM", how="right")

    result["ORIGEN"] = origen_label
    result["DIAS_HISTORICO"] = days
    result["MESES_HISTORICO"] = round(months, 2)

    # Tasas de consumo (basado en demanda ATENDIDA)
    result["PROM_MENSUAL"] = result["PROM_MENSUAL"].fillna(result["ATENDIDO_CANTIDAD_TOTAL"] / months)
    result["PROM_DIARIO"] = result["PROM_MENSUAL"] / 30.4375
    result["PROM_SEMANAL"] = result["PROM_DIARIO"] * 7
    result["PROM_QUINCENAL"] = result["PROM_DIARIO"] * 15
    result["PROM_MENSUAL_CALC"] = result["PROM_DIARIO"] * 30.4375

    # Min / Max sugeridos (und)
    min_mensual = result["MIN_MENSUAL"].fillna(0) if "MIN_MENSUAL" in result.columns else 0
    result["MIN_SUGERIDO"] = np.ceil(
        np.maximum(result["PROM_SEMANAL"].fillna(0), min_mensual / 4) * SAFETY_FACTOR
    )
    result["MAX_SUGERIDO"] = np.ceil(
        np.maximum(
            result["MAX_MENSUAL"].fillna(result["PROM_MENSUAL"]),
            result["PROM_MENSUAL"].fillna(0) * 1.5,
        )
        * SAFETY_FACTOR
    )
    result["MIN_SUGERIDO"] = result[["MIN_SUGERIDO"]].fillna(0).clip(lower=1)
    result["MAX_SUGERIDO"] = result["MAX_SUGERIDO"].fillna(result["MIN_SUGERIDO"])

    # Pedidos recomendados
    for days_p, col in [(7, "PEDIDO_7_DIAS"), (15, "PEDIDO_15_DIAS"), (30, "PEDIDO_30_DIAS")]:
        result[col] = np.ceil(result["PROM_DIARIO"].fillna(0) * days_p * SAFETY_FACTOR)
        result[col] = result[[col]].fillna(0).clip(lower=1)

    for col in ["MIN_SUGERIDO", "MAX_SUGERIDO", "PEDIDO_7_DIAS", "PEDIDO_15_DIAS", "PEDIDO_30_DIAS"]:
        result[col] = result[col].round(0).astype("Int64")

    result = result.sort_values("TOTAL_CANTIDAD_TOTAL", ascending=False)
    return result


def build_pedido_sheet(demand_df, dias, titulo_dias):
    if demand_df.empty:
        return pd.DataFrame()
    col = f"PEDIDO_{dias}_DIAS" if dias != 30 else "PEDIDO_30_DIAS"
    out = demand_df[
        [
            "ORIGEN",
            "CÓDIGO",
            "ARTICULO_DISPLAY",
            "CATEGORIA",
            col,
            "PROM_DIARIO",
            "PROM_MENSUAL",
            "MIN_SUGERIDO",
            "MAX_SUGERIDO",
            "TOTAL_N_SOLICITUDES",
            "ATENDIDO_CANTIDAD_TOTAL",
        ]
    ].copy()
    out = out.rename(columns={col: f"CANTIDAD_A_PEDIR_{titulo_dias}"})
    out = out[out[f"CANTIDAD_A_PEDIR_{titulo_dias}"] > 0]
    out["UNIDAD"] = "UND"
    out["PRIORIDAD"] = pd.cut(
        out["TOTAL_N_SOLICITUDES"].fillna(0),
        bins=[0, 2, 5, 1000],
        labels=["BAJA", "MEDIA", "ALTA"],
    )
    return out


def merge_catalog(demand_df, catalogo):
    if demand_df.empty:
        return demand_df
    m = demand_df.merge(catalogo, on="ARTICULO_NORM", how="left")
    m["ARTICULO_DISPLAY"] = m["ARTICULO_DISPLAY"].fillna(m["PRODUCTO_CAT"]).fillna(m["ARTICULO_NORM"])
    m["CÓDIGO"] = m["CÓDIGO"].fillna("")
    m["CATEGORIA"] = m["CATEGORIA"].fillna("SIN CATEGORÍA")
    return m


def build_resumen(taller, tiendas, unificado):
    def kpis(df, name):
        if df.empty:
            return {"Área": name}
        return {
            "Área": name,
            "Total solicitudes": len(df),
            "Atendidas": (df["CLASIF_ESTADO"] == "ATENDIDO").sum(),
            "No atendidas": (df["CLASIF_ESTADO"] == "NO ATENDIDO").sum(),
            "Otros estados": (df["CLASIF_ESTADO"].isin(["OTRO", "SIN ESTADO"])).sum(),
            "% Atención": round((df["CLASIF_ESTADO"] == "ATENDIDO").mean() * 100, 1),
            "Productos distintos": df["ARTICULO_NORM"].nunique(),
            "Unidades solicitadas": int(df["CANTIDAD"].sum()),
            "Unidades atendidas": int(df.loc[df["CLASIF_ESTADO"] == "ATENDIDO", "CANTIDAD"].sum()),
            "Fecha inicio": df["FECHA"].min().strftime("%Y-%m-%d"),
            "Fecha fin": df["FECHA"].max().strftime("%Y-%m-%d"),
            "Días de historial": period_days(df),
        }

    rows = [kpis(taller, "TALLER"), kpis(tiendas, "TIENDAS"), kpis(unificado, "UNIFICADO")]
    return pd.DataFrame(rows)


def build_instrucciones():
    lines = [
        ["ANÁLISIS DE CONSUMIBLES — GUÍA DE USO"],
        [""],
        ["Fecha de generación", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Archivos fuente", "Solicitudes (TALLER) consumibles.xlsx | Solicitudes (TIENDAS) consumibles.xlsx"],
        [""],
        ["1. OBJETIVO"],
        ["Este libro consolida el historial de solicitudes de consumibles para justificar cantidades de pedido."],
        ["Separa TALLER, TIENDAS y vista UNIFICADA, con tres frecuencias: semanal (7 días), quincenal (15 días) y mensual (30 días)."],
        [""],
        ["2. METODOLOGÍA"],
        ["• Demanda base: solicitudes ATENDIDAS (ENTREGADO/RECIBIDO/ENVIADO/PROCESO) — consumo real."],
        ["• Se calcula consumo diario = promedio mensual ÷ 30,44 días."],
        ["• Factor de seguridad: 15% sobre el promedio (colchón por variabilidad)."],
        ["• MIN sugerido: ~1 semana de consumo con factor de seguridad (mínimo 1 und)."],
        ["• MAX sugerido: máximo mensual histórico o 150% del promedio mensual (el mayor), × factor seguridad."],
        ["• Pedido 7 días = consumo diario × 7 × 1,15"],
        ["• Pedido 15 días = consumo diario × 15 × 1,15"],
        ["• Pedido 30 días = consumo diario × 30,44 × 1,15"],
        [""],
        ["3. HOJAS DEL ARCHIVO"],
        ["RESUMEN EJECUTIVO", "KPIs globales por área"],
        ["SOL. PROCESADAS / NO ATENDIDAS", "Detalle filtrado por estado"],
        ["ANÁLISIS DEMANDA", "Estadísticas por producto (TALLER | TIENDAS | UNIFICADO)"],
        ["PEDIDO RECOMENDADO", "Listas listas para compras por frecuencia"],
        ["DETALLE SOLICITUDES", "Base limpia completa"],
        [""],
        ["4. RECOMENDACIÓN DE USO"],
        ["• Pedido mensual rutinario: usar hoja PEDIDO RECOMENDADO MENSUAL (UNIFICADO o por área)."],
        ["• Revisar SOL. NO ATENDIDAS antes de pedir — incluir esas unidades si aún aplican."],
        ["• Ajustar MIN/MAX manualmente si hay lead time largo o estacionalidad (ej. bolsas navideñas)."],
        ["• PRIORIDAD en pedidos: ALTA = muchas solicitudes recurrentes; BAJA = demanda esporádica."],
        [""],
        ["5. NOTAS SOBRE DATOS"],
        ["• Fechas fuera de rango 2024-2027 fueron excluidas (errores de captura)."],
        ["• Cantidades se extrajeron del campo numérico; registros sin cantidad se excluyeron del cálculo."],
        ["• Productos se normalizaron (mayúsculas, sin tildes) para agrupar variantes de nombre."],
    ]
    return pd.DataFrame(lines, columns=["Concepto", "Detalle"])


def style_workbook(writer):
    wb = writer.book
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2" if ws.title != "INSTRUCCIONES" else None
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 1)):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 12
            for row in range(1, min(ws.max_row + 1, 200)):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    max_len = max(max_len, min(len(str(val)), 50))
            ws.column_dimensions[letter].width = max_len + 2


def main():
    print("Cargando datos...")
    taller = load_taller()
    tiendas = load_tiendas()
    unificado = pd.concat([taller, tiendas], ignore_index=True)
    catalogo = load_catalogo()

    # Filtrar registros con cantidad válida para análisis numérico
    taller_num = taller[taller["CANTIDAD"].notna() & (taller["CANTIDAD"] > 0)]
    tiendas_num = tiendas[tiendas["CANTIDAD"].notna() & (tiendas["CANTIDAD"] > 0)]
    unificado_num = pd.concat([taller_num, tiendas_num], ignore_index=True)

    print(f"TALLER: {len(taller_num)} | TIENDAS: {len(tiendas_num)} | UNIFICADO: {len(unificado_num)}")

    demand_taller = merge_catalog(build_demand_analysis(taller_num, "TALLER"), catalogo)
    demand_tiendas = merge_catalog(build_demand_analysis(tiendas_num, "TIENDAS"), catalogo)
    demand_uni = merge_catalog(build_demand_analysis(unificado_num, "UNIFICADO"), catalogo)

    proc_taller = taller[taller["CLASIF_ESTADO"] == "ATENDIDO"].sort_values("FECHA", ascending=False)
    no_taller = taller[taller["CLASIF_ESTADO"] == "NO ATENDIDO"].sort_values("FECHA", ascending=False)
    proc_tiendas = tiendas[tiendas["CLASIF_ESTADO"] == "ATENDIDO"].sort_values("FECHA", ascending=False)
    no_tiendas = tiendas[tiendas["CLASIF_ESTADO"] == "NO ATENDIDO"].sort_values("FECHA", ascending=False)

    resumen = build_resumen(taller_num, tiendas_num, unificado_num)
    instrucciones = build_instrucciones()

    def export_demand(df):
        cols = [
            "ORIGEN",
            "CÓDIGO",
            "ARTICULO_DISPLAY",
            "CATEGORIA",
            "DIAS_HISTORICO",
            "MESES_HISTORICO",
            "TOTAL_N_SOLICITUDES",
            "TOTAL_CANTIDAD_TOTAL",
            "ATENDIDO_N_SOLICITUDES",
            "ATENDIDO_CANTIDAD_TOTAL",
            "PROM_SOLICITUD",
            "MAX_SOLICITUD",
            "PROM_MENSUAL",
            "MAX_MENSUAL",
            "MIN_MENSUAL",
            "STD_MENSUAL",
            "PROM_DIARIO",
            "PROM_SEMANAL",
            "PROM_QUINCENAL",
            "MIN_SUGERIDO",
            "MAX_SUGERIDO",
            "PEDIDO_7_DIAS",
            "PEDIDO_15_DIAS",
            "PEDIDO_30_DIAS",
        ]
        existing = [c for c in cols if c in df.columns]
        out = df[existing].copy()
        rename = {
            "ARTICULO_DISPLAY": "ARTÍCULO",
            "TOTAL_N_SOLICITUDES": "N° SOLICITUDES (TOTAL)",
            "TOTAL_CANTIDAD_TOTAL": "UND SOLICITADAS (TOTAL)",
            "ATENDIDO_N_SOLICITUDES": "N° ATENDIDAS",
            "ATENDIDO_CANTIDAD_TOTAL": "UND ATENDIDAS",
            "PROM_SOLICITUD": "PROM POR SOLICITUD",
            "MAX_SOLICITUD": "MÁX EN 1 SOLICITUD",
            "PROM_MENSUAL": "CONSUMO PROM MENSUAL",
            "MAX_MENSUAL": "CONSUMO MÁX MENSUAL",
            "MIN_MENSUAL": "CONSUMO MÍN MENSUAL",
            "STD_MENSUAL": "DESV. EST. MENSUAL",
            "PROM_DIARIO": "CONSUMO PROM DIARIO",
            "PROM_SEMANAL": "CONSUMO PROM SEMANAL",
            "PROM_QUINCENAL": "CONSUMO PROM 15 DÍAS",
            "MIN_SUGERIDO": "STOCK MÍN (UND)",
            "MAX_SUGERIDO": "STOCK MÁX (UND)",
            "PEDIDO_7_DIAS": "PEDIDO SEMANAL (UND)",
            "PEDIDO_15_DIAS": "PEDIDO QUINCENAL (UND)",
            "PEDIDO_30_DIAS": "PEDIDO MENSUAL (UND)",
        }
        return out.rename(columns=rename)

    def export_solicitudes(df):
        return df[
            [
                "FECHA",
                "ORIGEN",
                "SUCURSAL",
                "TIENDA_HOJA",
                "SOLICITANTE",
                "ARTICULO",
                "CANTIDAD",
                "ESTADO_RAW",
                "CLASIF_ESTADO",
                "NOTAS",
            ]
        ].sort_values("FECHA", ascending=False)

    # Pedidos unificados por frecuencia (incluye desglose taller/tiendas)
    pedido_7_t = build_pedido_sheet(demand_taller, 7, "7_DIAS")
    pedido_7_s = build_pedido_sheet(demand_tiendas, 7, "7_DIAS")
    pedido_7_u = build_pedido_sheet(demand_uni, 7, "7_DIAS")

    pedido_15_t = build_pedido_sheet(demand_taller, 15, "15_DIAS")
    pedido_15_s = build_pedido_sheet(demand_tiendas, 15, "15_DIAS")
    pedido_15_u = build_pedido_sheet(demand_uni, 15, "15_DIAS")

    pedido_30_t = build_pedido_sheet(demand_taller, 30, "30_DIAS")
    pedido_30_s = build_pedido_sheet(demand_tiendas, 30, "30_DIAS")
    pedido_30_u = build_pedido_sheet(demand_uni, 30, "30_DIAS")

    print("Generando Excel...")
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        instrucciones.to_excel(writer, sheet_name="INSTRUCCIONES", index=False)
        resumen.to_excel(writer, sheet_name="RESUMEN EJECUTIVO", index=False)

        export_solicitudes(proc_taller).to_excel(writer, sheet_name="SOL PROCESADAS TALLER", index=False)
        export_solicitudes(no_taller).to_excel(writer, sheet_name="SOL NO ATENDIDAS TALLER", index=False)
        export_solicitudes(proc_tiendas).to_excel(writer, sheet_name="SOL PROCESADAS TIENDAS", index=False)
        export_solicitudes(no_tiendas).to_excel(writer, sheet_name="SOL NO ATENDIDAS TIENDAS", index=False)

        export_demand(demand_taller).to_excel(writer, sheet_name="ANÁLISIS DEMANDA TALLER", index=False)
        export_demand(demand_tiendas).to_excel(writer, sheet_name="ANÁLISIS DEMANDA TIENDAS", index=False)
        export_demand(demand_uni).to_excel(writer, sheet_name="ANÁLISIS DEMANDA UNIFICADO", index=False)

        pedido_7_u.to_excel(writer, sheet_name="PEDIDO SEMANAL (7 DÍAS)", index=False)
        pedido_15_u.to_excel(writer, sheet_name="PEDIDO QUINCENAL (15 DÍAS)", index=False)
        pedido_30_u.to_excel(writer, sheet_name="PEDIDO MENSUAL (30 DÍAS)", index=False)

        # Desglose por área en hojas adicionales
        pd.concat([pedido_7_t, pedido_7_s], ignore_index=True).to_excel(
            writer, sheet_name="PEDIDO 7D POR ÁREA", index=False
        )
        pd.concat([pedido_15_t, pedido_15_s], ignore_index=True).to_excel(
            writer, sheet_name="PEDIDO 15D POR ÁREA", index=False
        )
        pd.concat([pedido_30_t, pedido_30_s], ignore_index=True).to_excel(
            writer, sheet_name="PEDIDO 30D POR ÁREA", index=False
        )

        export_solicitudes(taller_num).to_excel(writer, sheet_name="DETALLE TALLER", index=False)
        export_solicitudes(tiendas_num).to_excel(writer, sheet_name="DETALLE TIENDAS", index=False)

        style_workbook(writer)

    print(f"Archivo generado: {OUTPUT_FILE}")
    return OUTPUT_FILE


if __name__ == "__main__":
    main()
