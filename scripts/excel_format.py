"""Formateo visual para analisis_microfibra_jabon.xlsx."""

from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2E75B6")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
ANSWER_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="B4C6E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

FILLS = {
    "ok": PatternFill("solid", fgColor="C6EFCE"),
    "warn": PatternFill("solid", fgColor="FFEB9C"),
    "crit": PatternFill("solid", fgColor="FFC7CE"),
    "info": PatternFill("solid", fgColor="DDEBF7"),
    "pedido": PatternFill("solid", fgColor="E2EFDA"),
    "abc_a": PatternFill("solid", fgColor="C6EFCE"),
    "abc_b": PatternFill("solid", fgColor="FFEB9C"),
    "abc_c": PatternFill("solid", fgColor="EDEDED"),
    "top5_yes": PatternFill("solid", fgColor="C6EFCE"),
    "top5_no": PatternFill("solid", fgColor="FFFFFF"),
}


def _auto_width(ws, max_width=48):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def _style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _add_table(ws, name: str, ref: str):
    tab = Table(displayName=name[:255], ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def _format_data_sheet(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    _style_header_row(ws, 1)
    ws.freeze_panes = "A2"
    if ws.max_row >= 2 and ws.max_column >= 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        safe = sheet_name.replace(" ", "").replace(".", "")[:20]
        try:
            _add_table(ws, f"T_{safe}", ref)
        except ValueError:
            pass
    _auto_width(ws)


def format_workbook(path: str) -> None:
    wb = load_workbook(path)

    # --- Hoja 0: Resumen Ejecutivo ---
    if "0. Resumen Ejecutivo" in wb.sheetnames:
        ws = wb["0. Resumen Ejecutivo"]
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 78
        # Fila 1 = encabezados Concepto / Valor
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
        ws.freeze_panes = "A2"
        for row in range(2, ws.max_row + 1):
            label = str(ws[f"A{row}"].value or "")
            ws[f"A{row}"].font = Font(bold=True, color="1F4E79")
            ws[f"B{row}"].alignment = LEFT
            if "RESPUESTA LOGÍSTICA" in label or "PLANIFICACIÓN TELA" in label:
                ws[f"A{row}"].font = TITLE_FONT
                ws[f"A{row}"].fill = SECTION_FILL
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            if label in {"Pregunta", "Respuesta", "Por qué estos 5"}:
                ws[f"A{row}"].fill = PatternFill("solid", fgColor="F2F2F2")
            if label == "Respuesta":
                ws[f"B{row}"].fill = ANSWER_FILL
                ws[f"B{row}"].font = Font(bold=True, size=12, color="006100")
            if "Pedido tela" in label or "Crítico operativo" in label:
                ws[f"B{row}"].fill = FILLS["pedido"]
                ws[f"B{row}"].font = Font(bold=True, color="006100")
            if label.startswith("Nota"):
                ws[f"B{row}"].fill = FILLS["warn"]

    # --- Hoja 1: Respuesta Logística (narrativa) ---
    if "1. Respuesta Logística" in wb.sheetnames:
        ws = wb["1. Respuesta Logística"]
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 72
        for row in range(1, ws.max_row + 1):
            val = str(ws.cell(row=row, column=1).value or "")
            if val.startswith("PREGUNTA") or val.startswith("RESPUESTA") or val.startswith("CÓMO") or val.startswith("QUÉ SIGNIFICAN"):
                ws.cell(row=row, column=1).font = SUBTITLE_FONT
                ws.cell(row=row, column=1).fill = SECTION_FILL
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            if val == "Los 5 colores recomendados":
                ws.cell(row=row, column=2).fill = ANSWER_FILL
                ws.cell(row=row, column=2).font = Font(bold=True, size=12, color="006100")
            if val == "En una frase":
                ws.cell(row=row, column=2).alignment = LEFT
        # Header tabla embebida Top 5
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "TOP 5 — DETALLE POR COLOR":
                ws.cell(row=row, column=1).font = SUBTITLE_FONT
                ws.cell(row=row, column=1).fill = SECTION_FILL
            if ws.cell(row=row, column=1).value == "#":
                _style_header_row(ws, row)
                break
        _auto_width(ws, max_width=52)

    logistics_sheets = [
        "2. Top 5 Detalle",
        "3. Ranking Colores",
        "4. Casos Fuera del Top 5",
        "5. Top 5 Logística",
        "6. Colores con Sobrestock",
    ]
    for sheet_name in logistics_sheets:
        _format_data_sheet(wb, sheet_name)

    # Top 5 highlight in ranking
    if "3. Ranking Colores" in wb.sheetnames:
        ws = wb["3. Ranking Colores"]
        top5_col = next((c.column for c in ws[1] if c.value == "¿Top 5?"), None)
        if top5_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=top5_col)
                if cell.value == "SÍ":
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=c).fill = FILLS["top5_yes"]
        score_col = next((c.column for c in ws[1] if c.value == "SCORE TOTAL"), None)
        if score_col:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=top5_col or 0).value == "SÍ":
                    ws.cell(row=row, column=score_col).font = Font(bold=True, color="1F4E79")

    data_sheets = [
        "7. Resumen por Color",
        "8. Riesgo y Reorden",
        "9. Pedido Tela",
        "10. Resumen Modelos",
        "15. Semáforo Integrado",
        "16. Verificación y Notas",
        # compatibilidad nombres viejos
        "1. Resumen por Color",
        "2. Top 5 Logística",
        "3. Riesgo y Reorden",
        "4. Pedido Tela",
        "10. Semáforo Integrado",
        "11. Verificación y Notas",
    ]
    for sheet_name in data_sheets:
        _format_data_sheet(wb, sheet_name)

    # ABC badges
    for resumen_name in ("7. Resumen por Color", "1. Resumen por Color"):
        if resumen_name not in wb.sheetnames:
            continue
        ws = wb[resumen_name]
        abc_col = next((c.column for c in ws[1] if c.value == "ABC"), None)
        if abc_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=abc_col)
                val = str(cell.value or "")
                cell.fill = FILLS.get(f"abc_{val.lower()}", FILLS["info"])
                cell.font = Font(bold=True)
                cell.alignment = CENTER

    # Riesgo colors
    for sheet_name, col_name in [
        ("7. Resumen por Color", "Riesgo"),
        ("1. Resumen por Color", "Riesgo"),
        ("8. Riesgo y Reorden", "Riesgo"),
        ("3. Riesgo y Reorden", "Riesgo"),
    ]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        col_idx = next((c.column for c in ws[1] if c.value == col_name), None)
        if not col_idx:
            continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = str(cell.value or "").upper()
            if any(x in val for x in ("CRÍTICA", "SIN TELA", "QUIEBRE")):
                cell.fill = FILLS["crit"]
            elif "SOBRE" in val:
                cell.fill = FILLS["warn"]
            elif "SALUDABLE" in val:
                cell.fill = FILLS["ok"]
            cell.alignment = LEFT

    # Pedido highlight
    for ped_name in ("9. Pedido Tela", "4. Pedido Tela"):
        if ped_name not in wb.sheetnames:
            continue
        ws = wb[ped_name]
        ped_col = next((c.column for c in ws[1] if c.value == "Pedido sugerido (kg)"), None)
        if ped_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=ped_col)
                if isinstance(cell.value, (int, float)) and cell.value > 0:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=c).fill = FILLS["pedido"]
                    cell.font = Font(bold=True, color="006100", size=12)

    # Semáforo
    for sem_name in ("15. Semáforo Integrado", "10. Semáforo Integrado"):
        if sem_name not in wb.sheetnames:
            continue
        ws = wb[sem_name]
        sem_map = {
            "OK": FILLS["ok"], "SALUDABLE": FILLS["ok"], "REGULAR": FILLS["ok"], "ALTA": FILLS["ok"],
            "CRÍTICO": FILLS["crit"], "SIN TELA": FILLS["crit"], "QUIEBRE": FILLS["crit"],
            "SOBRESTOCK": FILLS["warn"], "BAJO": FILLS["warn"], "VARIABLE": FILLS["warn"], "MEDIA": FILLS["warn"],
        }
        for row in range(2, ws.max_row + 1):
            for col_name in ["Rotación", "Regularidad", "Riesgo PT", "Riesgo Tela"]:
                col_idx = next((c.column for c in ws[1] if c.value == col_name), None)
                if not col_idx:
                    continue
                cell = ws.cell(row=row, column=col_idx)
                val = str(cell.value or "").upper()
                for key, fill in sem_map.items():
                    if key in val:
                        cell.fill = fill
                        cell.alignment = CENTER
                        break

    # Top 5 score column
    for top_name in ("5. Top 5 Logística", "2. Top 5 Logística"):
        if top_name not in wb.sheetnames:
            continue
        ws = wb[top_name]
        score_col = next((c.column for c in ws[1] if c.value == "SCORE TOTAL"), None)
        if score_col:
            for row in range(2, min(7, ws.max_row + 1)):
                ws.cell(row=row, column=score_col).fill = FILLS["info"]
                ws.cell(row=row, column=score_col).font = Font(bold=True, size=12, color="1F4E79")

    wb.save(path)
