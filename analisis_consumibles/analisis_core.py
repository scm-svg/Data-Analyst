# -*- coding: utf-8 -*-
"""
Núcleo de análisis: carga, limpieza y cálculos sobre las solicitudes de
consumibles de TALLER y TIENDAS (historial de 2 meses: 08-jun al 05-ago-2026).
"""
import math
import datetime as dt
import unicodedata
from collections import Counter

import openpyxl
import pandas as pd

UPLOAD_DIR = "/home/ubuntu/.cursor/projects/workspace/uploads"
F_TALLER = f"{UPLOAD_DIR}/solicitudes_taller_2_meses_45c7.xlsx"
F_TIENDAS = f"{UPLOAD_DIR}/_Solicitudes__TIENDAS__consumibles_2meses_8eab.xlsx"

HOJAS_TIENDAS = ["GRIETA", "SAMBIL VALENCIA", "GRAND PLAZ", "CERRO VERDE",
                 "SAMBIL CHACAO", "TOLON", "MARGARITA"]

# Clasificación de estados según lo pedido:
#   ATENDIDA  = recibido/entregado/enviado + solicitado
#   NO ATENDIDA = no disponible
ESTADO_MAP = {
    "ENTRGADO": "ENTREGADO",      # typo en el archivo origen
    "ENTREGADO": "ENTREGADO",
    "RECIBIDO": "RECIBIDO",
    "ENVIADO": "ENVIADO",
    "SOLICITADO": "SOLICITADO",
    "NO DISPONIBLE": "NO DISPONIBLE",
}
ESTADOS_ATENDIDOS = {"ENTREGADO", "RECIBIDO", "ENVIADO", "SOLICITADO", "SIN ESTADO"}

Z_SERVICIO = 1.65          # nivel de servicio 95% (clase X)
Z_X, Z_Y, Z_Z = 1.65, 1.28, 0.84   # servicio por clase de demanda: X 95%, Y 90%, Z 80%
LT_DEFAULT = 7             # lead time por defecto (días) - editable por producto
P_SEMANAL, P_QUINCENAL, P_MENSUAL = 7, 14, 30


def _norm_txt(s):
    return " ".join(str(s).strip().split()) if s is not None else None


def _norm_key(s):
    """Mayúsculas sin acentos para cruzar con el catálogo."""
    return "".join(c for c in unicodedata.normalize("NFD", s.upper().strip())
                   if unicodedata.category(c) != "Mn")


def _leer(path, hojas, fila_hdr, origen):
    filas = []
    wb = openpyxl.load_workbook(path, data_only=True)
    for sn in hojas:
        ws = wb[sn]
        hdr = [c.value for c in ws[fila_hdr]]
        for i, r in enumerate(ws.iter_rows(min_row=fila_hdr + 1, values_only=True),
                              start=fila_hdr + 1):
            if all(v is None for v in r):
                continue
            d = dict(zip(hdr[:len(r)], r))
            d["_hoja"] = sn
            d["_fila"] = i
            d["_origen"] = origen
            filas.append(d)
    return filas


def cargar_catalogo():
    wb = openpyxl.load_workbook(F_TIENDAS, data_only=True)
    ws = wb["INVENTARIO DE CONSUMIBLES"]
    cat = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1]:
            cat[_norm_key(str(r[1]))] = {
                "id": str(r[0]).strip() if r[0] else "",
                "producto": str(r[1]).strip(),
                "categoria": str(r[2]).strip() if r[2] else "SIN CATEGORÍA",
            }
    return cat


def limpiar():
    """Devuelve (df, excluidas, observaciones)."""
    crudas = (_leer(F_TALLER, ["CRECO SOLICITUDES"], 2, "TALLER")
              + _leer(F_TIENDAS, HOJAS_TIENDAS, 3, "TIENDAS"))
    catalogo = cargar_catalogo()

    excluidas, obs = [], []
    filas, sin_estado, typo_estado = [], 0, 0

    for r in crudas:
        ref = f"{r['_origen']} / hoja '{r['_hoja']}' / fila {r['_fila']}"
        art = _norm_txt(r.get("ARTICULO"))
        car = _norm_txt(r.get("CARACTER ADICIONAL"))
        cant = r.get("CANTIDAD")
        est_raw = _norm_txt(r.get("ESTADO"))
        fecha = r.get("FECHA")

        # --- filas basura: sin fecha o sin cantidad no son solicitudes válidas ---
        if fecha is None or cant is None:
            motivo = ("Fila vacía / basura (solo contiene '.' o nombre)"
                      if fecha is None and cant is None else
                      "Fila incompleta: sin FECHA" if fecha is None else
                      "Fila incompleta: sin CANTIDAD ni artículo (solo fecha y nombre)")
            excluidas.append([ref, r["_origen"], r["_hoja"], str(fecha), str(r.get("NOMBRE")),
                              str(art), str(car), str(cant), str(est_raw), motivo])
            continue

        # --- artículo: si ARTICULO está vacío se usa CARACTER ADICIONAL ---
        nota_art = ""
        if art is None or art == ".":
            if car and car != ".":
                art, nota_art = car, "Artículo tomado de CARACTER ADICIONAL"
            else:
                art, nota_art = "(SIN ARTÍCULO)", "Sin artículo: solo cuenta para KPIs generales"

        # --- estado ---
        if est_raw is None:
            estado, sin_estado = "SIN ESTADO", sin_estado + 1
        else:
            key = est_raw.upper()
            if key == "ENTRGADO":
                typo_estado += 1
            estado = ESTADO_MAP.get(key, key)

        filas.append({
            "FECHA": fecha.date() if isinstance(fecha, dt.datetime) else fecha,
            "ORIGEN": r["_origen"],
            "SUCURSAL": _norm_txt(r.get("SUCURSAL")) or r["_hoja"],
            "NOMBRE": _norm_txt(r.get("NOMBRE")),
            "ARTICULO": art.upper() if art != "(SIN ARTÍCULO)" else art,
            "CARACTER ADICIONAL": car if car not in (None, ".") else None,
            "CANTIDAD": int(cant) if isinstance(cant, (int, float)) else None,
            "ESTADO ORIGINAL": est_raw,
            "ESTADO": estado,
            "ATENDIDA": estado in ESTADOS_ATENDIDOS,
            "NOTA LIMPIEZA": nota_art,
        })

    df = pd.DataFrame(filas)
    df["SEMANA"] = df["FECHA"].apply(lambda d: d - dt.timedelta(days=d.weekday()))
    df["MES"] = df["FECHA"].apply(lambda d: dt.date(d.year, d.month, 1))
    df["CON_ARTICULO"] = df["ARTICULO"] != "(SIN ARTÍCULO)"

    cat_info = df["ARTICULO"].map(lambda a: catalogo.get(_norm_key(a)))
    df["CATEGORIA"] = [c["categoria"] if c else "SIN CATEGORÍA (NO REGISTRADO)"
                       for c in cat_info]

    obs.append(f"'ENTRGADO' (typo del origen) normalizado a 'ENTREGADO': {typo_estado} filas.")
    obs.append(f"Filas sin ESTADO registrado: {sin_estado} (SAMBIL VALENCIA, 04-ago-2026). "
               "Se tratan como SOLICITADO para el % de atención y la demanda.")
    obs.append("Artículo 'REGLETA' no existe en el catálogo de consumibles: "
               "fue solicitado vía CARACTER ADICIONAL. Categoría asignada: SIN CATEGORÍA (NO REGISTRADO).")
    obs.append("Las cantidades se mantienen en la unidad registrada en cada solicitud "
               "(und, paquete, galón, caja… según CARACTER ADICIONAL).")
    obs.append("No se encontraron duplicados reales: las repeticiones del mismo día/artículo "
               "corresponden a sucursales distintas.")
    return df, excluidas, obs


# ----------------------------------------------------------------------------
# CÁLCULOS
# ----------------------------------------------------------------------------
def _rango_dias(df):
    return (df["FECHA"].max() - df["FECHA"].min()).days + 1


def tabla_productos(df, dias=None):
    """Tabla por producto: demanda diaria y variabilidad SEMANAL (σ de semanas
    completas), coeficiente de variación y clase XYZ de demanda."""
    d = df[df["CON_ARTICULO"]]
    dias = dias or _rango_dias(df)
    idx = pd.date_range(d["FECHA"].min(), d["FECHA"].max(), freq="D")
    fmin = d["FECHA"].min()
    semanas_completas = dias // 7

    rows = []
    for art, g in d.groupby("ARTICULO"):
        diaria = g.groupby("FECHA")["CANTIDAD"].sum().reindex(idx.date, fill_value=0.0)
        n_sem = [(x - fmin).days // 7 for x in diaria.index]
        semanal = diaria.groupby(n_sem).sum()
        semanal = semanal.iloc[:semanas_completas]          # solo semanas completas
        mu_sem = float(semanal.mean()) if len(semanal) else 0.0
        sigma_sem = float(semanal.std(ddof=1)) if len(semanal) > 1 else 0.0
        cv = sigma_sem / mu_sem if mu_sem > 0 else 0.0
        unds = int(g["CANTIDAD"].sum())
        atend = int(g.loc[g["ATENDIDA"], "CANTIDAD"].sum())
        solicitudes = len(g)
        if (solicitudes <= 2 and unds <= 4) or cv > 1.0:
            clase = "Z · Intermitente"
        elif cv > 0.5:
            clase = "Y · Variable"
        else:
            clase = "X · Regular"
        rows.append({
            "ARTICULO": art,
            "CATEGORIA": g["CATEGORIA"].iloc[0],
            "SOLICITUDES": solicitudes,
            "UNDS": unds,
            "UNDS_ATEND": atend,
            "UNDS_NODISP": unds - atend,
            "PCT_ATENCION": atend / unds if unds else 0,
            "D_DIARIA": unds / dias,
            "SIGMA_SEM": sigma_sem,
            "CV": cv,
            "CLASE": clase,
        })
    t = pd.DataFrame(rows).sort_values("UNDS", ascending=False).reset_index(drop=True)
    return t, dias


def excel_round(x, nd=0):
    f = 10 ** nd
    return math.floor(x * f + 0.5) / f if x >= 0 else -math.floor(-x * f + 0.5) / f


def modelo(tabla, lt=LT_DEFAULT):
    """Agrega SS, MIN y MAX/PEDIDO para las 3 opciones de frecuencia.
    SS = z(clase) × σ_semanal × √(LT/7): cubre la variabilidad durante el lead time."""
    zmap = {"X · Regular": Z_X, "Y · Variable": Z_Y, "Z · Intermitente": Z_Z}
    t = tabla.copy()
    t["Z_VAL"] = t["CLASE"].map(zmap)
    t["SS"] = [excel_round(z * s * math.sqrt(lt / 7), 2)
               for z, s in zip(t["Z_VAL"], t["SIGMA_SEM"])]
    t["MIN"] = [math.ceil(d * lt + ss) for d, ss in zip(t["D_DIARIA"], t["SS"])]
    for nombre, p in [("SEM", P_SEMANAL), ("QUINC", P_QUINCENAL), ("MENS", P_MENSUAL)]:
        t[f"MAX_{nombre}"] = [math.ceil(d * (lt + p) + ss)
                              for d, ss in zip(t["D_DIARIA"], t["SS"])]
        t[f"PEDIDO_{nombre}"] = t[f"MAX_{nombre}"]   # stock actual = 0 (historial desde cero)
    return t


def kpis_alcance(df):
    out = {}
    for nombre, sub in [("TALLER", df[df["ORIGEN"] == "TALLER"]),
                        ("TIENDAS", df[df["ORIGEN"] == "TIENDAS"]),
                        ("UNIFICADO", df)]:
        sol, at = len(sub), int(sub["ATENDIDA"].sum())
        unds = int(sub["CANTIDAD"].sum())
        unds_at = int(sub.loc[sub["ATENDIDA"], "CANTIDAD"].sum())
        out[nombre] = {
            "solicitudes": sol, "atendidas": at, "no_disponible": sol - at,
            "pct_atendidas": at / sol, "pct_no_disponible": (sol - at) / sol,
            "unds": unds, "unds_atend": unds_at, "unds_nodisp": unds - unds_at,
            "pct_unds_atend": unds_at / unds,
        }
    return out


def tabla_semanal(df):
    rows = []
    for sem, g in df.groupby("SEMANA"):
        ta = g[g["ORIGEN"] == "TALLER"]
        ti = g[g["ORIGEN"] == "TIENDAS"]
        rows.append({
            "SEMANA": sem, "SOL_TALLER": len(ta), "UNDS_TALLER": int(ta["CANTIDAD"].sum()),
            "SOL_TIENDAS": len(ti), "UNDS_TIENDAS": int(ti["CANTIDAD"].sum()),
            "SOL_TOTAL": len(g), "UNDS_TOTAL": int(g["CANTIDAD"].sum()),
        })
    return pd.DataFrame(rows).sort_values("SEMANA").reset_index(drop=True)


def matriz(df, columna):
    d = df[df["CON_ARTICULO"]]
    m = d.pivot_table(index="ARTICULO", columns=columna, values="CANTIDAD",
                      aggfunc="sum", fill_value=0)
    m["TOTAL"] = m.sum(axis=1)
    return m.sort_values("TOTAL", ascending=False).astype(int)


def top_no_disponible(df, n=10):
    d = df[(~df["ATENDIDA"]) & df["CON_ARTICULO"]]
    g = d.groupby("ARTICULO").agg(VECES=("CANTIDAD", "size"),
                                  UNDS_NODISP=("CANTIDAD", "sum"))
    tot = df[df["CON_ARTICULO"]].groupby("ARTICULO")["CANTIDAD"].sum()
    g["UNDS_TOTAL"] = tot
    g = g.sort_values(["VECES", "UNDS_NODISP"], ascending=False).head(n).reset_index()
    g["PCT_NO_CUBIERTO"] = g["UNDS_NODISP"] / g["UNDS_TOTAL"]
    return g
