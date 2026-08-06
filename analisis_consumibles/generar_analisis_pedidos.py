#!/usr/bin/env python3
"""
Analisis de solicitudes de consumibles (Tiendas + Taller/CRECO)
Historial ~2 meses -> cantidades justificadas para pedido
Escenarios: Semanal / Quincenal / Mensual
"""

from __future__ import annotations

import math
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.chart import BarChart, Reference

# ---------------------------------------------------------------------------
# Rutas
# ---------------------------------------------------------------------------
OUT_DIR = Path(__file__).resolve().parent
TIENDAS_PATH = OUT_DIR / "fuente_tiendas_2meses.xlsx"
TALLER_SOLIC_PATH = OUT_DIR / "fuente_taller_2meses.xlsx"  # solicitudes CRECO (incompleto para Taller)
TALLER_MOV_PATH = OUT_DIR / "fuente_movimientos_taller.xlsx"  # movimientos Taller (fuente completa)
# fallback a uploads del agente si no estan las copias locales
if not TIENDAS_PATH.exists():
    TIENDAS_PATH = Path(
        "/home/ubuntu/.cursor/projects/workspace/uploads/"
        "_Solicitudes__TIENDAS__consumibles_2meses_09b9.xlsx"
    )
if not TALLER_SOLIC_PATH.exists():
    TALLER_SOLIC_PATH = Path(
        "/home/ubuntu/.cursor/projects/workspace/uploads/"
        "solicitudes_taller_2_meses_4432.xlsx"
    )
if not TALLER_MOV_PATH.exists():
    TALLER_MOV_PATH = Path(
        "/home/ubuntu/.cursor/projects/workspace/uploads/"
        "movimientos_taller_actualizado_30c9.xlsx"
    )
OUT_XLSX = OUT_DIR / "Analisis_Pedidos_Consumibles_2meses.xlsx"
ARTIFACT_XLSX = Path("/opt/cursor/artifacts/Analisis_Pedidos_Consumibles_2meses.xlsx")

# ---------------------------------------------------------------------------
# Normalizacion
# ---------------------------------------------------------------------------
PRODUCT_ALIASES = {
    "CAFE": "CAFE",
    "CAFÉ": "CAFE",
    "BOLSA DE BASURA": "BOLSAS DE BASURA",
    "BOLSAS DE BASURA": "BOLSAS DE BASURA",
    "BATERIA AAA": "BATERIAS AAA",
    "BATERIAS AAA": "BATERIAS AAA",
    "BATERIA AA": "BATERIAS AA",
    "BATERIAS AA": "BATERIAS AA",
    "CARPETA ACORDEÓN": "CARPETA ACORDEON",
    "CARPETA ACORDEON": "CARPETA ACORDEON",
    "ROLLO DE IMPRESORA FISCAL PAQ 5 UNDS": "ROLLO IMPRESORA FISCAL PAQ 5 UND",
    "ROLLO DE IMPRESORA FISCAL PAQ 5 UND": "ROLLO IMPRESORA FISCAL PAQ 5 UND",
    "CINTA TÉRMICA": "CINTA TERMICA",
    "CINTA TERMICA": "CINTA TERMICA",
    "BOLSAS PEQUEÑAS NAVIDAD 2024": "BOLSAS PEQUENAS NAVIDAD 2024",
    "BOLSAS EXTRA PEQUEÑA KRAFT": "BOLSAS EXTRA PEQUENA KRAFT",
    "BOLSAS PEQUEÑAS KRAFT": "BOLSAS PEQUENAS KRAFT",
    "PANO AMARILLO": "PANO AMARILLO",
    "PAÑO AMARILLO": "PANO AMARILLO",
    "GEL DE BAÑO": "GEL DE BANO",
    "GEL DE BANO": "GEL DE BANO",
    "ENTRGADO": "ENTREGADO",
}


def norm_text(x) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip().upper()
    s = re.sub(r"\s+", " ", s)
    for a, b in (
        ("Á", "A"),
        ("É", "E"),
        ("Í", "I"),
        ("Ó", "O"),
        ("Ú", "U"),
        ("Ü", "U"),
        ("Ñ", "N"),
    ):
        s = s.replace(a, b)
    return s.strip(" .")


def norm_product(x) -> str | None:
    s = norm_text(x)
    if not s:
        return None
    return PRODUCT_ALIASES.get(s, s)


def classify_estado(e) -> str:
    e = norm_text(e)
    if not e:
        return "SIN_ESTADO"
    if e == "NO DISPONIBLE":
        return "NO_DISPONIBLE"
    # Atendido = recibido o en ciclo de solicitud/envio (incluye SOLICITADO)
    if e in {
        "RECIBIDO",
        "SOLICITADO",
        "ENVIADO",
        "ENTRGADO",
        "ENTREGADO",
        "PROCESO",
        "EN PROCESO",
        "MOVIMIENTO SALIDA",
        "MOVIMIENTO_SALIDA",
    }:
        return "ATENDIDO"
    return e


def find_header_row(df: pd.DataFrame, max_scan: int = 6) -> int | None:
    for i in range(min(max_scan, len(df))):
        vals = [norm_text(v) for v in df.iloc[i].tolist()]
        if "FECHA" in vals and "ARTICULO" in vals:
            return i
    return None


def parse_solicitudes_sheet(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    header_idx = find_header_row(df)
    if header_idx is None:
        return pd.DataFrame()

    cols_raw = [norm_text(c) for c in df.iloc[header_idx].tolist()]
    cols = []
    for j, c in enumerate(cols_raw):
        cols.append(c if c else f"COL{j}")

    body = df.iloc[header_idx + 1 :].copy()
    body.columns = cols

    rename = {}
    for c in body.columns:
        if c == "FECHA" or c.startswith("FECHA"):
            rename[c] = "FECHA"
        elif c == "NOMBRE":
            rename[c] = "NOMBRE"
        elif "ARTICULO" in c:
            rename[c] = "ARTICULO"
        elif "CARACTER" in c:
            rename[c] = "CARACTER_ADICIONAL"
        elif "CANTIDAD" in c:
            rename[c] = "CANTIDAD"
        elif "ESTADO" in c:
            rename[c] = "ESTADO"
        elif "SUCURSAL" in c:
            rename[c] = "SUCURSAL"
        elif "NOTA" in c or "COMENT" in c:
            rename[c] = "NOTAS"
    body = body.rename(columns=rename)
    keep = [
        c
        for c in [
            "FECHA",
            "NOMBRE",
            "ARTICULO",
            "CARACTER_ADICIONAL",
            "CANTIDAD",
            "ESTADO",
            "SUCURSAL",
            "NOTAS",
        ]
        if c in body.columns
    ]
    body = body[keep].dropna(how="all")
    if "ARTICULO" not in body.columns:
        return pd.DataFrame()
    body = body[body["ARTICULO"].notna()]
    body = body[body["ARTICULO"].map(norm_text) != "ARTICULO"]
    body["HOJA_ORIGEN"] = sheet_name
    return body.reset_index(drop=True)


def ceil_pos(x: float) -> int:
    if pd.isna(x) or x <= 0:
        return 0
    return int(math.ceil(x))


# ---------------------------------------------------------------------------
# Carga
# ---------------------------------------------------------------------------
def parse_movimientos_taller(path: Path) -> pd.DataFrame:
    """Salidas de inventario de Taller (movimientos negativos -> demanda en und)."""
    raw = pd.read_excel(path)
    raw.columns = [str(c).strip() for c in raw.columns]
    # normalizar nombres de columnas
    colmap = {}
    for c in raw.columns:
        cn = norm_text(c)
        if cn == "FECHA":
            colmap[c] = "FECHA"
        elif cn in {"CODIGO", "CÓDIGO"} or "CODIGO" in cn:
            colmap[c] = "CODIGO"
        elif "MOVIMIEN" in cn:
            colmap[c] = "MOVIMIENTO"
        elif cn == "PRODUCTO":
            colmap[c] = "PRODUCTO"
        elif "CATEGOR" in cn:
            colmap[c] = "CATEGORIA_SRC"
        elif cn == "UND":
            colmap[c] = "UOM"
        elif cn == "SUCURSAL":
            colmap[c] = "SUCURSAL"
    df = raw.rename(columns=colmap)
    need = {"FECHA", "MOVIMIENTO", "PRODUCTO"}
    if not need.issubset(df.columns):
        raise ValueError(f"Movimientos taller sin columnas esperadas: {df.columns.tolist()}")

    df["FECHA"] = pd.to_datetime(df["FECHA"], errors="coerce")
    df["MOVIMIENTO"] = pd.to_numeric(df["MOVIMIENTO"], errors="coerce")
    df = df[df["FECHA"].notna() & df["MOVIMIENTO"].notna() & df["PRODUCTO"].notna()].copy()
    # Solo salidas (consumo). Valor absoluto = unidades movidas
    df = df[df["MOVIMIENTO"] < 0].copy()
    df["CANTIDAD"] = df["MOVIMIENTO"].abs()
    df["ARTICULO"] = df["PRODUCTO"]
    df["NOMBRE"] = ""
    df["CARACTER_ADICIONAL"] = df["UOM"] if "UOM" in df.columns else ""
    df["ESTADO"] = "MOVIMIENTO_SALIDA"  # consumo ejecutado en inventario
    if "SUCURSAL" not in df.columns:
        df["SUCURSAL"] = "TALLER"
    df["SUCURSAL"] = df["SUCURSAL"].fillna("TALLER")
    df["NOTAS"] = df.apply(
        lambda r: f"CODIGO={r['CODIGO']}" if "CODIGO" in df.columns and pd.notna(r.get("CODIGO")) else "",
        axis=1,
    )
    df["HOJA_ORIGEN"] = "MOVIMIENTOS TALLER"
    df["ARCHIVO"] = "MOVIMIENTOS_TALLER"
    df["GRUPO"] = "TALLER"
    df["FUENTE_TIPO"] = "MOVIMIENTO"
    if "CATEGORIA_SRC" in df.columns:
        df["CATEGORIA_SRC"] = df["CATEGORIA_SRC"].map(norm_text)
    else:
        df["CATEGORIA_SRC"] = ""
    if "UOM" in df.columns:
        df["UOM"] = df["UOM"].map(norm_text)
    else:
        df["UOM"] = "UND"
    keep = [
        "FECHA",
        "NOMBRE",
        "ARTICULO",
        "CARACTER_ADICIONAL",
        "CANTIDAD",
        "ESTADO",
        "SUCURSAL",
        "NOTAS",
        "HOJA_ORIGEN",
        "ARCHIVO",
        "GRUPO",
        "FUENTE_TIPO",
        "CATEGORIA_SRC",
        "UOM",
        "CODIGO",
    ]
    keep = [c for c in keep if c in df.columns]
    return df[keep].reset_index(drop=True)


def load_all() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Retorna (all_df principal, inventario, solicitudes_taller_incompletas_ref)."""
    xl = pd.ExcelFile(TIENDAS_PATH)
    store_sheets = [s for s in xl.sheet_names if s != "INVENTARIO DE CONSUMIBLES"]
    frames = []
    for s in store_sheets:
        raw = pd.read_excel(TIENDAS_PATH, sheet_name=s, header=None)
        p = parse_solicitudes_sheet(raw, s)
        if len(p):
            p["ARCHIVO"] = "TIENDAS"
            p["GRUPO"] = "TIENDAS"
            p["FUENTE_TIPO"] = "SOLICITUD"
            p["CATEGORIA_SRC"] = ""
            p["UOM"] = "UND"
            p["CODIGO"] = ""
            frames.append(p)

    inv = pd.read_excel(TIENDAS_PATH, sheet_name="INVENTARIO DE CONSUMIBLES")
    inv["PRODUCTO_NORM"] = inv["PRODUNTOS"].map(norm_product)
    inv["CATEGORIA"] = inv["CATEGORIA"].map(norm_text)

    # TALLER completo = movimientos actualizados (reemplaza solicitudes incompletas de Taller)
    if TALLER_MOV_PATH.exists():
        frames.append(parse_movimientos_taller(TALLER_MOV_PATH))

    # Solicitudes CRECO: conservar SOLO otras areas (I+D, Almacen, etc.).
    # Las lineas SUCURSAL=TALLER del archivo viejo se dejan fuera del principal
    # (incompletas) y se devuelven aparte como referencia de fill rate historico.
    taller_solic_ref = pd.DataFrame()
    if TALLER_SOLIC_PATH.exists():
        raw_t = pd.read_excel(TALLER_SOLIC_PATH, sheet_name=0, header=None)
        taller_solic = parse_solicitudes_sheet(raw_t, "CRECO SOLICITUDES")
        if len(taller_solic):
            taller_solic["ARCHIVO"] = "TALLER_CRECO_SOLICITUDES"
            suc = taller_solic["SUCURSAL"].map(norm_text)
            taller_solic["GRUPO"] = np.where(suc.eq("TALLER"), "TALLER", "CRECO_OTRAS_AREAS")
            taller_solic["FUENTE_TIPO"] = "SOLICITUD"
            taller_solic["CATEGORIA_SRC"] = ""
            taller_solic["UOM"] = "UND"
            taller_solic["CODIGO"] = ""
            otras = taller_solic[taller_solic["GRUPO"] == "CRECO_OTRAS_AREAS"].copy()
            if len(otras):
                frames.append(otras)
            taller_solic_ref = taller_solic[taller_solic["GRUPO"] == "TALLER"].copy()

    all_df = pd.concat(frames, ignore_index=True)
    all_df["FECHA"] = pd.to_datetime(all_df["FECHA"], errors="coerce")
    all_df["CANTIDAD"] = pd.to_numeric(all_df["CANTIDAD"], errors="coerce")
    all_df = all_df[all_df["CANTIDAD"].notna() & (all_df["CANTIDAD"] > 0)].copy()
    all_df = all_df[all_df["FECHA"].notna()].copy()

    all_df["ARTICULO_ORIG"] = all_df["ARTICULO"].astype(str).str.strip()
    all_df["ARTICULO_NORM"] = all_df["ARTICULO"].map(norm_product)
    all_df["CARACTER_NORM"] = all_df.get(
        "CARACTER_ADICIONAL", pd.Series([""] * len(all_df))
    ).map(norm_text)
    all_df["PRODUCTO_KEY"] = all_df.apply(
        lambda r: r["ARTICULO_NORM"]
        + ((" | " + r["CARACTER_NORM"]) if r["CARACTER_NORM"] else ""),
        axis=1,
    )
    all_df["ESTADO_RAW"] = all_df["ESTADO"].map(norm_text)
    all_df["ESTADO_CAT"] = all_df["ESTADO"].map(classify_estado)
    # Movimientos de salida = consumo ejecutado => cuentan como atendidos para fill rate
    mov_mask = all_df["FUENTE_TIPO"].eq("MOVIMIENTO")
    all_df.loc[mov_mask, "ESTADO_CAT"] = "ATENDIDO"
    all_df.loc[mov_mask, "ESTADO_RAW"] = "MOVIMIENTO_SALIDA"

    all_df["SUCURSAL"] = all_df["SUCURSAL"].map(norm_text)
    miss = all_df["SUCURSAL"].eq("")
    all_df.loc[miss, "SUCURSAL"] = all_df.loc[miss, "HOJA_ORIGEN"].map(norm_text)
    all_df["ATENDIDO_FLAG"] = all_df["ESTADO_CAT"].eq("ATENDIDO")
    all_df["NO_DISP_FLAG"] = all_df["ESTADO_CAT"].eq("NO_DISPONIBLE")
    all_df["CANT_ATENDIDA"] = np.where(all_df["ATENDIDO_FLAG"], all_df["CANTIDAD"], 0.0)
    all_df["CANT_NO_DISP"] = np.where(all_df["NO_DISP_FLAG"], all_df["CANTIDAD"], 0.0)

    # Categoria: preferir la del movimiento; si no, catalogo inventario
    cat_map = (
        inv.dropna(subset=["PRODUCTO_NORM"])
        .drop_duplicates("PRODUCTO_NORM")
        .set_index("PRODUCTO_NORM")["CATEGORIA"]
        .to_dict()
    )
    if "CATEGORIA_SRC" in all_df.columns:
        all_df["CATEGORIA"] = all_df["CATEGORIA_SRC"].where(
            all_df["CATEGORIA_SRC"].astype(str).str.len() > 0
        )
        all_df["CATEGORIA"] = all_df["CATEGORIA"].fillna(
            all_df["ARTICULO_NORM"].map(cat_map)
        )
    else:
        all_df["CATEGORIA"] = all_df["ARTICULO_NORM"].map(cat_map)
    all_df["CATEGORIA"] = all_df["CATEGORIA"].fillna("SIN CATEGORIA").map(norm_text)

    all_df["SEMANA"] = all_df["FECHA"].dt.to_period("W-SUN").astype(str)
    all_df["MES"] = all_df["FECHA"].dt.to_period("M").astype(str)
    all_df["DIA"] = all_df["FECHA"].dt.day
    all_df["QUINCENA"] = all_df.apply(
        lambda r: f"{r['MES']}-Q1" if r['DIA'] <= 15 else f"{r['MES']}-Q2", axis=1
    )

    # Normalizar referencia incompleta de solicitudes taller (si existe)
    if len(taller_solic_ref):
        taller_solic_ref["FECHA"] = pd.to_datetime(taller_solic_ref["FECHA"], errors="coerce")
        taller_solic_ref["CANTIDAD"] = pd.to_numeric(taller_solic_ref["CANTIDAD"], errors="coerce")
        taller_solic_ref = taller_solic_ref[
            taller_solic_ref["CANTIDAD"].notna() & (taller_solic_ref["CANTIDAD"] > 0)
        ].copy()
        taller_solic_ref["ARTICULO_NORM"] = taller_solic_ref["ARTICULO"].map(norm_product)
        taller_solic_ref["ESTADO_CAT"] = taller_solic_ref["ESTADO"].map(classify_estado)
        taller_solic_ref["ATENDIDO_FLAG"] = taller_solic_ref["ESTADO_CAT"].eq("ATENDIDO")
        taller_solic_ref["NO_DISP_FLAG"] = taller_solic_ref["ESTADO_CAT"].eq("NO_DISPONIBLE")
        taller_solic_ref["CANT_ATENDIDA"] = np.where(
            taller_solic_ref["ATENDIDO_FLAG"], taller_solic_ref["CANTIDAD"], 0.0
        )
        taller_solic_ref["CANT_NO_DISP"] = np.where(
            taller_solic_ref["NO_DISP_FLAG"], taller_solic_ref["CANTIDAD"], 0.0
        )
        taller_solic_ref["SUCURSAL"] = taller_solic_ref["SUCURSAL"].map(norm_text)

    return all_df, inv, taller_solic_ref


# ---------------------------------------------------------------------------
# Metricas
# ---------------------------------------------------------------------------
def period_stats(df: pd.DataFrame) -> dict:
    dmin, dmax = df["FECHA"].min(), df["FECHA"].max()
    dias = int((dmax - dmin).days) + 1
    semanas = dias / 7.0
    meses = dias / 30.0
    quincenas = dias / 15.0
    return {
        "fecha_min": dmin,
        "fecha_max": dmax,
        "dias": dias,
        "semanas": semanas,
        "quincenas": quincenas,
        "meses": meses,
    }


def agg_producto(df: pd.DataFrame, grupo_label: str, stats: dict) -> pd.DataFrame:
    g = (
        df.groupby(["ARTICULO_NORM", "CATEGORIA"], dropna=False)
        .agg(
            n_solicitudes=("CANTIDAD", "size"),
            und_solicitadas=("CANTIDAD", "sum"),
            und_atendidas=("CANT_ATENDIDA", "sum"),
            und_no_disponible=("CANT_NO_DISP", "sum"),
            n_atendidas=("ATENDIDO_FLAG", "sum"),
            n_no_disponible=("NO_DISP_FLAG", "sum"),
            n_sin_estado=("ESTADO_CAT", lambda s: (s == "SIN_ESTADO").sum()),
            cant_min_solicitud=("CANTIDAD", "min"),
            cant_mediana_solicitud=("CANTIDAD", "median"),
            cant_prom_solicitud=("CANTIDAD", "mean"),
            cant_max_solicitud=("CANTIDAD", "max"),
            n_sucursales=("SUCURSAL", "nunique"),
            primera_fecha=("FECHA", "min"),
            ultima_fecha=("FECHA", "max"),
            n_semanas_activas=("SEMANA", "nunique"),
            n_quincenas_activas=("QUINCENA", "nunique"),
        )
        .reset_index()
    )

    # Demanda semanal por producto (serie)
    weekly = (
        df.groupby(["ARTICULO_NORM", "SEMANA"])["CANTIDAD"]
        .sum()
        .reset_index()
        .groupby("ARTICULO_NORM")["CANTIDAD"]
        .agg(
            demanda_sem_prom="mean",
            demanda_sem_mediana="median",
            demanda_sem_max="max",
            demanda_sem_std="std",
        )
        .reset_index()
    )
    g = g.merge(weekly, on="ARTICULO_NORM", how="left")

    dias = stats["dias"]
    semanas = stats["semanas"]
    quincenas = stats["quincenas"]
    meses = stats["meses"]

    # Consumo promedio diario / semanal / mensual sobre el horizonte completo
    g["und_por_dia"] = g["und_solicitadas"] / dias
    g["und_por_semana"] = g["und_solicitadas"] / semanas
    g["und_por_quincena"] = g["und_solicitadas"] / quincenas
    g["und_por_mes"] = g["und_solicitadas"] / meses

    # % atendidas (por lineas y por unidades)
    g["pct_lineas_atendidas"] = np.where(
        g["n_solicitudes"] > 0, g["n_atendidas"] / g["n_solicitudes"], 0
    )
    g["pct_lineas_no_disponible"] = np.where(
        g["n_solicitudes"] > 0, g["n_no_disponible"] / g["n_solicitudes"], 0
    )
    g["pct_und_atendidas"] = np.where(
        g["und_solicitadas"] > 0, g["und_atendidas"] / g["und_solicitadas"], 0
    )
    g["pct_und_no_disponible"] = np.where(
        g["und_solicitadas"] > 0, g["und_no_disponible"] / g["und_solicitadas"], 0
    )

    # Demanda ajustada: si hubo stockouts, la demanda real pudo ser mayor.
    # Ajuste conservador: demanda_ajustada = solicitada + no_disponible * 0 (ya incluida)
    # Mejor: usar solicitada como demanda observada; senalar ruptura.
    g["demanda_observada_und"] = g["und_solicitadas"]
    g["senal_ruptura"] = np.where(g["pct_lineas_no_disponible"] >= 0.25, "ALTA",
                           np.where(g["pct_lineas_no_disponible"] >= 0.10, "MEDIA", "BAJA"))

    # Seguridad / cobertura
    # MIN = cobertura de 1 ciclo + buffer bajo (demanda mediana semanal o 0.5*prom)
    # MAX = cobertura de 2 ciclos + buffer alto (prom + 0.5*std o max semanal)
    dem_sem = g["und_por_semana"].fillna(0)
    dem_sem_max = g["demanda_sem_max"].fillna(dem_sem)
    dem_sem_std = g["demanda_sem_std"].fillna(0)

    # Escenarios de pedido (ceil a unidad)
    # Buffer normal +20%. Si ruptura ALTA: +35% y piso = cubrir pico reciente del ciclo
    # (sin extrapolar el pico a 4 semanas enteras, para no sobrepedir por un spike puntual)
    g["pedido_semanal_base"] = dem_sem
    g["pedido_semanal_sugerido"] = np.where(
        g["senal_ruptura"].eq("ALTA"),
        np.maximum(dem_sem * 1.35, dem_sem_max),
        dem_sem * 1.20,
    )
    g["pedido_quincenal_sugerido"] = np.where(
        g["senal_ruptura"].eq("ALTA"),
        np.maximum(g["und_por_quincena"] * 1.35, dem_sem_max * 1.5),
        g["und_por_quincena"] * 1.20,
    )
    g["pedido_mensual_sugerido"] = np.where(
        g["senal_ruptura"].eq("ALTA"),
        np.maximum(g["und_por_mes"] * 1.35, dem_sem_max * 2.0),
        g["und_por_mes"] * 1.20,
    )

    # Min / Max inventario operativo (en unidades, horizonte mensual de referencia)
    # MIN: ~1 semana de demanda (piso de seguridad)
    # MAX: techo = max(pedido mensual, ~2 semanas del pico) para no sobreestockar
    g["MIN_und"] = np.maximum(dem_sem, dem_sem_max * 0.5)
    g["MAX_und"] = np.maximum(g["pedido_mensual_sugerido"], dem_sem_max * 2)
    # Punto de reorden aprox (ROP) sin lead time especifico: 1 semana + 20%
    g["ROP_und"] = dem_sem * 1.20
    # Cantidad economica de pedido de referencia = pedido mensual sugerido
    g["Q_pedido_ref_mensual"] = g["pedido_mensual_sugerido"]

    for col in [
        "pedido_semanal_sugerido",
        "pedido_quincenal_sugerido",
        "pedido_mensual_sugerido",
        "MIN_und",
        "MAX_und",
        "ROP_und",
        "Q_pedido_ref_mensual",
    ]:
        g[col] = g[col].map(ceil_pos)

    g["GRUPO"] = grupo_label
    g = g.sort_values(["und_solicitadas", "n_solicitudes"], ascending=False)
    return g


def fill_rate_summary(df: pd.DataFrame, grupo: str) -> dict:
    n = len(df)
    n_att = int(df["ATENDIDO_FLAG"].sum())
    n_nd = int(df["NO_DISP_FLAG"].sum())
    n_se = int((df["ESTADO_CAT"] == "SIN_ESTADO").sum())
    und = float(df["CANTIDAD"].sum())
    und_att = float(df["CANT_ATENDIDA"].sum())
    und_nd = float(df["CANT_NO_DISP"].sum())
    return {
        "Grupo": grupo,
        "Lineas_solicitud": n,
        "Lineas_atendidas": n_att,
        "Lineas_no_disponible": n_nd,
        "Lineas_sin_estado": n_se,
        "Pct_lineas_atendidas": n_att / n if n else 0,
        "Pct_lineas_no_disponible": n_nd / n if n else 0,
        "Und_solicitadas": und,
        "Und_atendidas": und_att,
        "Und_no_disponible": und_nd,
        "Pct_und_atendidas": und_att / und if und else 0,
        "Pct_und_no_disponible": und_nd / und if und else 0,
        "SKUs": df["ARTICULO_NORM"].nunique(),
        "Sucursales_areas": df["SUCURSAL"].nunique(),
    }


def by_sucursal(df: pd.DataFrame) -> pd.DataFrame:
    g = (
        df.groupby(["GRUPO", "SUCURSAL"], dropna=False)
        .agg(
            lineas=("CANTIDAD", "size"),
            und=("CANTIDAD", "sum"),
            und_atendidas=("CANT_ATENDIDA", "sum"),
            und_no_disp=("CANT_NO_DISP", "sum"),
            lineas_atendidas=("ATENDIDO_FLAG", "sum"),
            lineas_no_disp=("NO_DISP_FLAG", "sum"),
            skus=("ARTICULO_NORM", "nunique"),
        )
        .reset_index()
    )
    g["pct_lineas_atendidas"] = g["lineas_atendidas"] / g["lineas"]
    g["pct_lineas_no_disponible"] = g["lineas_no_disp"] / g["lineas"]
    g["pct_und_atendidas"] = g["und_atendidas"] / g["und"]
    g["pct_und_no_disponible"] = g["und_no_disp"] / g["und"]
    return g.sort_values(["GRUPO", "und"], ascending=[True, False])


def top_no_disponible(df: pd.DataFrame, n: int = 30) -> pd.DataFrame:
    nd = df[df["NO_DISP_FLAG"]].copy()
    if nd.empty:
        return pd.DataFrame()
    g = (
        nd.groupby(["GRUPO", "ARTICULO_NORM", "CATEGORIA"])
        .agg(
            lineas_no_disp=("CANTIDAD", "size"),
            und_no_disp=("CANTIDAD", "sum"),
            sucursales=("SUCURSAL", lambda s: ", ".join(sorted(set(s)))),
        )
        .reset_index()
        .sort_values("und_no_disp", ascending=False)
        .head(n)
    )
    return g


def weekly_trend(df: pd.DataFrame) -> pd.DataFrame:
    g = (
        df.groupby(["GRUPO", "SEMANA"], dropna=False)
        .agg(
            lineas=("CANTIDAD", "size"),
            und=("CANTIDAD", "sum"),
            und_atendidas=("CANT_ATENDIDA", "sum"),
            und_no_disp=("CANT_NO_DISP", "sum"),
        )
        .reset_index()
    )
    g["pct_und_atendidas"] = np.where(g["und"] > 0, g["und_atendidas"] / g["und"], 0)
    return g.sort_values(["GRUPO", "SEMANA"])


def matrix_producto_sucursal(df: pd.DataFrame) -> pd.DataFrame:
    return (
        pd.pivot_table(
            df,
            index="ARTICULO_NORM",
            columns="SUCURSAL",
            values="CANTIDAD",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
        .rename(columns={"ARTICULO_NORM": "Producto"})
    )


# ---------------------------------------------------------------------------
# Excel formatting helpers
# ---------------------------------------------------------------------------
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_HEADER2 = PatternFill("solid", fgColor="2E75B6")
FILL_TITLE = PatternFill("solid", fgColor="0D2B45")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_WARN = PatternFill("solid", fgColor="FFEB9C")
FILL_BAD = PatternFill("solid", fgColor="FFC7CE")
FILL_SOFT = PatternFill("solid", fgColor="DDEBF7")
FILL_ALT = PatternFill("solid", fgColor="F2F2F2")
FONT_HEADER = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
FONT_TITLE = Font(color="FFFFFF", bold=True, name="Calibri", size=16)
FONT_SUB = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
FONT_BOLD = Font(bold=True, name="Calibri", size=11)
FONT_NORMAL = Font(name="Calibri", size=10)


def style_header_row(ws, row: int, start_col: int, end_col: int, fill=FILL_HEADER):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = FONT_HEADER
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = THIN


def autosize(ws, min_w=10, max_w=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value is None:
                continue
            length = max(length, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = max(min_w, length + 2)


def write_df(
    ws,
    df: pd.DataFrame,
    start_row: int = 1,
    start_col: int = 1,
    pct_cols: set | None = None,
    int_cols: set | None = None,
    float_cols: set | None = None,
):
    pct_cols = pct_cols or set()
    int_cols = int_cols or set()
    float_cols = float_cols or set()

    # header
    for j, col in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = THIN
    ws.row_dimensions[start_row].height = 32

    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, (col, val) in enumerate(zip(df.columns, row), start=start_col):
            cell = ws.cell(row=i, column=j, value=None if (isinstance(val, float) and np.isnan(val)) else val)
            cell.border = THIN
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(vertical="center")
            if i % 2 == 0:
                cell.fill = FILL_ALT
            if col in pct_cols and isinstance(val, (int, float)) and not pd.isna(val):
                cell.number_format = "0.0%"
            elif col in int_cols and isinstance(val, (int, float)) and not pd.isna(val):
                cell.number_format = "#,##0"
            elif col in float_cols and isinstance(val, (int, float)) and not pd.isna(val):
                cell.number_format = "#,##0.00"
    return start_row + len(df)


def add_notes(ws, row: int, notes: list[str], title="Notas / metodologia"):
    ws.cell(row=row, column=1, value=title).font = FONT_BOLD
    for i, n in enumerate(notes):
        ws.cell(row=row + 1 + i, column=1, value=f"• {n}").alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)


# ---------------------------------------------------------------------------
# Construccion del libro
# ---------------------------------------------------------------------------
def build_excel(
    all_df: pd.DataFrame,
    inv: pd.DataFrame,
    taller_solic_ref: pd.DataFrame | None = None,
):
    stats = period_stats(all_df)
    # Grupos
    df_tiendas = all_df[all_df["GRUPO"] == "TIENDAS"].copy()
    df_taller = all_df[all_df["GRUPO"] == "TALLER"].copy()  # movimientos completos
    df_creco_otras = all_df[all_df["GRUPO"] == "CRECO_OTRAS_AREAS"].copy()
    # Operaciones CRECO = Taller (movimientos) + otras areas (solicitudes)
    df_ops_creco = all_df[all_df["GRUPO"].isin(["TALLER", "CRECO_OTRAS_AREAS"])].copy()
    df_unificado = all_df.copy()
    taller_solic_ref = taller_solic_ref if taller_solic_ref is not None else pd.DataFrame()

    stats_t = period_stats(df_tiendas) if len(df_tiendas) else stats
    stats_taller = period_stats(df_taller) if len(df_taller) else stats
    stats_ops = period_stats(df_ops_creco) if len(df_ops_creco) else stats
    stats_u = stats

    prod_tiendas = agg_producto(df_tiendas, "TIENDAS", stats_t) if len(df_tiendas) else pd.DataFrame()
    prod_taller = agg_producto(df_taller, "TALLER", stats_taller) if len(df_taller) else pd.DataFrame()
    prod_creco_otras = (
        agg_producto(df_creco_otras, "CRECO_OTRAS_AREAS", stats_ops)
        if len(df_creco_otras)
        else pd.DataFrame()
    )
    prod_ops_creco = (
        agg_producto(df_ops_creco, "TALLER_Y_CRECO_OPS", stats_ops)
        if len(df_ops_creco)
        else pd.DataFrame()
    )

    # Unificado: sumar demandas por grupo (cada uno con su propio horizonte)
    # y recalcular tasas sobre 60 dias de referencia (= 2 meses), para no diluir
    # tiendas cuando el calendario de movimientos empieza antes.
    def build_unificado_from_groups(frames: list[pd.DataFrame]) -> pd.DataFrame:
        parts = [f for f in frames if f is not None and len(f)]
        if not parts:
            return pd.DataFrame()
        keys = ["ARTICULO_NORM", "CATEGORIA"]
        base = parts[0][keys].copy()
        for p in parts[1:]:
            base = base.merge(p[keys], on=keys, how="outer")
        base["CATEGORIA"] = base["CATEGORIA"].fillna("SIN CATEGORIA")
        sum_cols = [
            "n_solicitudes",
            "und_solicitadas",
            "und_atendidas",
            "und_no_disponible",
            "n_atendidas",
            "n_no_disponible",
            "n_sin_estado",
        ]
        out = base.copy()
        for c in sum_cols:
            out[c] = 0.0
            for p in parts:
                tmp = p[keys + [c]].copy()
                out = out.merge(tmp, on=keys, how="left", suffixes=("", "_x"))
                out[c] = out[c] + out[c + "_x"].fillna(0)
                out = out.drop(columns=[c + "_x"])
        # max pico semanal entre grupos; sucursales approx sum
        out["demanda_sem_max"] = 0.0
        out["n_sucursales"] = 0.0
        for p in parts:
            tmp = p[keys + ["demanda_sem_max", "n_sucursales"]].copy()
            out = out.merge(tmp, on=keys, how="left", suffixes=("", "_x"))
            out["demanda_sem_max"] = np.maximum(
                out["demanda_sem_max"], out["demanda_sem_max_x"].fillna(0)
            )
            out["n_sucursales"] = out["n_sucursales"] + out["n_sucursales_x"].fillna(0)
            out = out.drop(columns=["demanda_sem_max_x", "n_sucursales_x"])
        # horizonte de referencia 2 meses
        dias, semanas, quincenas, meses = 60.0, 60 / 7.0, 60 / 15.0, 2.0
        out["und_por_dia"] = out["und_solicitadas"] / dias
        out["und_por_semana"] = out["und_solicitadas"] / semanas
        out["und_por_quincena"] = out["und_solicitadas"] / quincenas
        out["und_por_mes"] = out["und_solicitadas"] / meses
        out["pct_lineas_atendidas"] = np.where(
            out["n_solicitudes"] > 0, out["n_atendidas"] / out["n_solicitudes"], 0
        )
        out["pct_lineas_no_disponible"] = np.where(
            out["n_solicitudes"] > 0, out["n_no_disponible"] / out["n_solicitudes"], 0
        )
        out["pct_und_atendidas"] = np.where(
            out["und_solicitadas"] > 0, out["und_atendidas"] / out["und_solicitadas"], 0
        )
        out["pct_und_no_disponible"] = np.where(
            out["und_solicitadas"] > 0, out["und_no_disponible"] / out["und_solicitadas"], 0
        )
        out["senal_ruptura"] = np.where(
            out["pct_lineas_no_disponible"] >= 0.25,
            "ALTA",
            np.where(out["pct_lineas_no_disponible"] >= 0.10, "MEDIA", "BAJA"),
        )
        dem_sem = out["und_por_semana"]
        dem_sem_max = out["demanda_sem_max"].fillna(dem_sem)
        out["pedido_semanal_sugerido"] = np.where(
            out["senal_ruptura"].eq("ALTA"),
            np.maximum(dem_sem * 1.35, dem_sem_max),
            dem_sem * 1.20,
        )
        out["pedido_quincenal_sugerido"] = np.where(
            out["senal_ruptura"].eq("ALTA"),
            np.maximum(out["und_por_quincena"] * 1.35, dem_sem_max * 1.5),
            out["und_por_quincena"] * 1.20,
        )
        out["pedido_mensual_sugerido"] = np.where(
            out["senal_ruptura"].eq("ALTA"),
            np.maximum(out["und_por_mes"] * 1.35, dem_sem_max * 2.0),
            out["und_por_mes"] * 1.20,
        )
        out["MIN_und"] = np.maximum(dem_sem, dem_sem_max * 0.5)
        out["MAX_und"] = np.maximum(out["pedido_mensual_sugerido"], dem_sem_max * 2)
        out["ROP_und"] = dem_sem * 1.20
        out["Q_pedido_ref_mensual"] = out["pedido_mensual_sugerido"]
        for col in [
            "pedido_semanal_sugerido",
            "pedido_quincenal_sugerido",
            "pedido_mensual_sugerido",
            "MIN_und",
            "MAX_und",
            "ROP_und",
            "Q_pedido_ref_mensual",
        ]:
            out[col] = out[col].map(ceil_pos)
        out["GRUPO"] = "UNIFICADO"
        out["demanda_sem_std"] = np.nan
        out["demanda_sem_prom"] = out["und_por_semana"]
        out["demanda_sem_mediana"] = out["und_por_semana"]
        return out.sort_values(["und_solicitadas", "n_solicitudes"], ascending=False)

    prod_unif = build_unificado_from_groups([prod_tiendas, prod_taller, prod_creco_otras])

    # Resumen fill rate
    fill_rows = [
        fill_rate_summary(df_tiendas, "TIENDAS (solicitudes)"),
        fill_rate_summary(
            df_taller,
            "TALLER (movimientos salidas — demanda completa)",
        ),
        fill_rate_summary(df_creco_otras, "CRECO otras areas (solicitudes I+D/Almacen/Mant./etc.)"),
        fill_rate_summary(df_ops_creco, "TALLER + CRECO OPS (combinado)"),
        fill_rate_summary(df_unificado, "UNIFICADO (Tiendas + Taller mov. + CRECO ops)"),
    ]
    if len(taller_solic_ref):
        # Referencia: solicitudes incompletas previas de Taller (NO usar para pedido)
        fill_rows.append(
            fill_rate_summary(
                taller_solic_ref,
                "REF — Taller solicitudes viejas (INCOMPLETO, solo fill rate historico)",
            )
        )
    fill_df = pd.DataFrame(fill_rows)

    # Pedido scenarios from unified + separate
    def scenario_sheet(prod: pd.DataFrame, freq: str) -> pd.DataFrame:
        if prod.empty:
            return prod
        col_map = {
            "semanal": "pedido_semanal_sugerido",
            "quincenal": "pedido_quincenal_sugerido",
            "mensual": "pedido_mensual_sugerido",
        }
        pedido_col = col_map[freq]
        out = prod[
            [
                "GRUPO",
                "ARTICULO_NORM",
                "CATEGORIA",
                "und_solicitadas",
                "und_atendidas",
                "und_no_disponible",
                "pct_lineas_atendidas",
                "pct_lineas_no_disponible",
                "pct_und_atendidas",
                "pct_und_no_disponible",
                "und_por_semana",
                "und_por_quincena",
                "und_por_mes",
                "demanda_sem_max",
                "senal_ruptura",
                "MIN_und",
                "ROP_und",
                "MAX_und",
                pedido_col,
                "n_solicitudes",
                "n_sucursales",
            ]
        ].copy()
        out = out.rename(
            columns={
                "ARTICULO_NORM": "Producto",
                "CATEGORIA": "Categoria",
                "und_solicitadas": "Und_solicitadas_2meses",
                "und_atendidas": "Und_atendidas",
                "und_no_disponible": "Und_no_disponible",
                "pct_lineas_atendidas": "%_lineas_atendidas",
                "pct_lineas_no_disponible": "%_lineas_no_disponible",
                "pct_und_atendidas": "%_und_atendidas",
                "pct_und_no_disponible": "%_und_no_disponible",
                "und_por_semana": "Prom_und_semana",
                "und_por_quincena": "Prom_und_quincena",
                "und_por_mes": "Prom_und_mes",
                "demanda_sem_max": "Pico_und_semana",
                "senal_ruptura": "Senal_ruptura_stock",
                "MIN_und": "MIN_und",
                "ROP_und": "Punto_reorden_und",
                "MAX_und": "MAX_und",
                pedido_col: f"Pedido_{freq}_sugerido_und",
                "n_solicitudes": "N_solicitudes",
                "n_sucursales": "N_sucursales_areas",
            }
        )
        # Justificacion textual corta
        out["Justificacion"] = out.apply(
            lambda r: (
                f"Base: {r['Und_solicitadas_2meses']:.0f} und en ~2 meses "
                f"(~{r['Prom_und_mes']:.1f}/mes). "
                f"Atencion und {r['%_und_atendidas']:.0%}; "
                f"No disp. {r['%_und_no_disponible']:.0%}. "
                f"Buffer {'+35% por ruptura alta' if r['Senal_ruptura_stock']=='ALTA' else '+20% operativo'}."
            ),
            axis=1,
        )
        return out

    # Comparativo escenarios unificado
    comp = prod_unif[
        [
            "ARTICULO_NORM",
            "CATEGORIA",
            "und_solicitadas",
            "pct_und_atendidas",
            "pct_und_no_disponible",
            "senal_ruptura",
            "MIN_und",
            "ROP_und",
            "MAX_und",
            "pedido_semanal_sugerido",
            "pedido_quincenal_sugerido",
            "pedido_mensual_sugerido",
            "und_por_semana",
            "und_por_mes",
        ]
    ].copy()
    comp = comp.rename(
        columns={
            "ARTICULO_NORM": "Producto",
            "CATEGORIA": "Categoria",
            "und_solicitadas": "Und_2meses",
            "pct_und_atendidas": "%_und_atendidas",
            "pct_und_no_disponible": "%_und_no_disponible",
            "senal_ruptura": "Senal_ruptura",
            "pedido_semanal_sugerido": "Opcion_A_Semanal_und",
            "pedido_quincenal_sugerido": "Opcion_B_Quincenal_und",
            "pedido_mensual_sugerido": "Opcion_C_Mensual_und",
            "und_por_semana": "Prom_semana",
            "und_por_mes": "Prom_mes",
        }
    )
    # Equivalente mensual de cada opcion (para comparar costo/volumen)
    comp["Equiv_mensual_si_semanal"] = comp["Opcion_A_Semanal_und"] * 4.3
    comp["Equiv_mensual_si_quincenal"] = comp["Opcion_B_Quincenal_und"] * 2
    comp["Equiv_mensual_si_mensual"] = comp["Opcion_C_Mensual_und"]
    for c in [
        "Equiv_mensual_si_semanal",
        "Equiv_mensual_si_quincenal",
        "Equiv_mensual_si_mensual",
    ]:
        comp[c] = comp[c].map(ceil_pos)

    # Recomendacion de frecuencia por producto
    def reco_freq(r):
        # Alta rotacion / alta ruptura -> semanal; media -> quincenal; baja -> mensual
        if r["Senal_ruptura"] == "ALTA" or r["Prom_semana"] >= 20:
            return "SEMANAL"
        if r["Prom_semana"] >= 5 or r["Senal_ruptura"] == "MEDIA":
            return "QUINCENAL"
        return "MENSUAL"

    comp["Frecuencia_recomendada"] = comp.apply(reco_freq, axis=1)
    comp["Pedido_und_segun_frecuencia"] = comp.apply(
        lambda r: r["Opcion_A_Semanal_und"]
        if r["Frecuencia_recomendada"] == "SEMANAL"
        else (
            r["Opcion_B_Quincenal_und"]
            if r["Frecuencia_recomendada"] == "QUINCENAL"
            else r["Opcion_C_Mensual_und"]
        ),
        axis=1,
    )

    # Detalle limpio
    det_cols = [
        "FECHA",
        "ARCHIVO",
        "FUENTE_TIPO",
        "GRUPO",
        "SUCURSAL",
        "NOMBRE",
        "ARTICULO_ORIG",
        "ARTICULO_NORM",
        "CARACTER_NORM",
        "UOM",
        "CATEGORIA",
        "CANTIDAD",
        "ESTADO_RAW",
        "ESTADO_CAT",
        "ATENDIDO_FLAG",
        "NO_DISP_FLAG",
        "NOTAS",
        "SEMANA",
        "QUINCENA",
        "MES",
    ]
    det_cols = [c for c in det_cols if c in all_df.columns]
    detalle = all_df[det_cols].sort_values(
        ["FECHA", "GRUPO", "SUCURSAL", "ARTICULO_NORM"]
    )

    # Catalogo
    catalog = inv[["ID", "PRODUNTOS", "CATEGORIA", "PRODUCTO_NORM"]].copy()
    used = set(all_df["ARTICULO_NORM"])
    catalog["Solicitado_en_periodo"] = catalog["PRODUCTO_NORM"].isin(used)
    catalog = catalog.rename(
        columns={
            "PRODUNTOS": "Producto_catalogo",
            "PRODUCTO_NORM": "Producto_norm",
            "CATEGORIA": "Categoria",
        }
    )

    wb = Workbook()

    # ========== 00_LEEME ==========
    ws = wb.active
    ws.title = "00_LEEME"
    ws["A1"] = "ANALISIS DE PEDIDOS DE CONSUMIBLES — TIENDAS + TALLER/CRECO"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 28

    bullets = [
        ("Periodo analizado", f"{stats['fecha_min'].date()} a {stats['fecha_max'].date()} ({stats['dias']} dias ≈ {stats['semanas']:.1f} semanas ≈ {stats['meses']:.2f} meses)"),
        ("Fuentes", "1) Solicitudes TIENDAS  |  2) MOVIMIENTOS TALLER actualizados (salidas)  |  3) Solicitudes CRECO otras areas"),
        ("Ajuste Taller", "La data previa de solicitudes de Taller estaba incompleta. Se reemplazo por movimientos de inventario (salidas = demanda real)."),
        ("Alcance", "Solo estas fuentes. Catalogo de inventario solo para categorizar (stock vacio en origen)."),
        ("Unidad de analisis", "UNIDADES de demanda (solicitudes en tiendas; |movimientos| en taller). Lead time variable; no se fija LT unico."),
        ("Atendido", "Tiendas/CRECO: RECIBIDO+SOLICITADO+ENVIADO+ENTREGADO. Taller movimientos: toda salida cuenta como ejecutada/atendida."),
        ("No atendido", "NO DISPONIBLE (aplica a solicitudes; en movimientos de taller no hay este estado)."),
        ("Objetivo", "Cuantificar demanda, fill rate, min/max y 3 opciones de pedido: Semanal / Quincenal / Mensual"),
        ("Grupos", "TIENDAS | TALLER (movimientos) | CRECO otras areas | UNIFICADO"),
    ]
    ws["A3"] = "RESUMEN EJECUTIVO DEL ARCHIVO"
    ws["A3"].font = FONT_SUB
    ws["A3"].fill = FILL_HEADER2
    ws.merge_cells("A3:G3")
    r = 5
    for k, v in bullets:
        ws.cell(row=r, column=1, value=k).font = FONT_BOLD
        ws.cell(row=r, column=2, value=v)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="HOJAS DEL LIBRO").font = FONT_SUB
    ws.cell(row=r, column=1).fill = FILL_HEADER2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    sheets_help = [
        ("00_LEEME", "Esta guia: alcance, definiciones y como usar el archivo"),
        ("01_KPIs_FillRate", "% atendidas vs no disponible (lineas y und), por grupo + ref. solicitudes taller incompletas"),
        ("02_Pedido_UNIFICADO", "Catalogo de pedido unificado con MIN/MAX/ROP y 3 opciones"),
        ("03_Pedido_TIENDAS", "Mismo analisis solo tiendas (solicitudes)"),
        ("04_Pedido_TALLER", "Pedido Taller basado en MOVIMIENTOS (salidas) — fuente completa"),
        ("04b_Pedido_CRECO_ops", "Taller movimientos + CRECO otras areas combinado"),
        ("04c_Pedido_CRECO_otras", "Solo CRECO otras areas (solicitudes)"),
        ("04d_Taller_viejo_vs_nuevo", "Contraste solicitudes incompletas vs movimientos actualizados"),
        ("05_Comparativo_3_Opciones", "Semanal vs Quincenal vs Mensual + frecuencia recomendada por SKU"),
        ("06_Por_Sucursal_Area", "Demanda y fill rate por sucursal/area"),
        ("07_Top_No_Disponible", "Productos con mas ruptura (prioridad compra; tipicamente tiendas/CRECO)"),
        ("08_Tendencia_Semanal", "Evolucion semanal de und de demanda"),
        ("09_Matriz_Prod_x_Sucursal", "Mapa producto x sucursal (unidades)"),
        ("10_Detalle_Limpio", "Base transaccional normalizada (auditoria)"),
        ("11_Catalogo_vs_Uso", "Catalogo de consumibles vs uso real"),
        ("12_Metodologia", "Formulas, supuestos de buffer, min/max y recomendaciones"),
    ]
    r += 2
    ws.cell(row=r, column=1, value="Hoja").font = FONT_HEADER
    ws.cell(row=r, column=1).fill = FILL_HEADER
    ws.cell(row=r, column=2, value="Contenido").font = FONT_HEADER
    ws.cell(row=r, column=2).fill = FILL_HEADER
    for name, desc in sheets_help:
        r += 1
        ws.cell(row=r, column=1, value=name).border = THIN
        ws.cell(row=r, column=2, value=desc).border = THIN

    r += 2
    ws.cell(row=r, column=1, value="COMO MONTAR EL PEDIDO (PASOS)").font = FONT_SUB
    ws.cell(row=r, column=1).fill = FILL_HEADER2
    steps = [
        "1) Revisar 01_KPIs_FillRate: entender nivel de servicio actual (% atendido vs no disponible).",
        "2) Ir a 05_Comparativo_3_Opciones: elegir politica global (semanal/quincenal/mensual) o usar Frecuencia_recomendada por producto.",
        "3) Usar 02_Pedido_UNIFICADO como lista maestra de compra (columnas Pedido_*_sugerido_und).",
        "4) Priorizar SKUs de 07_Top_No_Disponible (alta ruptura = comprar primero / subir MAX).",
        "5) Si opera separado: usar 03 para tiendas y 04 para taller (movimientos).",
        "6) Validar picos en 08_Tendencia_Semanal y concentracion por punto en 06/09.",
        "7) Ajustar MIN/MAX si el lead time de un producto es largo: subir ROP proporcional al LT en semanas.",
    ]
    for i, s in enumerate(steps):
        ws.cell(row=r + 1 + i, column=1, value=s)
        ws.merge_cells(start_row=r + 1 + i, start_column=1, end_row=r + 1 + i, end_column=7)

    # Totales clave en portada
    r = r + len(steps) + 3
    ws.cell(row=r, column=1, value="CIFRAS CLAVE DEL PERIODO").font = FONT_SUB
    ws.cell(row=r, column=1).fill = FILL_HEADER2
    key_stats = [
        ("Lineas totales", len(all_df)),
        ("Und demanda (tiendas+taller+creco)", int(all_df["CANTIDAD"].sum())),
        ("Und Taller (movimientos)", int(df_taller["CANTIDAD"].sum()) if len(df_taller) else 0),
        ("Und Tiendas (solicitudes)", int(df_tiendas["CANTIDAD"].sum()) if len(df_tiendas) else 0),
        ("Unidades atendidas", int(all_df["CANT_ATENDIDA"].sum())),
        ("Unidades no disponibles", int(all_df["CANT_NO_DISP"].sum())),
        ("% lineas atendidas", f"{all_df['ATENDIDO_FLAG'].mean():.1%}"),
        ("% lineas no disponible", f"{all_df['NO_DISP_FLAG'].mean():.1%}"),
        ("% und atendidas", f"{all_df['CANT_ATENDIDA'].sum()/all_df['CANTIDAD'].sum():.1%}"),
        ("% und no disponibles", f"{all_df['CANT_NO_DISP'].sum()/all_df['CANTIDAD'].sum():.1%}"),
        ("SKUs distintos", all_df["ARTICULO_NORM"].nunique()),
        ("Tiendas / areas", all_df["SUCURSAL"].nunique()),
        ("Lineas movimientos Taller", len(df_taller)),
    ]
    rr = r + 2
    for k, v in key_stats:
        ws.cell(row=rr, column=1, value=k).font = FONT_BOLD
        ws.cell(row=rr, column=2, value=v)
        rr += 1
    autosize(ws)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    # ========== 01 KPIs ==========
    ws = wb.create_sheet("01_KPIs_FillRate")
    ws["A1"] = "FILL RATE / % SOLICITUDES ATENDIDAS vs NO DISPONIBLE"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells("A1:N1")
    ws["A2"] = (
        "Definicion: ATENDIDO = RECIBIDO + SOLICITADO + ENVIADO + ENTREGADO. "
        "NO ATENDIDO = NO DISPONIBLE. Se reporta por lineas y por unidades."
    )
    ws.merge_cells("A2:N2")

    fill_view = fill_df.rename(
        columns={
            "Grupo": "Grupo",
            "Lineas_solicitud": "Lineas",
            "Lineas_atendidas": "Lineas_atendidas",
            "Lineas_no_disponible": "Lineas_no_disponible",
            "Lineas_sin_estado": "Lineas_sin_estado",
            "Pct_lineas_atendidas": "%_lineas_atendidas",
            "Pct_lineas_no_disponible": "%_lineas_no_disponible",
            "Und_solicitadas": "Und_solicitadas",
            "Und_atendidas": "Und_atendidas",
            "Und_no_disponible": "Und_no_disponible",
            "Pct_und_atendidas": "%_und_atendidas",
            "Pct_und_no_disponible": "%_und_no_disponible",
            "SKUs": "SKUs",
            "Sucursales_areas": "Sucursales_areas",
        }
    )
    write_df(
        ws,
        fill_view,
        start_row=4,
        pct_cols={"%_lineas_atendidas", "%_lineas_no_disponible", "%_und_atendidas", "%_und_no_disponible"},
        int_cols={
            "Lineas",
            "Lineas_atendidas",
            "Lineas_no_disponible",
            "Lineas_sin_estado",
            "Und_solicitadas",
            "Und_atendidas",
            "Und_no_disponible",
            "SKUs",
            "Sucursales_areas",
        },
    )
    # Estado raw breakdown
    estado_break = (
        all_df.groupby(["GRUPO", "ESTADO_RAW"])
        .agg(lineas=("CANTIDAD", "size"), und=("CANTIDAD", "sum"))
        .reset_index()
        .sort_values(["GRUPO", "lineas"], ascending=[True, False])
    )
    estado_break = estado_break.rename(
        columns={"ESTADO_RAW": "Estado_original", "lineas": "Lineas", "und": "Und"}
    )
    ws.cell(row=12, column=1, value="Desglose por estado original (auditoria)").font = FONT_BOLD
    write_df(ws, estado_break, start_row=13, int_cols={"Lineas", "Und"})

    add_notes(
        ws,
        13 + len(estado_break) + 2,
        [
            "Un % alto de NO DISPONIBLE implica demanda reprimida: el pedido debe cubrir al menos la demanda solicitada + buffer.",
            "SOLICITADO se cuenta como atendido/en ciclo (junto con RECIBIDO).",
            "Taller usa MOVIMIENTOS (salidas): no tiene NO DISPONIBLE; cada salida cuenta como demanda ejecutada/atendida.",
            "La fila REF de solicitudes Taller viejas es solo historica/incompleta: no usar para armar pedido.",
            "Lineas sin estado se excluyen del numerador atendido y del no disponible; revisar en Detalle.",
        ],
    )
    autosize(ws)

    # ========== helper to write pedido sheets ==========
    def write_pedido_sheet(name: str, title: str, prod: pd.DataFrame, freq_focus: str = "mensual"):
        ws = wb.create_sheet(name)
        ws["A1"] = title
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_TITLE
        ws.merge_cells("A1:V1")
        if prod.empty:
            ws["A3"] = "Sin datos para este grupo."
            return
        view = scenario_sheet(prod, freq_focus)
        # also add other pedido cols
        base = prod[
            [
                "ARTICULO_NORM",
                "pedido_semanal_sugerido",
                "pedido_quincenal_sugerido",
                "pedido_mensual_sugerido",
            ]
        ].rename(
            columns={
                "ARTICULO_NORM": "Producto",
                "pedido_semanal_sugerido": "Pedido_semanal_und",
                "pedido_quincenal_sugerido": "Pedido_quincenal_und",
                "pedido_mensual_sugerido": "Pedido_mensual_und",
            }
        )
        view = view.merge(base, on="Producto", how="left")
        # reorder
        cols = [
            "Producto",
            "Categoria",
            "N_solicitudes",
            "N_sucursales_areas",
            "Und_solicitadas_2meses",
            "Und_atendidas",
            "Und_no_disponible",
            "%_lineas_atendidas",
            "%_lineas_no_disponible",
            "%_und_atendidas",
            "%_und_no_disponible",
            "Prom_und_semana",
            "Prom_und_quincena",
            "Prom_und_mes",
            "Pico_und_semana",
            "Senal_ruptura_stock",
            "MIN_und",
            "Punto_reorden_und",
            "MAX_und",
            "Pedido_semanal_und",
            "Pedido_quincenal_und",
            "Pedido_mensual_und",
            "Justificacion",
        ]
        view = view[cols]
        write_df(
            ws,
            view,
            start_row=3,
            pct_cols={
                "%_lineas_atendidas",
                "%_lineas_no_disponible",
                "%_und_atendidas",
                "%_und_no_disponible",
            },
            int_cols={
                "N_solicitudes",
                "N_sucursales_areas",
                "Und_solicitadas_2meses",
                "Und_atendidas",
                "Und_no_disponible",
                "MIN_und",
                "Punto_reorden_und",
                "MAX_und",
                "Pedido_semanal_und",
                "Pedido_quincenal_und",
                "Pedido_mensual_und",
            },
            float_cols={
                "Prom_und_semana",
                "Prom_und_quincena",
                "Prom_und_mes",
                "Pico_und_semana",
            },
        )
        # color ruptura
        # find Senal col
        header = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
        if "Senal_ruptura_stock" in header:
            c = header["Senal_ruptura_stock"]
            for row in range(4, 4 + len(view)):
                val = ws.cell(row, c).value
                if val == "ALTA":
                    ws.cell(row, c).fill = FILL_BAD
                elif val == "MEDIA":
                    ws.cell(row, c).fill = FILL_WARN
                else:
                    ws.cell(row, c).fill = FILL_OK
        add_notes(
            ws,
            5 + len(view),
            [
                "MIN_und ≈ 1 semana de demanda (piso de seguridad).",
                "Punto_reorden_und ≈ demanda semanal × 1.20 (sin LT fijo; si LT=2 semanas, multiplicar por 2).",
                "MAX_und ≈ techo mensual sugerido (evita sobreestock).",
                "Pedidos: base = demanda del ciclo × 1.20; si Senal_ruptura=ALTA → ×1.35 con piso de pico (1× / 1.5× / 2× segun ciclo).",
                "Use Pedido_mensual_und para el pedido mensual recurrente de consumibles.",
            ],
        )
        autosize(ws)
        ws.column_dimensions["A"].width = 36
        ws.freeze_panes = "B4"

    write_pedido_sheet(
        "02_Pedido_UNIFICADO",
        "PEDIDO UNIFICADO (TIENDAS + TALLER movimientos + CRECO ops) — cantidades en UND",
        prod_unif,
    )
    write_pedido_sheet(
        "03_Pedido_TIENDAS",
        "PEDIDO SOLO TIENDAS (solicitudes) — cantidades en UND",
        prod_tiendas,
    )
    write_pedido_sheet(
        "04_Pedido_TALLER",
        "PEDIDO TALLER — basado en MOVIMIENTOS/SALIDAS (fuente completa) — UND",
        prod_taller,
    )
    if not prod_ops_creco.empty:
        write_pedido_sheet(
            "04b_Pedido_CRECO_ops",
            "PEDIDO TALLER (movimientos) + CRECO OTRAS AREAS — UND",
            prod_ops_creco,
        )
    if not prod_creco_otras.empty:
        write_pedido_sheet(
            "04c_Pedido_CRECO_otras",
            "PEDIDO CRECO OTRAS AREAS (I+D, Almacen, Mant., etc.) — UND",
            prod_creco_otras,
        )

    # ========== 04d contraste taller viejo vs nuevo ==========
    ws = wb.create_sheet("04d_Taller_viejo_vs_nuevo")
    ws["A1"] = "CONTRASTE: SOLICITUDES TALLER INCOMPLETAS vs MOVIMIENTOS ACTUALIZADOS"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells("A1:H1")
    ws["A2"] = (
        "Las solicitudes previas de Taller NO se usan para armar el pedido. "
        "Quedan solo como referencia de fill rate / cobertura incompleta. "
        "La demanda de pedido de Taller sale de movimientos (salidas)."
    )
    ws.merge_cells("A2:H2")

    if len(taller_solic_ref):
        old_agg = (
            taller_solic_ref.groupby("ARTICULO_NORM")
            .agg(
                und_solicitudes_viejas=("CANTIDAD", "sum"),
                lineas_viejas=("CANTIDAD", "size"),
                und_atendidas_viejas=("CANT_ATENDIDA", "sum"),
                und_no_disp_viejas=("CANT_NO_DISP", "sum"),
            )
            .reset_index()
        )
    else:
        old_agg = pd.DataFrame(
            columns=[
                "ARTICULO_NORM",
                "und_solicitudes_viejas",
                "lineas_viejas",
                "und_atendidas_viejas",
                "und_no_disp_viejas",
            ]
        )

    if len(df_taller):
        new_agg = (
            df_taller.groupby("ARTICULO_NORM")
            .agg(
                und_movimientos=("CANTIDAD", "sum"),
                lineas_movimientos=("CANTIDAD", "size"),
            )
            .reset_index()
        )
    else:
        new_agg = pd.DataFrame(
            columns=["ARTICULO_NORM", "und_movimientos", "lineas_movimientos"]
        )

    contrast = old_agg.merge(new_agg, on="ARTICULO_NORM", how="outer").fillna(0)
    contrast["delta_und_mov_menos_solic"] = (
        contrast["und_movimientos"] - contrast["und_solicitudes_viejas"]
    )
    contrast["en_movimientos"] = contrast["und_movimientos"] > 0
    contrast["en_solicitudes_viejas"] = contrast["und_solicitudes_viejas"] > 0
    contrast = contrast.sort_values("und_movimientos", ascending=False).rename(
        columns={"ARTICULO_NORM": "Producto"}
    )
    # resumen
    resumen_c = pd.DataFrame(
        [
            {
                "Metrica": "Lineas solicitudes Taller (viejo/incompleto)",
                "Valor": int(len(taller_solic_ref)),
            },
            {
                "Metrica": "Und solicitudes Taller (viejo/incompleto)",
                "Valor": int(taller_solic_ref["CANTIDAD"].sum()) if len(taller_solic_ref) else 0,
            },
            {
                "Metrica": "Lineas movimientos Taller (nuevo/completo)",
                "Valor": int(len(df_taller)),
            },
            {
                "Metrica": "Und movimientos Taller (nuevo/completo)",
                "Valor": int(df_taller["CANTIDAD"].sum()) if len(df_taller) else 0,
            },
            {
                "Metrica": "SKUs solo en movimientos (no estaban en solicitudes)",
                "Valor": int(
                    ((contrast["und_movimientos"] > 0) & (contrast["und_solicitudes_viejas"] == 0)).sum()
                ),
            },
            {
                "Metrica": "SKUs solo en solicitudes viejas (no salieron en movimientos)",
                "Valor": int(
                    ((contrast["und_solicitudes_viejas"] > 0) & (contrast["und_movimientos"] == 0)).sum()
                ),
            },
        ]
    )
    write_df(ws, resumen_c, start_row=4, int_cols={"Valor"})
    ws.cell(row=12, column=1, value="Detalle por producto").font = FONT_BOLD
    write_df(
        ws,
        contrast,
        start_row=13,
        int_cols={
            "und_solicitudes_viejas",
            "lineas_viejas",
            "und_atendidas_viejas",
            "und_no_disp_viejas",
            "und_movimientos",
            "lineas_movimientos",
            "delta_und_mov_menos_solic",
        },
    )
    autosize(ws)

    # ========== 05 Comparativo ==========
    ws = wb.create_sheet("05_Comparativo_3_Opciones")
    ws["A1"] = "TRES OPCIONES DE PEDIDO: SEMANAL / QUINCENAL / MENSUAL (UNIFICADO)"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells("A1:P1")
    ws["A2"] = (
        "Opcion A Semanal = mas reactivo, menos inventario parado, mas gestiones. "
        "Opcion B Quincenal = equilibrio. "
        "Opcion C Mensual = menos gestiones, mas stock de seguridad. "
        "Equiv_mensual_* permite comparar volumen mensual de cada politica."
    )
    ws.merge_cells("A2:P2")

    # Totales por opcion
    tot = pd.DataFrame(
        [
            {
                "Opcion": "A_SEMANAL",
                "Cadencia": "Cada 7 dias",
                "SKUs_con_pedido": int((comp["Opcion_A_Semanal_und"] > 0).sum()),
                "Und_por_ciclo": int(comp["Opcion_A_Semanal_und"].sum()),
                "Und_equiv_mes": int(comp["Equiv_mensual_si_semanal"].sum()),
                "Ventaja": "Menor stock / reaccion rapida a picos",
                "Riesgo": "Mas gestiones; sensible a lead time largo",
            },
            {
                "Opcion": "B_QUINCENAL",
                "Cadencia": "Cada 15 dias",
                "SKUs_con_pedido": int((comp["Opcion_B_Quincenal_und"] > 0).sum()),
                "Und_por_ciclo": int(comp["Opcion_B_Quincenal_und"].sum()),
                "Und_equiv_mes": int(comp["Equiv_mensual_si_quincenal"].sum()),
                "Ventaja": "Equilibrio operacion vs inventario",
                "Riesgo": "Buffer medio; vigilar SKUs de ruptura alta",
            },
            {
                "Opcion": "C_MENSUAL",
                "Cadencia": "Cada 30 dias",
                "SKUs_con_pedido": int((comp["Opcion_C_Mensual_und"] > 0).sum()),
                "Und_por_ciclo": int(comp["Opcion_C_Mensual_und"].sum()),
                "Und_equiv_mes": int(comp["Equiv_mensual_si_mensual"].sum()),
                "Ventaja": "Menos gestiones; ideal pedido mensual de consumibles",
                "Riesgo": "Mayor capital en stock; revisar MAX",
            },
        ]
    )
    write_df(ws, tot, start_row=4, int_cols={"SKUs_con_pedido", "Und_por_ciclo", "Und_equiv_mes"})

    # Mix recomendado
    mix = (
        comp.groupby("Frecuencia_recomendada")
        .agg(SKUs=("Producto", "size"), Und_pedido=("Pedido_und_segun_frecuencia", "sum"))
        .reset_index()
        .rename(columns={"Frecuencia_recomendada": "Frecuencia"})
    )
    ws.cell(row=9, column=1, value="Mix sugerido (frecuencia por producto, no una sola politica)").font = FONT_BOLD
    write_df(ws, mix, start_row=10, int_cols={"SKUs", "Und_pedido"})

    write_df(
        ws,
        comp,
        start_row=15,
        pct_cols={"%_und_atendidas", "%_und_no_disponible"},
        int_cols={
            "Und_2meses",
            "MIN_und",
            "ROP_und",
            "MAX_und",
            "Opcion_A_Semanal_und",
            "Opcion_B_Quincenal_und",
            "Opcion_C_Mensual_und",
            "Equiv_mensual_si_semanal",
            "Equiv_mensual_si_quincenal",
            "Equiv_mensual_si_mensual",
            "Pedido_und_segun_frecuencia",
        },
        float_cols={"Prom_semana", "Prom_mes"},
    )
    autosize(ws)
    ws.freeze_panes = "B16"

    # ========== 06 sucursal ==========
    ws = wb.create_sheet("06_Por_Sucursal_Area")
    ws["A1"] = "DEMANDA Y FILL RATE POR SUCURSAL / AREA"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells("A1:L1")
    suc = by_sucursal(all_df).rename(
        columns={
            "lineas": "Lineas",
            "und": "Und",
            "und_atendidas": "Und_atendidas",
            "und_no_disp": "Und_no_disponible",
            "lineas_atendidas": "Lineas_atendidas",
            "lineas_no_disp": "Lineas_no_disponible",
            "skus": "SKUs",
            "pct_lineas_atendidas": "%_lineas_atendidas",
            "pct_lineas_no_disponible": "%_lineas_no_disponible",
            "pct_und_atendidas": "%_und_atendidas",
            "pct_und_no_disponible": "%_und_no_disponible",
        }
    )
    write_df(
        ws,
        suc,
        start_row=3,
        pct_cols={
            "%_lineas_atendidas",
            "%_lineas_no_disponible",
            "%_und_atendidas",
            "%_und_no_disponible",
        },
        int_cols={
            "Lineas",
            "Und",
            "Und_atendidas",
            "Und_no_disponible",
            "Lineas_atendidas",
            "Lineas_no_disponible",
            "SKUs",
        },
    )
    autosize(ws)

    # ========== 07 top no disp ==========
    ws = wb.create_sheet("07_Top_No_Disponible")
    ws["A1"] = "PRIORIDAD DE COMPRA — TOP NO DISPONIBLE"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells("A1:G1")
    ws["A2"] = "Estos productos concentran la ruptura. Subir pedido/MIN/MAX y comprar primero."
    topnd = top_no_disponible(all_df, 40).rename(
        columns={
            "ARTICULO_NORM": "Producto",
            "CATEGORIA": "Categoria",
            "lineas_no_disp": "Lineas_no_disponible",
            "und_no_disp": "Und_no_disponible",
            "sucursales": "Sucursales_areas_afectadas",
        }
    )
    if topnd.empty:
        ws["A4"] = "No hubo registros NO DISPONIBLE."
    else:
        write_df(ws, topnd, start_row=4, int_cols={"Lineas_no_disponible", "Und_no_disponible"})
    autosize(ws)

    # ========== 08 tendencia ==========
    ws = wb.create_sheet("08_Tendencia_Semanal")
    ws["A1"] = "TENDENCIA SEMANAL DE SOLICITUDES"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells("A1:H1")
    trend = weekly_trend(all_df).rename(
        columns={
            "lineas": "Lineas",
            "und": "Und",
            "und_atendidas": "Und_atendidas",
            "und_no_disp": "Und_no_disponible",
            "pct_und_atendidas": "%_und_atendidas",
        }
    )
    write_df(
        ws,
        trend,
        start_row=3,
        pct_cols={"%_und_atendidas"},
        int_cols={"Lineas", "Und", "Und_atendidas", "Und_no_disponible"},
    )
    # chart for unificado total by week
    trend_u = (
        all_df.groupby("SEMANA")
        .agg(Und=("CANTIDAD", "sum"), Und_no_disp=("CANT_NO_DISP", "sum"))
        .reset_index()
        .sort_values("SEMANA")
    )
    start = 3 + len(trend) + 3
    ws.cell(row=start, column=1, value="Serie unificada (para grafico)").font = FONT_BOLD
    write_df(ws, trend_u, start_row=start + 1, int_cols={"Und", "Und_no_disp"})
    if len(trend_u) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Unidades solicitadas por semana (unificado)"
        chart.y_axis.title = "Und"
        data = Reference(
            ws,
            min_col=2,
            min_row=start + 1,
            max_col=3,
            max_row=start + 1 + len(trend_u),
        )
        cats = Reference(ws, min_col=1, min_row=start + 2, max_row=start + 1 + len(trend_u))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 18
        chart.height = 8
        ws.add_chart(chart, "F" + str(start + 1))
    autosize(ws)

    # ========== 09 matriz ==========
    ws = wb.create_sheet("09_Matriz_Prod_x_Sucursal")
    ws["A1"] = "MATRIZ PRODUCTO × SUCURSAL/AREA (UND SOLICITADAS)"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells("A1:P1")
    mat = matrix_producto_sucursal(all_df)
    # add total
    num_cols = [c for c in mat.columns if c != "Producto"]
    mat["TOTAL"] = mat[num_cols].sum(axis=1)
    mat = mat.sort_values("TOTAL", ascending=False)
    write_df(ws, mat, start_row=3, int_cols=set(num_cols + ["TOTAL"]))
    autosize(ws)
    ws.freeze_panes = "B4"

    # ========== 10 detalle ==========
    ws = wb.create_sheet("10_Detalle_Limpio")
    ws["A1"] = "BASE DETALLE NORMALIZADA (AUDITORIA)"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells("A1:R1")
    det_view = detalle.copy()
    det_view["FECHA"] = det_view["FECHA"].dt.strftime("%Y-%m-%d")
    det_view = det_view.rename(
        columns={
            "ARTICULO_ORIG": "Articulo_original",
            "ARTICULO_NORM": "Articulo_normalizado",
            "CARACTER_NORM": "Caracter_adicional",
            "ESTADO_RAW": "Estado_original",
            "ESTADO_CAT": "Estado_categoria",
            "ATENDIDO_FLAG": "Es_atendido",
            "NO_DISP_FLAG": "Es_no_disponible",
            "CANTIDAD": "Cantidad_und",
        }
    )
    write_df(
        ws,
        det_view,
        start_row=3,
        int_cols={"Cantidad_und"},
    )
    autosize(ws)
    ws.freeze_panes = "A4"

    # ========== 11 catalogo ==========
    ws = wb.create_sheet("11_Catalogo_vs_Uso")
    ws["A1"] = "CATALOGO DE CONSUMIBLES vs USO EN EL PERIODO"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells("A1:F1")
    # enrich with demand
    dem = (
        all_df.groupby("ARTICULO_NORM")["CANTIDAD"]
        .sum()
        .rename("Und_solicitadas_periodo")
        .reset_index()
        .rename(columns={"ARTICULO_NORM": "Producto_norm"})
    )
    cat_view = catalog.merge(dem, on="Producto_norm", how="left")
    cat_view["Und_solicitadas_periodo"] = cat_view["Und_solicitadas_periodo"].fillna(0).astype(int)
    cat_view = cat_view.sort_values(
        ["Solicitado_en_periodo", "Und_solicitadas_periodo"], ascending=[False, False]
    )
    write_df(ws, cat_view, start_row=3, int_cols={"Und_solicitadas_periodo"})
    autosize(ws)

    # ========== 12 metodologia ==========
    ws = wb.create_sheet("12_Metodologia")
    ws["A1"] = "METODOLOGIA, SUPUESTOS Y GUIA DE USO"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells("A1:B1")
    method = [
        ("Horizonte", f"{stats['dias']} dias corridos ({stats['fecha_min'].date()} → {stats['fecha_max'].date()})."),
        ("Demanda diaria", "Suma und solicitadas / dias del horizonte."),
        ("Demanda semanal", "Suma und / (dias/7). Tambien se calcula pico semanal observado."),
        ("Demanda quincenal", "Suma und / (dias/15)."),
        ("Demanda mensual", "Suma und / (dias/30). Es la base del pedido mensual de consumibles."),
        ("Fuente Taller", "MOVIMIENTOS actualizados (salidas de inventario). Reemplaza solicitudes incompletas de Taller."),
        ("Demanda Taller", "Cantidad = valor absoluto del movimiento negativo (salida). UOM puede ser UND/PACK/GAL/ROLLO."),
        ("Atendido (si)", "Solicitudes: RECIBIDO/SOLICITADO/ENVIADO/ENTREGADO. Movimientos Taller: toda salida = ejecutada."),
        ("No atendido (no)", "Estado NO DISPONIBLE en solicitudes. No aplica a movimientos de Taller."),
        ("% lineas atendidas", "Lineas con estado atendido / total lineas del grupo."),
        ("% und atendidas", "Suma und atendidas / suma und de demanda del grupo."),
        ("% no disponible", "Analogamente sobre lineas o und con estado NO DISPONIBLE (solicitudes)."),
        ("Buffer normal", "+20% sobre demanda del ciclo (variabilidad corta + redondeo operativo)."),
        ("Buffer ruptura alta", "Si ≥25% de lineas del SKU fueron NO DISPONIBLE: demanda×1.35, con piso = pico semanal (semanal), 1.5×pico (quincenal) o 2×pico (mensual). Evita extrapolar un spike a todo el mes."),
        ("MIN_und", "Piso ≈ demanda de 1 semana (seguridad minima en punto de uso/bodega)."),
        ("ROP_und", "Punto de reorden ≈ demanda semanal × 1.20. Si lead time = L semanas → ROP ≈ demanda_semanal × L × 1.20."),
        ("MAX_und", "Techo ≈ pedido mensual sugerido (o 2× pico semanal), para no sobreestockar."),
        ("Opcion Semanal", "Pedido_ciclo = demanda_semanal × buffer."),
        ("Opcion Quincenal", "Pedido_ciclo = demanda_quincenal × buffer."),
        ("Opcion Mensual", "Pedido_ciclo = demanda_mensual × buffer. Recomendada como default de consumibles."),
        ("Frecuencia recomendada", "ALTA rotacion o ruptura ALTA → semanal; media → quincenal; baja → mensual."),
        ("Lead time", "No hay LT por producto en la data. Las cantidades son en UND de demanda; ajustar ROP si LT > 7 dias."),
        ("Limitacion", "Solo 2 meses de historia: estacionalidad anual no visible. Revisar en 1-2 ciclos y recalibrar."),
        ("Separacion", "Tiendas (solicitudes), Taller (movimientos) y CRECO otras areas se analizan aparte y unificados."),
        ("Solicitudes Taller viejas", "Quedan en hoja 04d solo como referencia; NO alimentan el pedido."),
        ("Notas de envio parcial", "En tiendas, algunas notas indican envios parciales; la cantidad registrada es la solicitada."),
    ]
    ws.cell(row=3, column=1, value="Concepto").font = FONT_HEADER
    ws.cell(row=3, column=1).fill = FILL_HEADER
    ws.cell(row=3, column=2, value="Detalle").font = FONT_HEADER
    ws.cell(row=3, column=2).fill = FILL_HEADER
    for i, (k, v) in enumerate(method, start=4):
        ws.cell(row=i, column=1, value=k).font = FONT_BOLD
        ws.cell(row=i, column=1).border = THIN
        ws.cell(row=i, column=2, value=v).border = THIN
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110

    # Politica sugerida final
    r = 4 + len(method) + 2
    ws.cell(row=r, column=1, value="POLITICA SUGERIDA PARA EMPEZAR").font = FONT_SUB
    ws.cell(row=r, column=1).fill = FILL_HEADER2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    advice = [
        f"Pedido mensual unificado sugerido (suma SKUs): {int(comp['Opcion_C_Mensual_und'].sum())} und.",
        f"Equivalente si se pidiera semanal todo: {int(comp['Equiv_mensual_si_semanal'].sum())} und/mes.",
        f"Equivalente si se pidiera quincenal todo: {int(comp['Equiv_mensual_si_quincenal'].sum())} und/mes.",
        "Recomendacion practica: politica MENSUAL como default + excepcion SEMANAL/QUINCENAL para SKUs con Senal_ruptura=ALTA o alta rotacion (ver columna Frecuencia_recomendada).",
        "Separar compras si la logistica lo exige: hoja 03 (tiendas) y 04 (taller movimientos).",
    ]
    for i, a in enumerate(advice):
        ws.cell(row=r + 1 + i, column=1, value=a)
        ws.merge_cells(start_row=r + 1 + i, start_column=1, end_row=r + 1 + i, end_column=2)

    # Save also raw aggregates as CSV sidecar for traceability
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    detalle.to_csv(OUT_DIR / "detalle_limpio.csv", index=False)
    prod_unif.to_csv(OUT_DIR / "pedido_unificado.csv", index=False)
    fill_df.to_csv(OUT_DIR / "fill_rate_resumen.csv", index=False)

    wb.save(OUT_XLSX)
    wb.save(ARTIFACT_XLSX)
    return {
        "stats": stats,
        "fill_df": fill_df,
        "comp": comp,
        "n_tiendas": len(df_tiendas),
        "n_ops_creco": len(df_ops_creco),
        "n_taller": len(df_taller),
        "n_creco_otras": len(df_creco_otras),
        "n_total": len(all_df),
        "skus": all_df["ARTICULO_NORM"].nunique(),
        "und_taller": float(df_taller["CANTIDAD"].sum()) if len(df_taller) else 0,
    }


def main():
    all_df, inv, taller_solic_ref = load_all()
    print("Cargado:", len(all_df), "lineas")
    print(all_df["GRUPO"].value_counts().to_dict())
    print(all_df["ARCHIVO"].value_counts().to_dict())
    print(all_df["FUENTE_TIPO"].value_counts().to_dict())
    print("Estados:", all_df["ESTADO_CAT"].value_counts().to_dict())
    print("Fecha:", all_df["FECHA"].min(), "->", all_df["FECHA"].max())
    print(
        "Taller movimientos und:",
        int(all_df.loc[all_df["GRUPO"] == "TALLER", "CANTIDAD"].sum()),
        "| ref solicitudes incompletas lineas:",
        len(taller_solic_ref),
    )
    summary = build_excel(all_df, inv, taller_solic_ref)
    print("Excel:", OUT_XLSX)
    print("Artifact:", ARTIFACT_XLSX)
    print(summary["fill_df"].to_string(index=False))
    print(
        "Pedido mensual total und:",
        int(summary["comp"]["Opcion_C_Mensual_und"].sum()),
    )
    print(
        "Pedido semanal ciclo und:",
        int(summary["comp"]["Opcion_A_Semanal_und"].sum()),
    )
    print(
        "Pedido quincenal ciclo und:",
        int(summary["comp"]["Opcion_B_Quincenal_und"].sum()),
    )


if __name__ == "__main__":
    main()
