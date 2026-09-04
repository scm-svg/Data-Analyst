#!/usr/bin/env python3
"""Genera Excel de rango de acción (mín / solicitado / máx) para Short Playa Sublimado."""

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HTML_PATH = Path(__file__).resolve().parent / "SHORT PLAYA SUBL.html"
OUTPUT_PATH = Path(__file__).resolve().parent / "SHORT_PLAYA_SUBL_RANGO_PRODUCCION.xlsx"

MODELO = "SHORT PLAYA SUBLIMADO"
TORD = {
    "XS": 0, "S": 1, "M": 2, "L": 3, "XL": 4, "2XL": 5, "3XL": 6,
    "1": 10, "2": 11, "4": 12, "6": 13, "8": 14, "10": 15, "12": 16, "14": 17,
}
COLOR_ORDER = ["Playuela", "Sal", "Tucupido", "Sombrero", "Nuevo color"]
MAX_COVERAGE_MONTHS = 6
MAX_PCT_BUFFER = 1.10

CAB_CORE = {"M", "L", "XL"}
CAB_TAIL = {"S", "2XL"}
KIDS_CORE = {"6", "8", "10", "12", "14"}
KIDS_TAIL = {"1", "2", "4"}

title_fill = PatternFill("solid", fgColor="6ABF4A")
sub_fill = PatternFill("solid", fgColor="A8D5A2")
color_fill = PatternFill("solid", fgColor="E2F0E4")
tot_fill = PatternFill("solid", fgColor="C8E6C9")
sol_fill = PatternFill("solid", fgColor="FFF9C4")
min_fill = PatternFill("solid", fgColor="E3F2FD")
max_fill = PatternFill("solid", fgColor="FFE0B2")
white_fill = PatternFill("solid", fgColor="FFFFFF")
thin = Side(style="thin", color="2D5016")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def load_data() -> dict:
    html = HTML_PATH.read_text(encoding="utf-8")
    match = re.search(r"var DATA=(\{.*?\});", html, re.DOTALL)
    return json.loads(match.group(1))


def merge_rows(data: dict, genero: str) -> list:
    plan = [r for r in data["production_plan"] if r["modelo"] == MODELO and r["genero"] == genero]
    launch = [r for r in data.get("launch_production_plan", []) if r["modelo"] == MODELO and r["genero"] == genero]
    merged = {r["color"]: r for r in plan}
    for row in launch:
        merged[row["color"]] = row

    def sort_key(row):
        color = row["color"]
        if color in COLOR_ORDER:
            return (0, COLOR_ORDER.index(color))
        return (1, color)

    return sorted(merged.values(), key=sort_key)


def min_factor(genero: str, talla: str) -> float:
    if genero == "CAB":
        if talla in CAB_CORE:
            return 0.90
        if talla in CAB_TAIL:
            return 0.70
        return 0.85
    if talla in KIDS_CORE:
        return 0.85
    if talla in KIDS_TAIL:
        return 0.70
    return 0.85


def calc_range(genero: str, talla_row: dict) -> tuple[int, int, int]:
    sol = int(talla_row.get("produce", 0) or 0)
    if sol <= 0:
        return 0, 0, 0

    v_mes = float(talla_row.get("v_mes", 0) or 0)
    stk = int(talla_row.get("stk", 0) or 0)
    talla = talla_row["talla"]

    mn = max(0, round(sol * min_factor(genero, talla)))
    cap_coverage = max(0, round(v_mes * MAX_COVERAGE_MONTHS - stk)) if v_mes > 0 else sol
    cap_pct = max(0, round(sol * MAX_PCT_BUFFER))
    mx = min(cap_coverage, cap_pct)
    mx = max(mx, sol)
    return mn, sol, mx


def style_cell(cell, fill=None, bold=False, align=center):
    if fill:
        cell.fill = fill
    if bold:
        cell.font = Font(bold=True)
    cell.alignment = align
    cell.border = border


def write_genero_sheet(ws, genero: str, rows: list):
    talla_set = set()
    for row in rows:
        for t in row["tallas"]:
            talla_set.add(t["talla"])
    tallas = sorted(talla_set, key=lambda t: TORD.get(t, 99))
    ncol = 1 + len(tallas) * 3

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1, value="CANTIDADES POR COLORES — RANGO DE ACCIÓN")
    style_cell(c, title_fill, bold=True)

    ws.cell(row=2, column=1, value=f"SHORT PLAYA SUBLIMADO {genero}").font = Font(bold=True, italic=True)
    style_cell(ws.cell(row=2, column=1), sub_fill, align=left)

    col = 2
    for talla in tallas:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
        h = ws.cell(row=2, column=col, value=talla)
        style_cell(h, sub_fill, bold=True)
        for off, label, fill in [(0, "Mín", min_fill), (1, "Sol", sol_fill), (2, "Máx", max_fill)]:
            sub = ws.cell(row=3, column=col + off, value=label)
            style_cell(sub, fill, bold=True)
        col += 3

    totals = {"min": 0, "sol": 0, "max": 0}
    r_idx = 4
    for cp in rows:
        tmap = {t["talla"]: t for t in cp["tallas"]}
        label = cp["color"] + (" (lanzamiento)" if cp.get("is_launch") else "")
        style_cell(ws.cell(row=r_idx, column=1, value=label), color_fill, bold=True, align=left)

        row_min = row_sol = row_max = 0
        col = 2
        for talla in tallas:
            t = tmap.get(talla, {"talla": talla, "produce": 0, "v_mes": 0, "stk": 0})
            mn, sol, mx = calc_range(genero, t)
            for off, val, fill in [(0, mn, min_fill), (1, sol, sol_fill), (2, mx, max_fill)]:
                cell = ws.cell(row=r_idx, column=col + off, value=val if val else None)
                style_cell(cell, white_fill if val else fill)
                if val:
                    cell.font = Font(bold=True)
            row_min += mn
            row_sol += sol
            row_max += mx
            col += 3

        totals["min"] += row_min
        totals["sol"] += row_sol
        totals["max"] += row_max
        r_idx += 1

    style_cell(ws.cell(row=r_idx, column=1, value="TOTAL"), tot_fill, bold=True, align=left)
    style_cell(ws.cell(row=r_idx, column=2, value=totals["min"]), tot_fill, bold=True)
    style_cell(ws.cell(row=r_idx, column=3, value=totals["sol"]), tot_fill, bold=True)
    style_cell(ws.cell(row=r_idx, column=4, value=totals["max"]), tot_fill, bold=True)

    ws.column_dimensions["A"].width = 24
    for i in range(2, ncol + 1):
        ws.column_dimensions[get_column_letter(i)].width = 7

    return totals


def write_resumen(wb, data: dict, summary: dict):
    ws = wb.create_sheet("Resumen", 0)
    hs = data.get("high_season_factor", 1.4)
    rows = [
        ["SHORT PLAYA SUBLIMADO — RANGO DE PRODUCCIÓN"],
        [],
        ["Fuente", "Dashboard cursor/short-playa-launch-color-b710"],
        ["Factor temporada alta", hs],
        ["Peso diciembre en rotación base", data.get("dec_base_factor", 1.4)],
        ["Lead time / cobertura solicitada", f"{data.get('lead_months', 3)} meses"],
        ["Tiendas nuevas consideradas", ", ".join(data.get("new_stores", ["VELA", "BARQUISIMETO"]))],
        ["Ramp-up tiendas nuevas (lanzamiento)", f"{int((data.get('launch_new_store_uptake', 0.7) or 0.7) * 100)}%"],
        [],
        ["Género", "Mínimo", "Solicitado (dashboard)", "Máximo"],
        ["CAB", summary["CAB"]["min"], summary["CAB"]["sol"], summary["CAB"]["max"]],
        ["KIDS", summary["KIDS"]["min"], summary["KIDS"]["sol"], summary["KIDS"]["max"]],
        ["TOTAL", summary["TOTAL"]["min"], summary["TOTAL"]["sol"], summary["TOTAL"]["max"]],
        [],
        ["Notas"],
        ["• Solicitado = cantidades del dashboard (curva óptima, cobertura 3 meses, temporada alta)."],
        ["• Mínimo = piso operativo para no quedar cortos en tallas núcleo."],
        ["• Máximo = tope anti-sobrestock (menor entre +10% y 6 meses de cobertura por talla)."],
        ["• Incluye color de lanzamiento y proyección de red ampliada (VELA, Barquisimeto)."],
    ]
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True, size=13)
            if r == 10:
                cell.font = Font(bold=True)
                cell.fill = sub_fill
    ws.column_dimensions["A"].width = 42
    for col in "BCD":
        ws.column_dimensions[col].width = 18


def write_metodologia(wb, data: dict):
    ws = wb.create_sheet("Metodología")
    hs = data.get("high_season_factor", 1.4)
    text = [
        "METODOLOGÍA — RANGO DE ACCIÓN SHORT PLAYA SUBLIMADO",
        "",
        "1. SOLICITADO (referencia del dashboard)",
        "   Cantidad por color/talla según rotación ajustada × factor temporada alta (×{hs}).".format(hs=hs),
        "   Rotación base: promedio últimos 6 meses completos, con diciembre ponderado ×1.4.",
        "   Cobertura objetivo: 3 meses (lead time 90 días).",
        "   Color nuevo: benchmark top 3 colores (Sal, Playuela, Tucupido) × participación de red",
        "   incluyendo VELA (1.5× GRIETA) y Barquisimeto, con ramp-up 70% en apertura.",
        "",
        "2. MÍNIMO (piso — producción no debería bajar de aquí)",
        "   CAB tallas M/L/XL: 90% del solicitado.",
        "   CAB tallas S/2XL: 70% del solicitado.",
        "   KIDS tallas 6–14: 85% del solicitado.",
        "   KIDS tallas 1–4: 70% del solicitado.",
        "   Si solicitado = 0 → mínimo = 0.",
        "",
        "3. MÁXIMO (techo — anti sobrestock)",
        "   Menor valor entre:",
        "   a) Solicitado + 10%",
        "   b) Unidades para no superar 6 meses de cobertura (stock actual + producción).",
        "   El máximo nunca es menor que el solicitado.",
        "",
        "4. CONTEXTO TEMPORADA Y TIENDAS",
        "   Proyección orientada a picos: diciembre, enero y carnaval (factor temporada alta).",
        "   Uso inteligente de tela disponible: fabricar según rango y redistribuir entre tiendas.",
        "   Tiendas nuevas absorben parte del volumen de lanzamiento sin sobrecargar stock en una sola tienda.",
        "",
        "5. CÓMO USA PRODUCCIÓN EL RANGO",
        "   • Fabricar preferentemente en el SOLICITADO.",
        "   • Si hay restricción de tela/capacidad: no bajar del MÍNIMO en tallas núcleo.",
        "   • Si hay tela adicional y demanda confirma pico festivo: puede subir hasta el MÁXIMO sin aprobación extra.",
    ]
    for r, line in enumerate(text, start=1):
        cell = ws.cell(row=r, column=1, value=line)
        if line.startswith(("METODOLOGÍA", "1.", "2.", "3.", "4.", "5.")):
            cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 95


def main():
    data = load_data()
    wb = Workbook()
    wb.remove(wb.active)

    summary = {}
    for genero in ["CAB", "KIDS"]:
        rows = merge_rows(data, genero)
        ws = wb.create_sheet(genero)
        totals = write_genero_sheet(ws, genero, rows)
        summary[genero] = totals

    summary["TOTAL"] = {
        k: summary["CAB"][k] + summary["KIDS"][k] for k in ("min", "sol", "max")
    }
    write_resumen(wb, data, summary)
    write_metodologia(wb, data)
    wb.save(OUTPUT_PATH)
    print(f"Guardado: {OUTPUT_PATH}")
    print(f"CAB:  {summary['CAB']['min']} — {summary['CAB']['sol']} — {summary['CAB']['max']}")
    print(f"KIDS: {summary['KIDS']['min']} — {summary['KIDS']['sol']} — {summary['KIDS']['max']}")
    print(f"TOTAL:{summary['TOTAL']['min']} — {summary['TOTAL']['sol']} — {summary['TOTAL']['max']}")


if __name__ == "__main__":
    main()
