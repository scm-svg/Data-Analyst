#!/usr/bin/env python3
"""Exporta SPOTS_PRODUCCION_EXPANSION.xlsx — matriz color/diseño × talla por zona (estilo Short Playa)."""

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from build_spots_dashboard import (
    DEFAULT_PROD_MONTHS,
    FABRIC_KG_PER_UNIT,
    FABRIC_SAFETY_PCT,
    ZONE_ADICIONAL_COLOR,
    rebuild_data,
)

OUT_PATH = Path(__file__).resolve().parent / "SPOTS_PRODUCCION_EXPANSION.xlsx"
TALLA_ORDER = ["XS", "S", "M", "L", "XL", "2XL"]
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F2937")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=12)
TOTAL_FILL = PatternFill("solid", fgColor="E5E7EB")
TOTAL_FONT = Font(bold=True)
SECTION_FILL = PatternFill("solid", fgColor="F3F4F6")


def _style_range(ws, row, col_start, col_end, fill=None, font=None, border=True):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if border:
            cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_matrix(ws, start_row, title, tallas, rows_data):
    """rows_data: list of (label, {talla: qty})"""
    ws.cell(row=start_row, column=1, value="CANTIDADES POR DISEÑO / COLOR").font = TITLE_FONT
    hdr_row = start_row + 1
    ws.cell(row=hdr_row, column=1, value=title).font = Font(bold=True)
    for i, t in enumerate(tallas, start=2):
        ws.cell(row=hdr_row, column=i, value=t)
    ws.cell(row=hdr_row, column=len(tallas) + 2, value="Tot")
    _style_range(ws, hdr_row, 1, len(tallas) + 2, fill=HDR_FILL, font=HDR_FONT)

    r = hdr_row + 1
    col_totals = defaultdict(int)
    grand = 0
    for label, by_talla in rows_data:
        ws.cell(row=r, column=1, value=label)
        row_sum = 0
        for i, t in enumerate(tallas, start=2):
            qty = int(by_talla.get(t, 0) or 0)
            if qty:
                ws.cell(row=r, column=i, value=qty)
            col_totals[t] += qty
            row_sum += qty
        ws.cell(row=r, column=len(tallas) + 2, value=row_sum)
        grand += row_sum
        _style_range(ws, r, 1, len(tallas) + 2)
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    for i, t in enumerate(tallas, start=2):
        if col_totals[t]:
            ws.cell(row=r, column=i, value=col_totals[t])
    ws.cell(row=r, column=len(tallas) + 2, value=grand)
    _style_range(ws, r, 1, len(tallas) + 2, fill=TOTAL_FILL, font=TOTAL_FONT)
    return r + 2, grand


def _zone_genero_rows(store, genero, months):
    """Agrupa SKUs por diseño+color para un género."""
    key = f"need_{months}m"
    grouped = defaultdict(lambda: defaultdict(int))
    for sku in store.get("skus", []):
        if sku["genero"] != genero:
            continue
        label = f"{sku['diseno']} ({sku['color']})"
        grouped[label][sku["talla"]] += sku.get(key, sku.get("need_3m", 0))
    tallas = [t for t in TALLA_ORDER if any(grouped[l].get(t) for l in grouped)]
    rows = [(label, dict(tallas)) for label, tallas in sorted(grouped.items())]
    return tallas, rows


def _all_skus_by_genero(data, genero, months):
    """Agrega SKUs de todas las zonas para un género."""
    key = f"need_{months}m"
    by_label = defaultdict(lambda: defaultdict(int))
    meta = {}
    for store in data["expansion"]["by_store"]:
        zone = store["store"]
        for sku in store.get("skus", []):
            if sku["genero"] != genero:
                continue
            if sku["color"] == "Blanco":
                label = f"{sku['diseno']} · Blanco ({zone.title()})"
            else:
                label = f"{sku['color']} · {zone.title()}"
            by_label[label][sku["talla"]] += sku.get(key, sku.get("need_3m", 0))
            meta[label] = {"color": sku["color"], "zona": zone, "diseno": sku["diseno"]}
    return by_label, meta


def _consolidated_rows(by_label, meta):
    """Blanco primero, luego colores — orden Short Playa."""
    blanco = []
    colores = []
    for label, tallas in by_label.items():
        m = meta[label]
        row = (label, dict(tallas), m)
        if m["color"] == "Blanco":
            blanco.append(row)
        else:
            colores.append(row)
    blanco.sort(key=lambda x: (x[2]["zona"], x[2]["diseno"]))
    color_order = {c: i for i, c in enumerate(ZONE_ADICIONAL_COLOR.values())}
    colores.sort(key=lambda x: (color_order.get(x[2]["color"], 99), x[2]["zona"]))
    return blanco, colores


def _write_colores_sheet(wb, data, genero, months):
    """Pestaña estilo Short Playa: blanco arriba, colores abajo, tallas en columnas."""
    ws = wb.create_sheet(genero)
    ws.column_dimensions["A"].width = 36
    for i in range(2, 12):
        ws.column_dimensions[get_column_letter(i)].width = 8

    by_label, meta = _all_skus_by_genero(data, genero, months)
    blanco_rows, color_rows = _consolidated_rows(by_label, meta)
    all_tallas = [t for t in TALLA_ORDER if any(by_label[l].get(t) for l in by_label)]

    row = 1
    ws.cell(row=row, column=1, value="CANTIDADES POR COLORES").font = Font(bold=True, size=13)
    row += 1
    hdr = row
    ws.cell(row=hdr, column=1, value=f"SPOTS MANGA CORTA {genero}")
    for i, t in enumerate(all_tallas, start=2):
        ws.cell(row=hdr, column=i, value=t)
    ws.cell(row=hdr, column=len(all_tallas) + 2, value="Tot")
    _style_range(ws, hdr, 1, len(all_tallas) + 2, fill=HDR_FILL, font=HDR_FONT)
    row = hdr + 1

    def write_rows(rows_data, section_total_label=None):
        nonlocal row
        col_totals = defaultdict(int)
        section_total = 0
        for label, tallas, _m in rows_data:
            ws.cell(row=row, column=1, value=label)
            row_sum = 0
            for i, t in enumerate(all_tallas, start=2):
                qty = int(tallas.get(t, 0) or 0)
                if qty:
                    ws.cell(row=row, column=i, value=qty)
                col_totals[t] += qty
                row_sum += qty
            ws.cell(row=row, column=len(all_tallas) + 2, value=row_sum)
            section_total += row_sum
            _style_range(ws, row, 1, len(all_tallas) + 2)
            row += 1
        if section_total_label and rows_data:
            ws.cell(row=row, column=1, value=section_total_label).font = TOTAL_FONT
            for i, t in enumerate(all_tallas, start=2):
                if col_totals[t]:
                    ws.cell(row=row, column=i, value=col_totals[t])
            ws.cell(row=row, column=len(all_tallas) + 2, value=section_total)
            _style_range(ws, row, 1, len(all_tallas) + 2, fill=TOTAL_FILL, font=TOTAL_FONT)
            row += 1
        return section_total, col_totals

    grand_cols = defaultdict(int)
    grand = 0

    b_tot, b_cols = write_rows(blanco_rows, "TOTAL BLANCO")
    grand += b_tot
    for t, v in b_cols.items():
        grand_cols[t] += v
    row += 1

    c_tot, c_cols = write_rows(color_rows, "TOTAL COLORES")
    grand += c_tot
    for t, v in c_cols.items():
        grand_cols[t] += v

    ws.cell(row=row, column=1, value="TOTAL GENERAL").font = Font(bold=True, size=11)
    for i, t in enumerate(all_tallas, start=2):
        if grand_cols[t]:
            ws.cell(row=row, column=i, value=grand_cols[t])
    ws.cell(row=row, column=len(all_tallas) + 2, value=grand)
    _style_range(ws, row, 1, len(all_tallas) + 2, fill=PatternFill("solid", fgColor="D1FAE5"), font=TOTAL_FONT)


def _write_zone_sheet(wb, store, months):
    zone = store["store"]
    ws = wb.create_sheet(zone.title())
    ws.column_dimensions["A"].width = 34
    for i in range(2, 10):
        ws.column_dimensions[get_column_letter(i)].width = 8

    row = 1
    ws.cell(row=row, column=1, value=f"SPOTS MANGA CORTA — {zone} · Proyección {months} meses")
    ws.cell(row=row, column=1).font = Font(bold=True, size=13)
    row += 2
    ws.cell(row=row, column=1, value=store.get("label", ""))
    row += 2

    zone_total = 0
    for genero in ("CAB", "DAMA"):
        tallas, rows = _zone_genero_rows(store, genero, months)
        if not rows:
            continue
        ws.cell(row=row, column=1, value=genero).font = Font(bold=True, size=11)
        ws.cell(row=row, column=1).fill = SECTION_FILL
        row += 1
        row, sub = _write_matrix(
            ws, row, f"SPOTS MANGA CORTA {genero} — {zone}", tallas, rows,
        )
        zone_total += sub

    ws.cell(row=row, column=1, value=f"TOTAL ZONA {zone}").font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=zone_total).font = Font(bold=True, size=11)


def _write_resumen(wb, data, months):
    ws = wb.create_sheet("RESUMEN", 0)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 40

    exp = data["expansion"]
    fab = data.get("fabric", exp.get("fabric", {}))
    row = 1
    ws.cell(row=row, column=1, value="SPOTS — Resumen producción expansión").font = Font(bold=True, size=14)
    row += 2
    meta = [
        ("Horizonte", f"{months} meses"),
        ("Base rotación", data.get("velocity_months_label", "")),
        ("Temporada alta", f"×{data.get('high_season_factor', 1.2)}"),
        ("Virgen BQT (dic)", f"×{data.get('december_hs_factor', 1.4)}"),
        ("Colores zona", "Caracas Azul Marino · Valencia Vinotinto · Barquisimeto Verde"),
        ("Factor color", f"{int((data.get('additional_color_factor') or 0.7) * 100)}% del blanco principal"),
    ]
    for label, val in meta:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=val)
        row += 1
    row += 1

    headers = ["Zona", "Blanco", "Color zona", "Total"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_range(ws, row, 1, 4, fill=HDR_FILL, font=HDR_FONT)
    row += 1
    for store in exp.get("by_store", []):
        ws.cell(row=row, column=1, value=store["store"])
        ws.cell(row=row, column=2, value=store["blanco"])
        ws.cell(row=row, column=3, value=f"{store.get('adicional_color', '')} ({store['adicional']})")
        ws.cell(row=row, column=4, value=store["total"])
        _style_range(ws, row, 1, 4)
        row += 1
    ws.cell(row=row, column=1, value="TOTAL EXPANSIÓN").font = TOTAL_FONT
    ws.cell(row=row, column=2, value=exp.get("total_blanco", 0))
    ws.cell(row=row, column=3, value=exp.get("total_adicional", 0))
    ws.cell(row=row, column=4, value=exp.get("total_expansion", 0))
    _style_range(ws, row, 1, 4, fill=TOTAL_FILL, font=TOTAL_FONT)
    row += 2

    ws.cell(row=row, column=1, value="Nota metodológica").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value=exp.get("nota", ""))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    row += 2

    ws.cell(row=row, column=1, value="Compra de tela (referencia)").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value=f"Consumo {FABRIC_KG_PER_UNIT} kg/und · stock seguridad +{int(FABRIC_SAFETY_PCT * 100)}%")
    row += 1
    ws.cell(row=row, column=1, value="Material")
    ws.cell(row=row, column=2, value="Unidades")
    ws.cell(row=row, column=3, value="Tela (kg)")
    _style_range(ws, row, 1, 3, fill=HDR_FILL, font=HDR_FONT)
    row += 1
    blanco = fab.get("blanco", {})
    ws.cell(row=row, column=1, value="Blanco (total)")
    ws.cell(row=row, column=2, value=blanco.get("units", 0))
    ws.cell(row=row, column=3, value=blanco.get("kg", 0))
    _style_range(ws, row, 1, 3)
    row += 1
    for r in fab.get("adicional_by_zone", []):
        ws.cell(row=row, column=1, value=r.get("color", r.get("tipo", "Color")))
        ws.cell(row=row, column=2, value=r.get("units", 0))
        ws.cell(row=row, column=3, value=r.get("kg", 0))
        _style_range(ws, row, 1, 3)
        row += 1
    total = fab.get("total", {})
    ws.cell(row=row, column=1, value="TOTAL TELA").font = TOTAL_FONT
    ws.cell(row=row, column=2, value=total.get("units", 0))
    ws.cell(row=row, column=3, value=total.get("kg", 0))
    _style_range(ws, row, 1, 3, fill=TOTAL_FILL, font=TOTAL_FONT)


def _write_tela(wb, data):
    fab = data.get("fabric", data["expansion"].get("fabric", {}))
    ws = wb.create_sheet("TELA")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12

    row = 1
    ws.cell(row=row, column=1, value="Pedido de tela — Expansión SPOTS").font = Font(bold=True, size=13)
    row += 2
    headers = ["Material", "Zona / diseño", "Unidades", "Consumo", "Stock seg.", "Tela (kg)"]
    for i, h in enumerate(headers,  start=1):
        ws.cell(row=row, column=i, value=h)
    _style_range(ws, row, 1, 6, fill=HDR_FILL, font=HDR_FONT)
    row += 1

    blanco = fab.get("blanco", {})
    ws.cell(row=row, column=1, value="Blanco")
    ws.cell(row=row, column=2, value="Total expansión")
    ws.cell(row=row, column=3, value=blanco.get("units", 0))
    ws.cell(row=row, column=4, value=f"{FABRIC_KG_PER_UNIT} kg/und")
    ws.cell(row=row, column=5, value=f"+{int(FABRIC_SAFETY_PCT * 100)}%")
    ws.cell(row=row, column=6, value=blanco.get("kg", 0))
    _style_range(ws, row, 1, 6)
    row += 1

    for d in fab.get("blanco_detail", []):
        ws.cell(row=row, column=1, value="↳ Blanco")
        ws.cell(row=row, column=2, value=f"Barquisimeto · {d['detalle']}")
        ws.cell(row=row, column=3, value=d.get("units", 0))
        ws.cell(row=row, column=4, value=f"{FABRIC_KG_PER_UNIT} kg/und")
        ws.cell(row=row, column=5, value=f"+{int(FABRIC_SAFETY_PCT * 100)}%")
        ws.cell(row=row, column=6, value=d.get("kg", 0))
        _style_range(ws, row, 1, 6)
        row += 1

    for r in fab.get("adicional_by_zone", []):
        color = r.get("color", r.get("tipo", "Color"))
        ws.cell(row=row, column=1, value=color)
        ws.cell(row=row, column=2, value=f"{r['zona']} · {r.get('detalle', '')}")
        ws.cell(row=row, column=3, value=r.get("units", 0))
        ws.cell(row=row, column=4, value=f"{FABRIC_KG_PER_UNIT} kg/und")
        ws.cell(row=row, column=5, value=f"+{int(FABRIC_SAFETY_PCT * 100)}%")
        ws.cell(row=row, column=6, value=r.get("kg", 0))
        _style_range(ws, row, 1, 6)
        row += 1

    total = fab.get("total", {})
    ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=row, column=2, value="Caracas + Valencia + Barquisimeto")
    ws.cell(row=row, column=3, value=total.get("units", 0))
    ws.cell(row=row, column=4, value=f"{FABRIC_KG_PER_UNIT} kg/und")
    ws.cell(row=row, column=5, value=f"+{int(FABRIC_SAFETY_PCT * 100)}%")
    ws.cell(row=row, column=6, value=total.get("kg", 0))
    _style_range(ws, row, 1, 6, fill=TOTAL_FILL, font=TOTAL_FONT)


def export_produccion_xlsx(months: int = DEFAULT_PROD_MONTHS, out_path: Path = OUT_PATH) -> Path:
    data = rebuild_data()
    wb = Workbook()
    wb.remove(wb.active)

    _write_resumen(wb, data, months)
    _write_colores_sheet(wb, data, "CAB", months)
    _write_colores_sheet(wb, data, "DAMA", months)
    for store in data["expansion"]["by_store"]:
        _write_zone_sheet(wb, store, months)
    _write_tela(wb, data)

    wb.save(out_path)
    return out_path


def main():
    path = export_produccion_xlsx()
    data = rebuild_data()
    exp = data["expansion"]
    print(f"Wrote {path}")
    print(f"Expansión {exp['total_expansion']} und · Tela {data['fabric']['total']['kg']} kg")


if __name__ == "__main__":
    main()
